from flask import Blueprint, render_template, request, session, redirect, url_for, flash, jsonify
from app.models import db, CaseBenefit, Case, FapReason
from datetime import datetime, date
from functools import wraps
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import unicodedata

def _normalize_header(value):
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    for token in [' ', '-', '.', '/', '\\', ':', '(', ')', '$']:
        text = text.replace(token, '_')
    while '__' in text:
        text = text.replace('__', '_')
    return text.strip('_')

def _parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None

def _read_excel_rows(file_storage):
    workbook = load_workbook(filename=file_storage, data_only=True)
    data_rows = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = rows[0]
        if not headers or all(header is None or str(header).strip() == '' for header in headers):
            continue
        for row_index, row in enumerate(rows[1:], start=2):
            if row is None:
                continue
            row_dict = {'__sheet__': sheet.title, '__row__': row_index}
            for idx, header in enumerate(headers):
                if header is None:
                    continue
                row_dict[str(header)] = row[idx] if idx < len(row) else None
            data_rows.append(row_dict)
    return data_rows

benefits_bp = Blueprint('benefits', __name__, url_prefix='/benefits')

def require_law_firm(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('law_firm_id'):
            if request.is_json:
                return jsonify({"error": "Unauthorized"}), 401
            else:
                return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function

@benefits_bp.route('/')
def benefits_list():
    """Lista todos os benefícios do sistema para visualização geral"""
    benefits = CaseBenefit.query.join(Case).order_by(CaseBenefit.created_at.desc()).all()
    return render_template('benefits/list.html', benefits=benefits)

@benefits_bp.route('/<int:benefit_id>')
def benefit_detail(benefit_id):
    """Visualiza detalhes de um benefício específico"""
    benefit = CaseBenefit.query.get_or_404(benefit_id)
    
    related_cases = Case.query.filter(
        Case.benefits.any(CaseBenefit.id == benefit_id)
    ).order_by(Case.created_at.desc()).all()
    
    from app.models import Document
    related_documents = Document.query.filter_by(related_benefit_id=benefit_id).all()
    
    return render_template('benefits/detail.html', 
                         benefit=benefit, 
                         related_cases=related_cases,
                         documents=related_documents)

@benefits_bp.route('/case/<int:case_id>')
def case_benefits_list(case_id):
    case = Case.query.get_or_404(case_id)
    benefits = CaseBenefit.query.filter_by(case_id=case_id).order_by(CaseBenefit.created_at.desc()).all()
    return render_template('cases/benefits_list.html', case=case, benefits=benefits)

@benefits_bp.route('/case/<int:case_id>/import', methods=['GET', 'POST'])
def case_benefits_import(case_id):
    case = Case.query.get_or_404(case_id)
    errors = []
    summary = None

    header_map = {
        'item': 'item',
        'numero_do_beneficio': 'benefit_number',
        'numero_da_cat': 'numero_cat',
        'cnpj_do_empregador': 'cnpj_empregador',
        'tipo': 'benefit_type',
        'nit_do_empregado': 'insured_nit',
        'cpf_do_beneficiario': 'cpf_beneficiario',
        'data_de_nascimento_do_beneficiario': 'data_nascimento_beneficiario',
        'renda_mensal_inicial_rmi_r': 'rmi',
        'data_de_despacho_do_beneficio_ddb': 'ddb',
        'data_de_inicio_do_beneficio_dib': 'data_inicio_beneficio',
        'data_de_cessacao_do_beneficio_dcb': 'data_fim_beneficio',
        'custo': 'custo',
        'data_da_cat': 'accident_date',
        'nome': 'insured_name',
        'obs': 'notes',
        'tela_fap': 'tela_fap',
        'calculo': 'calculo',
        'gerid': 'gerid'
    }

    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename:
            flash('Selecione um arquivo Excel (.xlsx) para importar.', 'warning')
            return render_template('cases/benefits_import.html', case=case, errors=errors, summary=summary)

        filename = secure_filename(file.filename)
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

        if ext == 'xlsx':
            rows = _read_excel_rows(file)
        else:
            flash('Formato não suportado. Use Excel (.xlsx).', 'danger')
            return render_template('cases/benefits_import.html', case=case, errors=errors, summary=summary)

        if not rows:
            flash('Arquivo vazio ou sem linhas de dados.', 'warning')
            return render_template('cases/benefits_import.html', case=case, errors=errors, summary=summary)

        created = 0
        updated = 0
        skipped = 0

        for index, row in enumerate(rows, start=2):
            if not row:
                continue
            sheet_name = row.pop('__sheet__', 'Planilha')
            row_number = row.pop('__row__', index)
            vigencia_year = sheet_name.strip() if isinstance(sheet_name, str) else None
            if not (vigencia_year and vigencia_year.isdigit()):
                vigencia_year = None
            if all(value is None or str(value).strip() == '' for value in row.values()):
                continue

            raw_data = {}
            for header, value in row.items():
                normalized = _normalize_header(header)
                field = header_map.get(normalized)
                if not field and 'obs' in normalized:
                    field = 'notes'
                if not field:
                    continue
                if field == 'notes':
                    if value not in (None, ''):
                        existing_notes = raw_data.get('notes')
                        if existing_notes:
                            raw_data['notes'] = f"{existing_notes}\n{value}"
                        else:
                            raw_data['notes'] = value
                    continue
                raw_data[field] = value

            benefit_number = str(raw_data.get('benefit_number', '')).strip()
            benefit_type = str(raw_data.get('benefit_type', '')).strip()
            insured_name = str(raw_data.get('insured_name', '')).strip()

            if not benefit_number or not benefit_type or not insured_name:
                skipped += 1
                errors.append(
                    f'Planilha {sheet_name} - linha {row_number}: campos obrigatórios ausentes (Número do Benefício, TIPO, NOME).'
                )
                continue

            existing_benefit = CaseBenefit.query.filter_by(
                case_id=case_id,
                benefit_number=benefit_number
            ).first()

            if existing_benefit:
                if vigencia_year:
                    existing_years = []
                    if existing_benefit.fap_vigencia_years:
                        existing_years = [
                            year.strip() for year in existing_benefit.fap_vigencia_years.split(',') if year.strip()
                        ]
                    if vigencia_year not in existing_years:
                        existing_years.append(vigencia_year)
                        existing_benefit.fap_vigencia_years = ','.join(existing_years)
                updated += 1
                continue

            extra_notes = []
            extra_map = {
                'item': 'ITEM',
                'cnpj_empregador': 'CNPJ do Empregador',
                'cpf_beneficiario': 'CPF do Beneficiário',
                'data_nascimento_beneficiario': 'Data de Nascimento',
                'rmi': 'RMI (R$)',
                'ddb': 'Data de Despacho (DDB)',
                'custo': 'Custo',
                'tela_fap': 'Tela FAP',
                'calculo': 'Cálculo',
                'gerid': 'GERID'
            }
            for key, label in extra_map.items():
                value = raw_data.get(key)
                if value not in (None, ''):
                    extra_notes.append(f'{label}: {value}')

            base_notes = str(raw_data.get('notes')).strip() if raw_data.get('notes') else ''
            if base_notes and extra_notes:
                notes_value = f"{base_notes}\n" + "\n".join(extra_notes)
            elif base_notes:
                notes_value = base_notes
            elif extra_notes:
                notes_value = "\n".join(extra_notes)
            else:
                notes_value = None

            print(base_notes)
            exit()

            benefit = CaseBenefit(
                case_id=case_id,
                benefit_number=benefit_number,
                benefit_type=benefit_type,
                insured_name=insured_name,
                insured_nit=str(raw_data.get('insured_nit')).strip() if raw_data.get('insured_nit') else None,
                numero_cat=str(raw_data.get('numero_cat')).strip() if raw_data.get('numero_cat') else None,
                data_inicio_beneficio=_parse_date(raw_data.get('data_inicio_beneficio')),
                data_fim_beneficio=_parse_date(raw_data.get('data_fim_beneficio')),
                accident_date=_parse_date(raw_data.get('accident_date')),
                fap_vigencia_years=vigencia_year,
                notes=notes_value
            )

            db.session.add(benefit)
            created += 1

        if created == 0 and updated == 0:
            flash('Nenhum benefício foi importado.', 'warning')
            return render_template('cases/benefits_import.html', case=case, errors=errors, summary=summary)

        db.session.commit()

        summary = {
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'errors_count': len(errors)
        }

        if errors:
            flash('Importação concluída com avisos. Verifique os detalhes abaixo.', 'warning')
            return render_template('cases/benefits_import.html', case=case, errors=errors, summary=summary)

        flash(f'Importação concluída: {created} benefício(s) adicionados, {updated} atualizado(s).', 'success')
        return redirect(url_for('benefits.case_benefits_list', case_id=case_id))

    return render_template('cases/benefits_import.html', case=case, errors=errors, summary=summary)

@benefits_bp.route('/case/<int:case_id>/new', methods=['GET', 'POST'])
def case_benefit_new(case_id):
    from app.form import CaseBenefitContextForm
    case = Case.query.get_or_404(case_id)
    form = CaseBenefitContextForm()
    
    # Populate fap_reason_id choices
    fap_reasons = FapReason.query.filter_by(
        law_firm_id=case.law_firm_id,
        is_active=True
    ).order_by(FapReason.display_name).all()
    fap_reason_choices = [('', 'Nenhum motivo selecionado')] + [(str(r.id), r.display_name) for r in fap_reasons]
    form.fap_reason_id.choices = fap_reason_choices
    
    # Populate fap_vigencia_years choices based on case dates
    if case.fap_start_year and case.fap_end_year:
        years = [str(year) for year in range(case.fap_start_year, case.fap_end_year + 1)]
        form.fap_vigencia_years.choices = [(year, year) for year in years]
    else:
        form.fap_vigencia_years.choices = []
    
    if form.validate_on_submit():
        benefit = CaseBenefit(
            case_id=case_id,
            benefit_number=form.benefit_number.data,
            benefit_type=form.benefit_type.data,
            insured_name=form.insured_name.data,
            insured_nit=form.insured_nit.data,
            numero_cat=form.numero_cat.data,
            numero_bo=form.numero_bo.data,
            data_inicio_beneficio=form.data_inicio_beneficio.data,
            data_fim_beneficio=form.data_fim_beneficio.data,
            accident_date=form.accident_date.data,
            accident_company_name=form.accident_company_name.data,
            accident_summary=form.accident_summary.data,
            fap_reason_id=int(form.fap_reason_id.data) if form.fap_reason_id.data else None,
            fap_vigencia_years=','.join(form.fap_vigencia_years.data) if form.fap_vigencia_years.data else None,
            notes=form.notes.data
        )
        
        db.session.add(benefit)
        try:
            db.session.commit()
            flash('Benefício cadastrado com sucesso!', 'success')
            return redirect(url_for('benefits.case_benefits_list', case_id=case_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar benefício: {str(e)}', 'danger')
    
    return render_template('cases/benefit_form.html', form=form, case=case, title='Novo Benefício')

@benefits_bp.route('/case/<int:case_id>/<int:benefit_id>/edit', methods=['GET', 'POST'])
def case_benefit_edit(case_id, benefit_id):
    from app.form import CaseBenefitContextForm
    case = Case.query.get_or_404(case_id)
    benefit = CaseBenefit.query.get_or_404(benefit_id)
    
    if benefit.case_id != case_id:
        flash('Benefício não encontrado neste caso.', 'error')
        return redirect(url_for('benefits.case_benefits_list', case_id=case_id))
    
    # Populate fap_reason_id choices FIRST
    fap_reasons = FapReason.query.filter_by(
        law_firm_id=case.law_firm_id,
        is_active=True
    ).order_by(FapReason.display_name).all()
    fap_reason_choices = [('', 'Nenhum motivo selecionado')] + [(str(r.id), r.display_name) for r in fap_reasons]
    
    # Populate fap_vigencia_years choices based on case dates
    fap_vigencia_choices = []
    if case.fap_start_year and case.fap_end_year:
        years = [str(year) for year in range(case.fap_start_year, case.fap_end_year + 1)]
        fap_vigencia_choices = [(year, year) for year in years]
    
    # Now create the form with object data
    form = CaseBenefitContextForm(obj=benefit)
    
    # Set the choices after creating the form
    form.fap_reason_id.choices = fap_reason_choices
    form.fap_vigencia_years.choices = fap_vigencia_choices
    
    # Pre-fill fap_vigencia_years with existing values
    if benefit.fap_vigencia_years:
        form.fap_vigencia_years.data = benefit.fap_vigencia_years.split(',')
    
    # Ensure fap_reason_id is properly set as string
    if benefit.fap_reason_id:
        form.fap_reason_id.data = str(benefit.fap_reason_id)
    
    if form.validate_on_submit():
        benefit.benefit_number = form.benefit_number.data
        benefit.benefit_type = form.benefit_type.data
        benefit.insured_name = form.insured_name.data
        benefit.insured_nit = form.insured_nit.data
        benefit.numero_cat = form.numero_cat.data
        benefit.numero_bo = form.numero_bo.data
        benefit.data_inicio_beneficio = form.data_inicio_beneficio.data
        benefit.data_fim_beneficio = form.data_fim_beneficio.data
        benefit.accident_date = form.accident_date.data
        benefit.accident_company_name = form.accident_company_name.data
        benefit.accident_summary = form.accident_summary.data
        benefit.fap_reason_id = int(form.fap_reason_id.data) if form.fap_reason_id.data else None
        benefit.fap_vigencia_years = ','.join(form.fap_vigencia_years.data) if form.fap_vigencia_years.data else None
        benefit.notes = form.notes.data
        benefit.updated_at = datetime.utcnow()
        
        try:
            db.session.commit()
            flash('Benefício atualizado com sucesso!', 'success')
            return redirect(url_for('benefits.case_benefits_list', case_id=case_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar benefício: {str(e)}', 'danger')
    
    return render_template('cases/benefit_form.html', form=form, case=case, title='Editar Benefício', benefit_id=benefit_id)

@benefits_bp.route('/case/<int:case_id>/<int:benefit_id>/delete', methods=['POST'])
def case_benefit_delete(case_id, benefit_id):
    case = Case.query.get_or_404(case_id)
    benefit = CaseBenefit.query.get_or_404(benefit_id)
    
    if benefit.case_id != case_id:
        flash('Benefício não encontrado neste caso.', 'error')
        return redirect(url_for('benefits.case_benefits_list', case_id=case_id))
    
    try:
        db.session.delete(benefit)
        db.session.commit()
        flash('Benefício excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir benefício: {str(e)}', 'danger')
    
    return redirect(url_for('benefits.case_benefits_list', case_id=case_id))
