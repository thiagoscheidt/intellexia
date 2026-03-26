from datetime import datetime
from functools import wraps
from io import BytesIO
from decimal import Decimal
import hashlib
import json
import os

from flask import Blueprint, current_app, flash, jsonify, redirect, render_template, request, session, url_for, send_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import String, and_, case, cast, func, or_
from werkzeug.utils import secure_filename

from app.models import (
    Benefit,
    BenefitFapSourceHistory,
    BenefitFapVigenciaCnpj,
    Client,
    FapContestationJudgmentReport,
    KnowledgeBase,
    db,
)


central_benefits_bp = Blueprint('central_benefits', __name__, url_prefix='/central-benefits')


FILTER_FIELD_MAP = {
    'id': Benefit.id,
    'benefit_number': Benefit.benefit_number,
    'client_name': Client.name,
    'insured_name': Benefit.insured_name,
    'benefit_type': Benefit.benefit_type,
    'insured_cpf': Benefit.insured_cpf,
    'insured_nit': Benefit.insured_nit,
    'insured_date_of_birth': Benefit.insured_date_of_birth,
    'employer_cnpj': Benefit.employer_cnpj,
    'employer_name': Benefit.employer_name,
    'status': Benefit.status,
    'first_instance_status': Benefit.first_instance_status,
    'second_instance_status': Benefit.second_instance_status,
    'fap_vigencia_years': Benefit.fap_vigencia_years,
    'benefit_start_date': Benefit.benefit_start_date,
    'benefit_end_date': Benefit.benefit_end_date,
    'accident_date': Benefit.accident_date,
    'accident_company_name': Benefit.accident_company_name,
    'cat_number': Benefit.cat_number,
    'bo_number': Benefit.bo_number,
    'request_type': Benefit.request_type,
    'initial_monthly_benefit': Benefit.initial_monthly_benefit,
    'total_paid': Benefit.total_paid,
    'justification': Benefit.justification,
    'opinion': Benefit.opinion,
    'notes': Benefit.notes,
}


ORDER_COLUMN_MAP = {
    0: Benefit.id,
    1: Benefit.benefit_number,
    2: Benefit.insured_name,
    3: Benefit.employer_cnpj,
    4: Benefit.first_instance_status,
    5: Benefit.second_instance_status,
    6: Benefit.fap_vigencia_years,
    7: Benefit.benefit_start_date,
}


def _format_date(value):
    return value.strftime('%d/%m/%Y') if value else ''


def _format_decimal(value):
    if value is None:
        return ''
    if isinstance(value, Decimal):
        return f'{value:.2f}'
    return str(value)


def _format_datetime(value):
    return value.strftime('%d/%m/%Y %H:%M:%S') if value else None


def _normalize_text(value):
    return (value or '').strip()


def _normalize_cnpj_digits(value):
    return ''.join(ch for ch in (value or '') if ch.isdigit())


def _normalize_status_key(value):
    return _normalize_text(value).lower()


def _extract_cnpj_root(cnpj):
    digits = _normalize_cnpj_digits(cnpj)
    return digits[:8] if len(digits) >= 8 else ''


def _extract_cnpj_branch(cnpj):
    digits = _normalize_cnpj_digits(cnpj)
    return digits[8:12] if len(digits) >= 12 else ''


def _format_cnpj(value):
    digits = _normalize_cnpj_digits(value)
    if len(digits) != 14:
        return value or ''
    return f'{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:14]}'


def _build_client_cnpj_lookup(clients):
    exact_lookup = {}
    root_lookup = {}

    for client in clients:
        digits = _normalize_cnpj_digits(client.cnpj)
        if len(digits) != 14:
            continue

        exact_lookup[digits] = client

        root = digits[:8]
        branch = digits[8:12]
        if root not in root_lookup or branch == '0001':
            root_lookup[root] = client

    return exact_lookup, root_lookup


def _resolve_client_for_vigencia(employer_cnpj, clients_by_exact, clients_by_root):
    digits = _normalize_cnpj_digits(employer_cnpj)
    if len(digits) != 14:
        return None

    client = clients_by_exact.get(digits)
    if client is not None:
        return client

    return clients_by_root.get(digits[:8])


def _compute_file_hash_from_path(file_path):
    hasher = hashlib.sha256()

    with open(file_path, 'rb') as stream:
        while True:
            chunk = stream.read(8192)
            if not chunk:
                break
            hasher.update(chunk)

    return hasher.hexdigest()


def _resolve_existing_file_path(file_path: str | None) -> str | None:
    if not file_path:
        return None

    candidates = [file_path]
    if not os.path.isabs(file_path):
        candidates.append(os.path.abspath(os.path.join(current_app.root_path, '..', file_path)))
        candidates.append(os.path.abspath(file_path))

    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            return candidate
    return None


def _ensure_fap_report_in_knowledge_base(law_firm_id, user_id, filename, file_path, file_size, file_type):
    file_hash = _compute_file_hash_from_path(file_path)

    duplicate = KnowledgeBase.query.filter_by(
        law_firm_id=law_firm_id,
        file_hash=file_hash,
        is_active=True,
    ).first()

    if duplicate:
        # Se já existe hash igual, garante que o caminho armazenado continue apontando para um arquivo real.
        if (not duplicate.file_path or not os.path.exists(duplicate.file_path)) and os.path.exists(file_path):
            duplicate.file_path = file_path
            duplicate.file_size = file_size
            duplicate.file_type = file_type
            duplicate.updated_at = datetime.utcnow()
        return duplicate

    knowledge_file = KnowledgeBase(
        user_id=user_id,
        law_firm_id=law_firm_id,
        original_filename=filename,
        file_path=file_path,
        file_size=file_size,
        file_type=file_type,
        file_hash=file_hash,
        description='Relatório de julgamento de contestação do FAP importado pelo painel de benefícios.',
        category='Relatórios FAP',
        tags='fap,contestacao,julgamento',
        lawsuit_number=None,
        processing_status='pending',
        is_active=True,
    )
    db.session.add(knowledge_file)
    db.session.flush()
    return knowledge_file


def _apply_text_operator(column, operator, value):
    column_text = func.lower(cast(column, String))
    value_text = (value or '').strip().lower()

    if operator == 'empty':
        return or_(column.is_(None), func.trim(cast(column, String)) == '')
    if operator == 'not_empty':
        return and_(column.is_not(None), func.trim(cast(column, String)) != '')
    if operator == 'equals':
        return column_text == value_text
    if operator == 'starts_with':
        return column_text.like(f'{value_text}%')
    if operator == 'ends_with':
        return column_text.like(f'%{value_text}')

    return column_text.like(f'%{value_text}%')


def _parse_custom_filters(raw_filters):
    if not raw_filters:
        return []

    if isinstance(raw_filters, str):
        try:
            loaded = json.loads(raw_filters)
        except json.JSONDecodeError:
            return []
    elif isinstance(raw_filters, list):
        loaded = raw_filters
    else:
        return []

    valid_filters = []
    for item in loaded:
        if not isinstance(item, dict):
            continue
        field = (item.get('field') or '').strip()
        operator = (item.get('operator') or 'contains').strip()
        value = item.get('value')
        if field not in FILTER_FIELD_MAP:
            continue
        if operator not in {'contains', 'equals', 'starts_with', 'ends_with', 'empty', 'not_empty'}:
            continue
        if operator in {'contains', 'equals', 'starts_with', 'ends_with'} and not str(value or '').strip():
            continue
        valid_filters.append({'field': field, 'operator': operator, 'value': value})

    return valid_filters


def _base_benefits_query(law_firm_id):
    return (
        db.session.query(Benefit, Client.name.label('client_name'))
        .outerjoin(Client, Benefit.client_id == Client.id)
        .filter(Benefit.law_firm_id == law_firm_id)
    )


def _group_count_by_status(query, column):
    status_expr = func.lower(func.coalesce(cast(column, String), ''))
    rows = query.with_entities(status_expr.label('status_key'), func.count(Benefit.id)).group_by(status_expr).all()
    return {_normalize_status_key(status): int(count or 0) for status, count in rows}


def _build_instance_stats(total_count, status_counts):
    approved = int(status_counts.get('deferido', 0) or 0)
    rejected = int(status_counts.get('indeferido', 0) or 0)
    analyzing = int(status_counts.get('analyzing', 0) or 0)
    pending = max(int(total_count) - approved - rejected - analyzing, 0)
    return {
        'approved': approved,
        'rejected': rejected,
        'analyzing': analyzing,
        'pending': pending,
    }


def _apply_benefits_filters(query, search_value='', custom_filters=None, quick_client='', quick_root='', vigencia_id=None):
    search_text = (search_value or '').strip().lower()
    if search_text:
        like_term = f'%{search_text}%'
        query = query.filter(
            or_(
                func.lower(cast(Benefit.id, String)).like(like_term),
                func.lower(cast(Benefit.benefit_number, String)).like(like_term),
                func.lower(cast(Client.name, String)).like(like_term),
                func.lower(cast(Benefit.insured_name, String)).like(like_term),
                func.lower(cast(Benefit.benefit_type, String)).like(like_term),
                func.lower(cast(Benefit.insured_cpf, String)).like(like_term),
                func.lower(cast(Benefit.insured_nit, String)).like(like_term),
                func.lower(cast(Benefit.employer_cnpj, String)).like(like_term),
                func.lower(cast(Benefit.employer_name, String)).like(like_term),
                func.lower(cast(Benefit.status, String)).like(like_term),
                func.lower(cast(Benefit.fap_vigencia_years, String)).like(like_term),
            )
        )

    if quick_client:
        query = query.filter(func.lower(cast(Client.name, String)) == quick_client.strip().lower())

    if quick_root:
        root = ''.join(ch for ch in quick_root if ch.isdigit())[:8]
        if root:
            sanitized_cnpj = func.replace(
                func.replace(
                    func.replace(func.replace(cast(Benefit.employer_cnpj, String), '.', ''), '/', ''),
                    '-',
                    '',
                ),
                ' ',
                '',
            )
            query = query.filter(sanitized_cnpj.like(f'{root}%'))

    if vigencia_id:
        try:
            query = query.filter(Benefit.fap_vigencia_cnpj_id == int(vigencia_id))
        except (TypeError, ValueError):
            pass

    for item in custom_filters or []:
        column = FILTER_FIELD_MAP.get(item['field'])
        if column is None:
            continue
        query = query.filter(_apply_text_operator(column, item['operator'], item.get('value')))

    return query


def _serialize_benefit_row(benefit, client_name):
    return {
        'id': benefit.id,
        'benefit_number': benefit.benefit_number or '',
        'client_name': client_name or '',
        'insured_name': benefit.insured_name or '',
        'benefit_type': benefit.benefit_type or '',
        'insured_cpf': benefit.insured_cpf or '',
        'insured_nit': benefit.insured_nit or '',
        'insured_date_of_birth': _format_date(benefit.insured_date_of_birth),
        'employer_cnpj': benefit.employer_cnpj or '',
        'employer_name': benefit.employer_name or '',
        'status': benefit.status or '',
        'first_instance_status': benefit.first_instance_status or '',
        'first_instance_justification': benefit.first_instance_justification or '',
        'first_instance_opinion': benefit.first_instance_opinion or '',
        'second_instance_status': benefit.second_instance_status or '',
        'second_instance_justification': benefit.second_instance_justification or '',
        'second_instance_opinion': benefit.second_instance_opinion or '',
        'fap_vigencia_years': benefit.fap_vigencia_years or '',
        'benefit_start_date': _format_date(benefit.benefit_start_date),
        'benefit_end_date': _format_date(benefit.benefit_end_date),
        'accident_date': _format_date(benefit.accident_date),
        'accident_company_name': benefit.accident_company_name or '',
        'cat_number': benefit.cat_number or '',
        'bo_number': benefit.bo_number or '',
        'request_type': benefit.request_type or '',
        'initial_monthly_benefit': _format_decimal(benefit.initial_monthly_benefit),
        'total_paid': _format_decimal(benefit.total_paid),
        'justification': benefit.justification or '',
        'opinion': benefit.opinion or '',
        'notes': benefit.notes or '',
        'timeline_url': url_for('central_benefits.benefit_file_timeline', benefit_id=benefit.id),
        'edit_url': url_for('central_benefits.edit_central_benefit', benefit_id=benefit.id),
        'delete_url': url_for('central_benefits.delete_central_benefit', benefit_id=benefit.id),
    }


def _collect_listing_payload(default_length=25):
    if request.is_json:
        payload = request.get_json(silent=True) or {}
        return {
            'draw': int(payload.get('draw', 1) or 1),
            'start': max(int(payload.get('start', 0) or 0), 0),
            'length': max(int(payload.get('length', default_length) or default_length), 1),
            'search': _normalize_text(payload.get('search', '')),
            'order_column': int(payload.get('order_column', 0) or 0),
            'order_dir': 'desc' if str(payload.get('order_dir', 'desc')).lower() == 'desc' else 'asc',
            'filters': _parse_custom_filters(payload.get('filters')),
            'quick_client': _normalize_text(payload.get('quick_client', '')),
            'quick_root': _normalize_text(payload.get('quick_root', '')),
            'vigencia_id': payload.get('vigencia_id'),
        }

    draw = int(request.args.get('draw', 1) or 1)
    start = max(int(request.args.get('start', 0) or 0), 0)
    length = max(int(request.args.get('length', default_length) or default_length), 1)
    order_column = int(request.args.get('order[0][column]', 0) or 0)
    order_dir = 'desc' if (request.args.get('order[0][dir]', 'desc') or '').lower() == 'desc' else 'asc'
    search_value = _normalize_text(request.args.get('search[value]', ''))

    filters = _parse_custom_filters(request.args.get('custom_filters', '[]'))
    quick_client = _normalize_text(request.args.get('quick_client', ''))
    quick_root = _normalize_text(request.args.get('quick_root', ''))

    return {
        'draw': draw,
        'start': start,
        'length': length,
        'search': search_value,
        'order_column': order_column,
        'order_dir': order_dir,
        'filters': filters,
        'quick_client': quick_client,
        'quick_root': quick_root,
        'vigencia_id': request.args.get('vigencia_id'),
    }


def get_current_law_firm_id():
    return session.get('law_firm_id')


def require_law_firm(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not get_current_law_firm_id():
            if request.is_json:
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)

    return decorated_function


@central_benefits_bp.route('/<int:benefit_id>/file-timeline', methods=['GET'])
@require_law_firm
def benefit_file_timeline(benefit_id):
    law_firm_id = get_current_law_firm_id()
    benefit = Benefit.query.filter_by(id=benefit_id, law_firm_id=law_firm_id).first_or_404()

    history_items = (
        BenefitFapSourceHistory.query.filter_by(
            law_firm_id=law_firm_id,
            benefit_id=benefit_id,
        )
        .order_by(
            func.coalesce(
                BenefitFapSourceHistory.publication_datetime,
                BenefitFapSourceHistory.transmission_datetime,
            ).is_(None).asc(),
            func.coalesce(
                BenefitFapSourceHistory.publication_datetime,
                BenefitFapSourceHistory.transmission_datetime,
            ).desc(),
            BenefitFapSourceHistory.created_at.desc(),
        )
        .all()
    )

    events = []
    for item in history_items:
        report = item.report
        events.append(
            {
                'history_id': item.id,
                'report_id': item.report_id,
                'knowledge_base_id': item.knowledge_base_id,
                'action': item.action,
                'transmission_datetime': _format_datetime(item.transmission_datetime),
                'publication_datetime': _format_datetime(item.publication_datetime or item.transmission_datetime),
                'created_at': _format_datetime(item.created_at),
                'report_uploaded_at': _format_datetime(report.uploaded_at if report else None),
                'report_filename': (report.original_filename if report else None) or '-',
                'knowledge_details_url': (
                    url_for('central_benefits.view_fap_contestation_report', report_id=item.report_id)
                    if report else None
                ),
            }
        )

    return jsonify(
        {
            'benefit_id': benefit.id,
            'benefit_number': benefit.benefit_number,
            'events': events,
        }
    )


@central_benefits_bp.route('/fap-contestation-reports/<int:report_id>/view', methods=['GET'])
@require_law_firm
def view_fap_contestation_report(report_id):
    law_firm_id = get_current_law_firm_id()
    report = FapContestationJudgmentReport.query.filter_by(
        id=report_id,
        law_firm_id=law_firm_id,
    ).first_or_404()

    resolved_path = _resolve_existing_file_path(report.file_path)
    if not resolved_path:
        flash('Arquivo não encontrado no servidor.', 'error')
        return redirect(url_for('central_benefits.fap_contestation_reports'))

    mimetype = 'application/pdf'
    ext = (report.file_type or '').strip().lower()
    if ext in {'doc', 'docx'}:
        mimetype = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    elif ext in {'xls', 'xlsx'}:
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif ext == 'txt':
        mimetype = 'text/plain'

    return send_file(resolved_path, as_attachment=False, mimetype=mimetype)


@central_benefits_bp.route('/')
@require_law_firm
def list_central_benefits():
    law_firm_id = get_current_law_firm_id()
    current_vigencia_filter = None
    total_count = Benefit.query.filter_by(law_firm_id=law_firm_id).count()
    approved_count = Benefit.query.filter_by(law_firm_id=law_firm_id, status='approved').count()
    in_review_count = Benefit.query.filter(
        Benefit.law_firm_id == law_firm_id,
        func.lower(cast(Benefit.status, String)).in_(['in_review', 'analyzing']),
    ).count()
    rejected_count = Benefit.query.filter_by(law_firm_id=law_firm_id, status='rejected').count()
    pending_count = max(total_count - approved_count - in_review_count - rejected_count, 0)

    general_query = _base_benefits_query(law_firm_id)
    first_instance_stats = _build_instance_stats(
        total_count,
        _group_count_by_status(general_query, Benefit.first_instance_status),
    )
    second_instance_stats = _build_instance_stats(
        total_count,
        _group_count_by_status(general_query, Benefit.second_instance_status),
    )

    clients_data = (
        Client.query.with_entities(Client.cnpj, Client.name)
        .filter_by(law_firm_id=law_firm_id)
        .all()
    )

    roots_map = {}
    for cnpj, name in clients_data:
        root = _extract_cnpj_root(cnpj)
        if not root:
            continue

        branch = _extract_cnpj_branch(cnpj)
        clean_name = (name or '').strip()

        if root not in roots_map:
            roots_map[root] = {'root': root, 'company_name': '', 'is_main': False}

        current = roots_map[root]
        if branch == '0001' and clean_name:
            current['company_name'] = clean_name
            current['is_main'] = True
            continue

        if not current['is_main'] and clean_name and not current['company_name']:
            current['company_name'] = clean_name

    cnpj_roots = [
        {'root': item['root'], 'company_name': item['company_name']}
        for _, item in sorted(roots_map.items(), key=lambda entry: entry[0])
    ]

    client_options = sorted({(name or '').strip() for _, name in clients_data if (name or '').strip()})

    current_vigencia_id = _normalize_text(request.args.get('vigencia_id', ''))
    if current_vigencia_id:
        try:
            vigencia = BenefitFapVigenciaCnpj.query.filter_by(
                id=int(current_vigencia_id),
                law_firm_id=law_firm_id,
            ).first()
        except (TypeError, ValueError):
            vigencia = None

        if vigencia is not None:
            clients = Client.query.filter_by(law_firm_id=law_firm_id).all()
            clients_by_exact, clients_by_root = _build_client_cnpj_lookup(clients)
            resolved_client = _resolve_client_for_vigencia(
                vigencia.employer_cnpj,
                clients_by_exact,
                clients_by_root,
            )
            company_name = ''
            if resolved_client is not None:
                company_name = (resolved_client.name or '').strip()
            if not company_name:
                company_name = (
                    db.session.query(Benefit.employer_name)
                    .filter(
                        Benefit.law_firm_id == law_firm_id,
                        Benefit.fap_vigencia_cnpj_id == vigencia.id,
                        Benefit.employer_name.is_not(None),
                        func.trim(Benefit.employer_name) != '',
                    )
                    .order_by(Benefit.updated_at.desc(), Benefit.id.desc())
                    .scalar()
                    or ''
                ).strip()

            current_vigencia_filter = {
                'id': vigencia.id,
                'year': (vigencia.vigencia_year or '').strip(),
                'company_name': company_name,
                'company_cnpj': _format_cnpj(vigencia.employer_cnpj),
            }

    return render_template(
        'central_benefits/list.html',
        total_count=total_count,
        approved_count=approved_count,
        in_review_count=in_review_count,
        rejected_count=rejected_count,
        pending_count=pending_count,
        first_instance_stats=first_instance_stats,
        second_instance_stats=second_instance_stats,
        cnpj_roots=cnpj_roots,
        client_options=client_options,
        current_vigencia_filter=current_vigencia_filter,
    )


@central_benefits_bp.route('/vigencias', methods=['GET'])
@require_law_firm
def list_fap_vigencias():
    law_firm_id = get_current_law_firm_id()
    first_instance_key = func.lower(func.coalesce(cast(Benefit.first_instance_status, String), ''))
    second_instance_key = func.lower(func.coalesce(cast(Benefit.second_instance_status, String), ''))

    clients = (
        Client.query.filter_by(law_firm_id=law_firm_id)
        .order_by(Client.name.asc())
        .all()
    )
    clients_by_exact, clients_by_root = _build_client_cnpj_lookup(clients)

    vigencia_rows = (
        db.session.query(
            BenefitFapVigenciaCnpj,
            func.count(Benefit.id).label('benefits_count'),
            func.max(Benefit.updated_at).label('last_benefit_update'),
            func.sum(
                case((and_(Benefit.id.is_not(None), first_instance_key == 'deferido'), 1), else_=0)
            ).label('first_approved_count'),
            func.sum(
                case((and_(Benefit.id.is_not(None), first_instance_key == 'indeferido'), 1), else_=0)
            ).label('first_rejected_count'),
            func.sum(
                case((and_(Benefit.id.is_not(None), first_instance_key == 'analyzing'), 1), else_=0)
            ).label('first_in_review_count'),
            func.sum(
                case(
                    (
                        and_(
                            Benefit.id.is_not(None),
                            ~first_instance_key.in_(['deferido', 'indeferido', 'analyzing']),
                        ),
                        1,
                    ),
                    else_=0,
                )
            ).label('first_pending_count'),
            func.sum(
                case((and_(Benefit.id.is_not(None), second_instance_key == 'deferido'), 1), else_=0)
            ).label('second_approved_count'),
            func.sum(
                case((and_(Benefit.id.is_not(None), second_instance_key == 'indeferido'), 1), else_=0)
            ).label('second_rejected_count'),
            func.sum(
                case((and_(Benefit.id.is_not(None), second_instance_key == 'analyzing'), 1), else_=0)
            ).label('second_in_review_count'),
            func.sum(
                case(
                    (
                        and_(
                            Benefit.id.is_not(None),
                            ~second_instance_key.in_(['deferido', 'indeferido', 'analyzing']),
                        ),
                        1,
                    ),
                    else_=0,
                )
            ).label('second_pending_count'),
            func.sum(
                case(
                    (
                        and_(
                            Benefit.id.is_not(None),
                            second_instance_key.in_(['deferido', 'indeferido', 'analyzing']),
                        ),
                        1,
                    ),
                    else_=0,
                )
            ).label('second_instance_activity_count'),
            func.sum(
                case(
                    (
                        and_(
                            Benefit.id.is_not(None),
                            ~first_instance_key.in_(['deferido', 'indeferido']),
                        ),
                        1,
                    ),
                    else_=0,
                )
            ).label('first_instance_eligible_count'),
        )
        .outerjoin(Benefit, Benefit.fap_vigencia_cnpj_id == BenefitFapVigenciaCnpj.id)
        .filter(BenefitFapVigenciaCnpj.law_firm_id == law_firm_id)
        .group_by(BenefitFapVigenciaCnpj.id)
        .order_by(BenefitFapVigenciaCnpj.vigencia_year.desc(), BenefitFapVigenciaCnpj.employer_cnpj.asc())
        .all()
    )

    grouped_clients = {}
    total_benefits_linked = 0

    for (
        vigencia,
        benefits_count,
        last_benefit_update,
        first_approved_count,
        first_rejected_count,
        first_in_review_count,
        first_pending_count,
        second_approved_count,
        second_rejected_count,
        second_in_review_count,
        second_pending_count,
        second_instance_activity_count,
        first_instance_eligible_count,
    ) in vigencia_rows:
        resolved_client = _resolve_client_for_vigencia(
            vigencia.employer_cnpj,
            clients_by_exact,
            clients_by_root,
        )

        if resolved_client is not None:
            group_key = f'client:{resolved_client.id}'
            client_name = resolved_client.name or 'Cliente sem nome'
            client_cnpj = resolved_client.cnpj or _format_cnpj(vigencia.employer_cnpj)
        else:
            group_key = f'unlinked:{_normalize_cnpj_digits(vigencia.employer_cnpj) or vigencia.id}'
            client_name = 'Sem cliente vinculado'
            client_cnpj = _format_cnpj(vigencia.employer_cnpj)

        if group_key not in grouped_clients:
            grouped_clients[group_key] = {
                'client_id': resolved_client.id if resolved_client is not None else None,
                'client_name': client_name,
                'client_cnpj': client_cnpj,
                'is_unlinked': resolved_client is None,
                'total_vigencias': 0,
                'total_benefits': 0,
                'vigencias': [],
            }

        benefits_count = int(benefits_count or 0)
        first_approved_count = int(first_approved_count or 0)
        first_rejected_count = int(first_rejected_count or 0)
        first_in_review_count = int(first_in_review_count or 0)
        first_pending_count = int(first_pending_count or 0)
        second_approved_count = int(second_approved_count or 0)
        second_rejected_count = int(second_rejected_count or 0)
        second_in_review_count = int(second_in_review_count or 0)
        second_pending_count = int(second_pending_count or 0)
        second_instance_activity_count = int(second_instance_activity_count or 0)
        first_instance_eligible_count = int(first_instance_eligible_count or 0)
        grouped_clients[group_key]['total_vigencias'] += 1
        grouped_clients[group_key]['total_benefits'] += benefits_count
        grouped_clients[group_key]['vigencias'].append(
            {
                'id': vigencia.id,
                'vigencia_year': vigencia.vigencia_year,
                'employer_cnpj': _format_cnpj(vigencia.employer_cnpj),
                'benefits_count': benefits_count,
                'first_approved_count': first_approved_count,
                'first_rejected_count': first_rejected_count,
                'first_in_review_count': first_in_review_count,
                'first_pending_count': first_pending_count,
                'second_approved_count': second_approved_count,
                'second_rejected_count': second_rejected_count,
                'second_in_review_count': second_in_review_count,
                'second_pending_count': second_pending_count,
                'can_mark_first_instance_deferred': (
                    second_instance_activity_count > 0 and first_instance_eligible_count > 0
                ),
                'first_instance_eligible_count': first_instance_eligible_count,
                'created_at': _format_datetime(vigencia.created_at),
                'updated_at': _format_datetime(vigencia.updated_at),
                'last_benefit_update': _format_datetime(last_benefit_update),
                'benefits_view_url': url_for('central_benefits.list_central_benefits', vigencia_id=vigencia.id),
            }
        )
        total_benefits_linked += benefits_count

    grouped_client_list = sorted(
        grouped_clients.values(),
        key=lambda item: (
            item['is_unlinked'],
            (item['client_name'] or '').lower(),
            item['client_cnpj'] or '',
        ),
    )

    for group in grouped_client_list:
        group['vigencias'].sort(
            key=lambda item: (
                str(item['vigencia_year'] or ''),
                item['employer_cnpj'] or '',
            ),
            reverse=True,
        )

    linked_clients_count = sum(1 for item in grouped_client_list if not item['is_unlinked'])
    unlinked_groups_count = sum(1 for item in grouped_client_list if item['is_unlinked'])

    return render_template(
        'central_benefits/vigencias.html',
        grouped_clients=grouped_client_list,
        total_vigencias=len(vigencia_rows),
        total_benefits_linked=total_benefits_linked,
        linked_clients_count=linked_clients_count,
        unlinked_groups_count=unlinked_groups_count,
    )


@central_benefits_bp.route('/vigencias/<int:vigencia_id>/mark-first-instance-deferred', methods=['POST'])
@require_law_firm
def mark_vigencia_first_instance_deferred(vigencia_id):
    law_firm_id = get_current_law_firm_id()

    vigencia = BenefitFapVigenciaCnpj.query.filter_by(
        id=vigencia_id,
        law_firm_id=law_firm_id,
    ).first_or_404()

    second_instance_activity_exists = db.session.query(Benefit.id).filter(
        Benefit.law_firm_id == law_firm_id,
        Benefit.fap_vigencia_cnpj_id == vigencia.id,
        func.lower(func.coalesce(cast(Benefit.second_instance_status, String), '')).in_(['deferido', 'indeferido', 'analyzing']),
    ).first() is not None

    if not second_instance_activity_exists:
        flash(
            'A ação em lote só pode ser aplicada quando houver decisão em 2ª instância ou benefício em análise.',
            'warning',
        )
        return redirect(url_for('central_benefits.list_fap_vigencias'))

    eligible_benefits = Benefit.query.filter(
        Benefit.law_firm_id == law_firm_id,
        Benefit.fap_vigencia_cnpj_id == vigencia.id,
        ~func.lower(func.coalesce(cast(Benefit.first_instance_status, String), '')).in_(['deferido', 'indeferido']),
    ).all()

    if not eligible_benefits:
        flash('Não há benefícios elegíveis para marcar como deferido na 1ª instância.', 'info')
        return redirect(url_for('central_benefits.list_fap_vigencias'))

    try:
        now = datetime.utcnow()
        for benefit in eligible_benefits:
            benefit.first_instance_status = 'deferido'
            benefit.updated_at = now

        db.session.commit()
        flash(
            f'{len(eligible_benefits)} benefício(s) da vigência {vigencia.vigencia_year or "-"} marcado(s) como deferido na 1ª instância.',
            'success',
        )
    except Exception as exc:
        db.session.rollback()
        flash(f'Erro ao aplicar atualização em lote: {str(exc)}', 'danger')

    return redirect(url_for('central_benefits.list_fap_vigencias'))


@central_benefits_bp.route('/api/list', methods=['GET'])
@require_law_firm
def list_central_benefits_api():
    law_firm_id = get_current_law_firm_id()
    payload = _collect_listing_payload(default_length=25)

    total_query = db.session.query(func.count(Benefit.id)).filter(Benefit.law_firm_id == law_firm_id)
    records_total = total_query.scalar() or 0

    filtered_query = _apply_benefits_filters(
        _base_benefits_query(law_firm_id),
        search_value=payload['search'],
        custom_filters=payload['filters'],
        quick_client=payload['quick_client'],
        quick_root=payload['quick_root'],
        vigencia_id=payload.get('vigencia_id'),
    )

    records_filtered = filtered_query.with_entities(func.count(Benefit.id)).scalar() or 0

    status_counts = _group_count_by_status(filtered_query, Benefit.status)
    approved_filtered = int(status_counts.get('approved', 0) or 0)
    in_review_filtered = int(status_counts.get('in_review', 0) or 0) + int(status_counts.get('analyzing', 0) or 0)
    rejected_filtered = int(status_counts.get('rejected', 0) or 0)
    pending_filtered = max(int(records_filtered) - approved_filtered - in_review_filtered - rejected_filtered, 0)
    filtered_first_instance_stats = _build_instance_stats(
        records_filtered,
        _group_count_by_status(filtered_query, Benefit.first_instance_status),
    )
    filtered_second_instance_stats = _build_instance_stats(
        records_filtered,
        _group_count_by_status(filtered_query, Benefit.second_instance_status),
    )

    order_column = ORDER_COLUMN_MAP.get(payload['order_column'], Benefit.id)
    if payload['order_dir'] == 'asc':
        filtered_query = filtered_query.order_by(order_column.asc(), Benefit.id.asc())
    else:
        filtered_query = filtered_query.order_by(order_column.desc(), Benefit.id.desc())

    paged_results = (
        filtered_query
        .offset(payload['start'])
        .limit(payload['length'])
        .all()
    )

    data = [_serialize_benefit_row(benefit, client_name) for benefit, client_name in paged_results]

    return jsonify(
        {
            'draw': payload['draw'],
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'filtered_stats': {
                'total': int(records_filtered),
                'approved': approved_filtered,
                'rejected': rejected_filtered,
                'in_review': in_review_filtered,
                'pending': pending_filtered,
                'first_instance': filtered_first_instance_stats,
                'second_instance': filtered_second_instance_stats,
            },
            'data': data,
        }
    )


@central_benefits_bp.route('/export-excel', methods=['POST'])
@require_law_firm
def export_central_benefits_excel():
    law_firm_id = get_current_law_firm_id()
    payload = _collect_listing_payload(default_length=1000)

    filtered_query = _apply_benefits_filters(
        _base_benefits_query(law_firm_id),
        search_value=payload['search'],
        custom_filters=payload['filters'],
        quick_client=payload['quick_client'],
        quick_root=payload['quick_root'],
        vigencia_id=payload.get('vigencia_id'),
    )

    order_column = ORDER_COLUMN_MAP.get(payload['order_column'], Benefit.id)
    if payload['order_dir'] == 'asc':
        filtered_query = filtered_query.order_by(order_column.asc(), Benefit.id.asc())
    else:
        filtered_query = filtered_query.order_by(order_column.desc(), Benefit.id.desc())

    benefits = filtered_query.all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Beneficios'

    headers = [
        'ID',
        'Número do benefício',
        'Cliente',
        'Segurado',
        'Tipo benefício',
        'CPF segurado',
        'NIT segurado',
        'Data nascimento segurado',
        'CNPJ empregador',
        'Nome empregador',
        'Status geral',
        'Status 1ª instância',
        'Justificativa 1ª instância',
        'Parecer 1ª instância',
        'Status 2ª instância',
        'Justificativa 2ª instância',
        'Parecer 2ª instância',
        'Vigência FAP',
        'DIB',
        'DCB',
        'Data acidente',
        'Empresa acidente',
        'CAT',
        'BO',
        'Tipo solicitação',
        'RMI',
        'Total pago',
        'Justificativa geral',
        'Parecer geral',
        'Observações',
    ]
    sheet.append(headers)

    for benefit, client_name in benefits:
        sheet.append(
            [
                benefit.id,
                benefit.benefit_number or '',
                client_name or '',
                benefit.insured_name or '',
                benefit.benefit_type or '',
                benefit.insured_cpf or '',
                benefit.insured_nit or '',
                _format_date(benefit.insured_date_of_birth),
                benefit.employer_cnpj or '',
                benefit.employer_name or '',
                benefit.status or '',
                benefit.first_instance_status or '',
                benefit.first_instance_justification or '',
                benefit.first_instance_opinion or '',
                benefit.second_instance_status or '',
                benefit.second_instance_justification or '',
                benefit.second_instance_opinion or '',
                benefit.fap_vigencia_years or '',
                _format_date(benefit.benefit_start_date),
                _format_date(benefit.benefit_end_date),
                _format_date(benefit.accident_date),
                benefit.accident_company_name or '',
                benefit.cat_number or '',
                benefit.bo_number or '',
                benefit.request_type or '',
                _format_decimal(benefit.initial_monthly_benefit),
                _format_decimal(benefit.total_paid),
                benefit.justification or '',
                benefit.opinion or '',
                benefit.notes or '',
            ]
        )

    for idx, _ in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = 22

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'beneficios_centralizados_{timestamp}.xlsx'

    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@central_benefits_bp.route('/fap-contestation-reports', methods=['GET', 'POST'])
@require_law_firm
def fap_contestation_reports():
    from app.form import FapContestationJudgmentReportForm

    law_firm_id = get_current_law_firm_id()
    form = FapContestationJudgmentReportForm()
    allowed_extensions = {'pdf', 'doc', 'docx', 'txt', 'xlsx', 'xls'}

    if request.method == 'POST' and form.validate_on_submit():
        files = request.files.getlist('file') if 'file' in request.files else []
        files = [f for f in files if f and f.filename and f.filename.strip()]

        if not files:
            flash('Selecione ao menos um arquivo para envio.', 'warning')
            return redirect(url_for('central_benefits.fap_contestation_reports'))

        invalid_files = []
        success_count = 0
        upload_dir = os.path.abspath(
            os.path.join(current_app.root_path, '..', 'uploads', 'fap_contestation_reports')
        )
        os.makedirs(upload_dir, exist_ok=True)

        for file in files:
            filename = secure_filename(file.filename)
            extension = os.path.splitext(filename)[1].lower().replace('.', '')
            if extension not in allowed_extensions:
                invalid_files.append(filename)
                continue

            try:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
                unique_filename = f'{timestamp}_{filename}'
                file_path = os.path.join(upload_dir, unique_filename)
                file.save(file_path)

                file_size = os.path.getsize(file_path)
                file_type = extension.upper()

                report = FapContestationJudgmentReport(
                    user_id=session.get('user_id'),
                    law_firm_id=law_firm_id,
                    original_filename=filename,
                    file_path=file_path,
                    file_size=file_size,
                    file_type=file_type,
                    status='pending',
                )

                db.session.add(report)
                db.session.flush()
                knowledge_file = _ensure_fap_report_in_knowledge_base(
                    law_firm_id=law_firm_id,
                    user_id=session.get('user_id'),
                    filename=filename,
                    file_path=file_path,
                    file_size=file_size,
                    file_type=file_type,
                )
                report.knowledge_base_id = knowledge_file.id
                db.session.commit()
                success_count += 1
            except Exception as e:
                db.session.rollback()
                flash(f'Erro ao enviar {filename}: {str(e)}', 'danger')

        if success_count:
            flash(
                f'{success_count} relatório(s) enviado(s) com sucesso! Eles ficarão pendentes até processamento via script e também foram adicionados à base de conhecimento sem vínculo com processo.',
                'success'
            )

        if invalid_files:
            flash(
                'Arquivos ignorados por extensão inválida: ' + ', '.join(invalid_files),
                'warning'
            )

        return redirect(url_for('central_benefits.fap_contestation_reports'))

    reports = (
        FapContestationJudgmentReport.query.filter_by(law_firm_id=law_firm_id)
        .order_by(FapContestationJudgmentReport.uploaded_at.desc())
        .all()
    )

    return render_template('central_benefits/fap_contestation_reports.html', form=form, reports=reports)


@central_benefits_bp.route('/fap-contestation-reports/<int:report_id>/delete', methods=['POST'])
@require_law_firm
def delete_fap_contestation_report(report_id):
    law_firm_id = get_current_law_firm_id()
    report = FapContestationJudgmentReport.query.filter_by(id=report_id, law_firm_id=law_firm_id).first_or_404()

    try:
        if report.file_path and os.path.exists(report.file_path):
            os.remove(report.file_path)
        db.session.delete(report)
        db.session.commit()
        flash('Relatório excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir relatório: {str(e)}', 'danger')

    return redirect(url_for('central_benefits.fap_contestation_reports'))


@central_benefits_bp.route('/new', methods=['GET', 'POST'])
@require_law_firm
def new_central_benefit():
    from app.form import CentralBenefitForm

    law_firm_id = get_current_law_firm_id()
    form = CentralBenefitForm()

    clients = (
        Client.query.filter_by(law_firm_id=law_firm_id)
        .order_by(Client.name.asc())
        .all()
    )
    form.client_id.choices = [('', 'Sem cliente vinculado')] + [(c.id, c.name) for c in clients]

    if form.validate_on_submit():
        benefit = Benefit(
            law_firm_id=law_firm_id,
            client_id=form.client_id.data,
            benefit_number=form.benefit_number.data,
            benefit_type=form.benefit_type.data,
            insured_name=form.insured_name.data,
            insured_nit=form.insured_nit.data,
            insured_cpf=form.insured_cpf.data,
            insured_date_of_birth=form.insured_date_of_birth.data,
            employer_cnpj=form.employer_cnpj.data,
            employer_name=form.employer_name.data,
            benefit_start_date=form.benefit_start_date.data,
            benefit_end_date=form.benefit_end_date.data,
            initial_monthly_benefit=form.initial_monthly_benefit.data,
            total_paid=form.total_paid.data,
            accident_date=form.accident_date.data,
            accident_company_name=form.accident_company_name.data,
            accident_summary=form.accident_summary.data,
            cat_number=form.cat_number.data,
            bo_number=form.bo_number.data,
            fap_vigencia_years=form.fap_vigencia_years.data,
            request_type=form.request_type.data or None,
            status=form.status.data,
            first_instance_status=form.first_instance_status.data or None,
            first_instance_justification=form.first_instance_justification.data,
            first_instance_opinion=form.first_instance_opinion.data,
            second_instance_status=form.second_instance_status.data or None,
            second_instance_justification=form.second_instance_justification.data,
            second_instance_opinion=form.second_instance_opinion.data,
            justification=form.justification.data,
            opinion=form.opinion.data,
            notes=form.notes.data,
        )

        db.session.add(benefit)
        try:
            db.session.commit()
            flash('Benefício centralizado cadastrado com sucesso!', 'success')
            return redirect(url_for('central_benefits.list_central_benefits'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar benefício: {str(e)}', 'danger')

    return render_template('central_benefits/form.html', form=form, title='Novo Benefício Centralizado')


@central_benefits_bp.route('/<int:benefit_id>/edit', methods=['GET', 'POST'])
@require_law_firm
def edit_central_benefit(benefit_id):
    from app.form import CentralBenefitForm

    law_firm_id = get_current_law_firm_id()
    benefit = Benefit.query.filter_by(id=benefit_id, law_firm_id=law_firm_id).first_or_404()
    form = CentralBenefitForm(obj=benefit)

    clients = (
        Client.query.filter_by(law_firm_id=law_firm_id)
        .order_by(Client.name.asc())
        .all()
    )
    form.client_id.choices = [('', 'Sem cliente vinculado')] + [(c.id, c.name) for c in clients]

    if request.method == 'GET':
        form.client_id.data = benefit.client_id

    if form.validate_on_submit():
        benefit.client_id = form.client_id.data
        benefit.benefit_number = form.benefit_number.data
        benefit.benefit_type = form.benefit_type.data
        benefit.insured_name = form.insured_name.data
        benefit.insured_nit = form.insured_nit.data
        benefit.insured_cpf = form.insured_cpf.data
        benefit.insured_date_of_birth = form.insured_date_of_birth.data
        benefit.employer_cnpj = form.employer_cnpj.data
        benefit.employer_name = form.employer_name.data
        benefit.benefit_start_date = form.benefit_start_date.data
        benefit.benefit_end_date = form.benefit_end_date.data
        benefit.initial_monthly_benefit = form.initial_monthly_benefit.data
        benefit.total_paid = form.total_paid.data
        benefit.accident_date = form.accident_date.data
        benefit.accident_company_name = form.accident_company_name.data
        benefit.accident_summary = form.accident_summary.data
        benefit.cat_number = form.cat_number.data
        benefit.bo_number = form.bo_number.data
        benefit.fap_vigencia_years = form.fap_vigencia_years.data
        benefit.request_type = form.request_type.data or None
        benefit.status = form.status.data
        benefit.first_instance_status = form.first_instance_status.data or None
        benefit.first_instance_justification = form.first_instance_justification.data
        benefit.first_instance_opinion = form.first_instance_opinion.data
        benefit.second_instance_status = form.second_instance_status.data or None
        benefit.second_instance_justification = form.second_instance_justification.data
        benefit.second_instance_opinion = form.second_instance_opinion.data
        benefit.justification = form.justification.data
        benefit.opinion = form.opinion.data
        benefit.notes = form.notes.data
        benefit.updated_at = datetime.utcnow()

        try:
            db.session.commit()
            flash('Benefício centralizado atualizado com sucesso!', 'success')
            return redirect(url_for('central_benefits.list_central_benefits'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar benefício: {str(e)}', 'danger')

    import json as _json
    clients_data = {str(c.id): {'name': c.name, 'cnpj': c.cnpj or ''} for c in clients}

    return render_template(
        'central_benefits/form.html',
        form=form,
        title='Editar Benefício Centralizado',
        benefit_id=benefit_id,
        clients_data=_json.dumps(clients_data),
    )


@central_benefits_bp.route('/<int:benefit_id>/delete', methods=['POST'])
@require_law_firm
def delete_central_benefit(benefit_id):
    law_firm_id = get_current_law_firm_id()
    benefit = Benefit.query.filter_by(id=benefit_id, law_firm_id=law_firm_id).first_or_404()

    try:
        db.session.delete(benefit)
        db.session.commit()
        flash('Benefício centralizado excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir benefício: {str(e)}', 'danger')

    return redirect(url_for('central_benefits.list_central_benefits'))
