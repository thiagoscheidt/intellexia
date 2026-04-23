from datetime import datetime
from functools import wraps
from io import BytesIO
from decimal import Decimal
import hashlib
import json
import os

from app.services.fap_web_service import FapWebAuthPayload, FapWebService

from app.utils.timezone import now_sp

from flask import Blueprint, current_app, flash, jsonify, redirect, render_template, request, session, url_for, send_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import String, and_, case, cast, func, or_
from werkzeug.utils import secure_filename

from app.models import (
    Benefit,
    BenefitManualHistory,
    BenefitFapSourceHistory,
    FapVigenciaCnpj,
    FapCompany,
    FapAutoImportedContestacao,
    FapWebContestacao,
    Client,
    FapContestationCat,
    FapContestationCatSourceHistory,
    FapContestationCatManualHistory,
    FapContestationJudgmentReport,
    FapContestationPayrollMass,
    FapContestationPayrollMassSourceHistory,
    FapContestationPayrollMassManualHistory,
    FapContestationEmploymentLink,
    FapContestationEmploymentLinkSourceHistory,
    FapContestationEmploymentLinkManualHistory,
    FapContestationTurnoverRate,
    FapContestationTurnoverRateSourceHistory,
    FapContestationTurnoverRateManualHistory,
    KnowledgeBase,
    db,
)


disputes_center_bp = Blueprint('disputes_center', __name__, url_prefix='/disputes-center')


FILTER_FIELD_MAP = {
    'id': Benefit.id,
    'benefit_number': Benefit.benefit_number,
    'client_name': Client.name,
    'insured_name': Benefit.insured_name,
    'benefit_type': Benefit.benefit_type,
    'insured_cpf': Benefit.insured_cpf,
    'insured_nit': Benefit.insured_nit,
    'insured_date_of_birth': Benefit.insured_date_of_birth,
    'employer_cnpj': Benefit.employer_cnpj,
    'employer_name': Benefit.employer_name,
    'status': Benefit.status,
    'first_instance_status': Benefit.first_instance_status,
    'first_instance_status_raw': Benefit.first_instance_status_raw,
    'second_instance_status': Benefit.second_instance_status,
    'second_instance_status_raw': Benefit.second_instance_status_raw,
    'fap_vigencia_years': Benefit.fap_vigencia_years,
    'fap_contestation_topic': Benefit.fap_contestation_topic,
    'benefit_start_date': Benefit.benefit_start_date,
    'benefit_end_date': Benefit.benefit_end_date,
    'accident_date': Benefit.accident_date,
    'accident_company_name': Benefit.accident_company_name,
    'cat_number': Benefit.cat_number,
    'bo_number': Benefit.bo_number,
    'request_type': Benefit.request_type,
    'initial_monthly_benefit': Benefit.initial_monthly_benefit,
    'total_paid': Benefit.total_paid,
    'justification': Benefit.justification,
    'opinion': Benefit.opinion,
    'notes': Benefit.notes,
}


ORDER_COLUMN_MAP = {
    0: Benefit.id,
    1: Benefit.benefit_number,
    2: Benefit.insured_name,
    3: Benefit.employer_cnpj,
    4: Benefit.first_instance_status,
    5: Benefit.second_instance_status,
    6: Benefit.fap_vigencia_years,
    7: Benefit.fap_contestation_topic,
    8: Benefit.benefit_start_date,
}


CAT_FILTER_FIELD_MAP = {
    'id': FapContestationCat.id,
    'cat_number': FapContestationCat.cat_number,
    'employer_cnpj': FapContestationCat.employer_cnpj,
    'employer_cnpj_assigned': FapContestationCat.employer_cnpj_assigned,
    'insured_nit': FapContestationCat.insured_nit,
    'cat_block': FapContestationCat.cat_block,
    'status': FapContestationCat.status,
    'first_instance_status': FapContestationCat.first_instance_status,
    'first_instance_status_raw': FapContestationCat.first_instance_status_raw,
    'second_instance_status': FapContestationCat.second_instance_status,
    'second_instance_status_raw': FapContestationCat.second_instance_status_raw,
}

CAT_ORDER_COLUMN_MAP = {
    0: FapContestationCat.id,
    1: FapContestationCat.cat_number,
    2: FapContestationCat.employer_cnpj,
    3: FapContestationCat.insured_nit,
    4: FapContestationCat.first_instance_status,
    5: FapContestationCat.second_instance_status,
    6: FapContestationCat.accident_date,
    7: FapContestationCat.cat_registration_date,
}

PAYROLL_MASS_FILTER_FIELD_MAP = {
    'id': FapContestationPayrollMass.id,
    'employer_cnpj': FapContestationPayrollMass.employer_cnpj,
    'competence': FapContestationPayrollMass.competence,
    'status': FapContestationPayrollMass.status,
    'first_instance_status': FapContestationPayrollMass.first_instance_status,
    'first_instance_status_raw': FapContestationPayrollMass.first_instance_status_raw,
    'second_instance_status': FapContestationPayrollMass.second_instance_status,
    'second_instance_status_raw': FapContestationPayrollMass.second_instance_status_raw,
}

PAYROLL_MASS_ORDER_COLUMN_MAP = {
    0: FapContestationPayrollMass.id,
    1: FapContestationPayrollMass.employer_cnpj,
    2: FapContestationPayrollMass.competence,
    3: FapContestationPayrollMass.total_remuneration,
    4: FapContestationPayrollMass.first_instance_requested_value,
    5: FapContestationPayrollMass.first_instance_status,
    6: FapContestationPayrollMass.second_instance_status,
}

EMPLOYMENT_LINK_FILTER_FIELD_MAP = {
    'id': FapContestationEmploymentLink.id,
    'employer_cnpj': FapContestationEmploymentLink.employer_cnpj,
    'competence': FapContestationEmploymentLink.competence,
    'status': FapContestationEmploymentLink.status,
    'first_instance_status': FapContestationEmploymentLink.first_instance_status,
    'first_instance_status_raw': FapContestationEmploymentLink.first_instance_status_raw,
    'second_instance_status': FapContestationEmploymentLink.second_instance_status,
    'second_instance_status_raw': FapContestationEmploymentLink.second_instance_status_raw,
}

EMPLOYMENT_LINK_ORDER_COLUMN_MAP = {
    0: FapContestationEmploymentLink.id,
    1: FapContestationEmploymentLink.employer_cnpj,
    2: FapContestationEmploymentLink.competence,
    3: FapContestationEmploymentLink.quantity,
    4: FapContestationEmploymentLink.first_instance_requested_quantity,
    5: FapContestationEmploymentLink.first_instance_status,
    6: FapContestationEmploymentLink.second_instance_status,
}

TURNOVER_RATE_FILTER_FIELD_MAP = {
    'id': FapContestationTurnoverRate.id,
    'employer_cnpj': FapContestationTurnoverRate.employer_cnpj,
    'year': FapContestationTurnoverRate.year,
    'status': FapContestationTurnoverRate.status,
    'first_instance_status': FapContestationTurnoverRate.first_instance_status,
    'first_instance_status_raw': FapContestationTurnoverRate.first_instance_status_raw,
    'second_instance_status': FapContestationTurnoverRate.second_instance_status,
    'second_instance_status_raw': FapContestationTurnoverRate.second_instance_status_raw,
}

TURNOVER_RATE_ORDER_COLUMN_MAP = {
    0: FapContestationTurnoverRate.id,
    1: FapContestationTurnoverRate.employer_cnpj,
    2: FapContestationTurnoverRate.year,
    3: FapContestationTurnoverRate.turnover_rate,
    4: FapContestationTurnoverRate.admissions,
    5: FapContestationTurnoverRate.first_instance_status,
    6: FapContestationTurnoverRate.second_instance_status,
}


def _format_date(value):
    return value.strftime('%d/%m/%Y') if value else ''


def _format_decimal(value):
    if value is None:
        return ''
    if isinstance(value, Decimal):
        return f'{value:.2f}'
    return str(value)


def _format_datetime(value):
    return value.strftime('%d/%m/%Y %H:%M:%S') if value else None


def _normalize_text(value):
    return (value or '').strip()


def _normalize_cnpj_digits(value):
    return ''.join(ch for ch in (value or '') if ch.isdigit())


def _normalize_status_key(value):
    return _normalize_text(value).lower()


def _normalize_optional_status(value):
    normalized = _normalize_text(value)
    return normalized or None


STATUS_LABEL_PT_MAP = {
    'analyzing': 'Em análise',
    'in_review': 'Em análise',
    'approved': 'Deferido',
    'deferido': 'Deferido',
    'rejected': 'Indeferido',
    'indeferido': 'Indeferido',
    'pending': 'Pendente',
    'pendente': 'Pendente',
}


def _status_label_pt(value):
    normalized = _normalize_text(value)
    if not normalized:
        return 'Sem status'

    translated = STATUS_LABEL_PT_MAP.get(normalized.lower())
    return translated or normalized


def _resolve_general_status_key(first_instance_status, second_instance_status, fallback_status=None):
    valid_instance_statuses = {'deferido', 'indeferido', 'analyzing'}

    normalized_second = _normalize_status_key(second_instance_status)
    if normalized_second in valid_instance_statuses:
        return normalized_second

    normalized_first = _normalize_status_key(first_instance_status)
    if normalized_first in valid_instance_statuses:
        return normalized_first

    normalized_fallback = _normalize_status_key(fallback_status)
    return normalized_fallback or 'pending'


def _resolve_general_status_excel_value(
    first_instance_status,
    second_instance_status,
    first_instance_status_raw=None,
    second_instance_status_raw=None,
    fallback_status=None,
):
    general_status_key = _resolve_general_status_key(
        first_instance_status,
        second_instance_status,
        fallback_status,
    )

    if general_status_key != 'deferido':
        return _status_label_pt(general_status_key)

    if _normalize_status_key(second_instance_status) == 'deferido':
        return _normalize_text(second_instance_status_raw) or _status_label_pt(general_status_key)

    if _normalize_status_key(first_instance_status) == 'deferido':
        return _normalize_text(first_instance_status_raw) or _status_label_pt(general_status_key)

    return _status_label_pt(general_status_key)


def _extract_cnpj_root(cnpj):
    digits = _normalize_cnpj_digits(cnpj)
    return digits[:8] if len(digits) >= 8 else ''


def _extract_cnpj_branch(cnpj):
    digits = _normalize_cnpj_digits(cnpj)
    return digits[8:12] if len(digits) >= 12 else ''


def _get_cnpj_establishment_type(cnpj):
    branch = _extract_cnpj_branch(cnpj)
    if not branch:
        return ''
    return 'Matriz' if branch == '0001' else 'Filial'


def _format_cnpj(value):
    digits = _normalize_cnpj_digits(value)
    if len(digits) != 14:
        return value or ''
    return f'{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:14]}'


def _matches_vigencia_filters(vigencia, resolved_client, client_id=None, client_cnpj='', client_root=''):
    vigencia_cnpj_digits = _normalize_cnpj_digits(vigencia.employer_cnpj)
    resolved_client_cnpj_digits = _normalize_cnpj_digits(resolved_client.cnpj if resolved_client else '')

    if client_id is not None and (resolved_client is None or resolved_client.id != client_id):
        return False

    if client_cnpj and client_cnpj not in {vigencia_cnpj_digits, resolved_client_cnpj_digits}:
        return False

    if client_root:
        matching_roots = {
            vigencia_cnpj_digits[:8],
            resolved_client_cnpj_digits[:8],
        }
        if client_root not in matching_roots:
            return False

    return True


def _build_unique_root_options(clients, vigencia_rows):
    root_map = {}

    for client in clients:
        root = _extract_cnpj_root(client.cnpj)
        if not root:
            continue

        current_name = (client.name or '').strip()
        if root not in root_map or current_name:
            root_map[root] = current_name

    for vigencia, *_ in vigencia_rows:
        root = _extract_cnpj_root(vigencia.employer_cnpj)
        if not root or root in root_map:
            continue
        root_map[root] = ''

    return [
        {
            'root': root,
            'label': f'{root} - {name}' if name else root,
        }
        for root, name in sorted(root_map.items(), key=lambda item: item[0])
    ]


def _build_client_cnpj_lookup(clients):
    exact_lookup = {}
    root_lookup = {}

    for client in clients:
        digits = _normalize_cnpj_digits(client.cnpj)
        if len(digits) != 14:
            continue

        exact_lookup[digits] = client

        root = digits[:8]
        branch = digits[8:12]
        if root not in root_lookup or branch == '0001':
            root_lookup[root] = client

    return exact_lookup, root_lookup


def _resolve_client_for_vigencia(employer_cnpj, clients_by_exact, clients_by_root):
    digits = _normalize_cnpj_digits(employer_cnpj)
    if len(digits) != 14:
        return None

    client = clients_by_exact.get(digits)
    if client is not None:
        return client

    return clients_by_root.get(digits[:8])


def _compute_file_hash_from_path(file_path):
    hasher = hashlib.sha256()

    with open(file_path, 'rb') as stream:
        while True:
            chunk = stream.read(8192)
            if not chunk:
                break
            hasher.update(chunk)

    return hasher.hexdigest()


def _resolve_existing_file_path(file_path: str | None) -> str | None:
    if not file_path:
        return None

    candidates = [file_path]
    if not os.path.isabs(file_path):
        candidates.append(os.path.abspath(os.path.join(current_app.root_path, '..', file_path)))
        candidates.append(os.path.abspath(file_path))

    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            return candidate
    return None


def _ensure_fap_report_in_knowledge_base(law_firm_id, user_id, filename, file_path, file_size, file_type):
    file_hash = _compute_file_hash_from_path(file_path)

    duplicate = KnowledgeBase.query.filter_by(
        law_firm_id=law_firm_id,
        file_hash=file_hash,
        is_active=True,
    ).first()

    if duplicate:
        # Se já existe hash igual, garante que o caminho armazenado continue apontando para um arquivo real.
        if (not duplicate.file_path or not os.path.exists(duplicate.file_path)) and os.path.exists(file_path):
            duplicate.file_path = file_path
            duplicate.file_size = file_size
            duplicate.file_type = file_type
            duplicate.updated_at = datetime.utcnow()
        return duplicate

    knowledge_file = KnowledgeBase(
        user_id=user_id,
        law_firm_id=law_firm_id,
        original_filename=filename,
        file_path=file_path,
        file_size=file_size,
        file_type=file_type,
        file_hash=file_hash,
        description='Relatório de julgamento de contestação do FAP importado pelo painel de benefícios.',
        category='Relatórios FAP',
        tags='fap,contestacao,julgamento',
        lawsuit_number=None,
        processing_status='pending',
        is_active=True,
    )
    db.session.add(knowledge_file)
    db.session.flush()
    return knowledge_file


def _apply_text_operator(column, operator, value):
    column_text = func.lower(cast(column, String))
    value_text = (value or '').strip().lower()

    if operator == 'empty':
        return or_(column.is_(None), func.trim(cast(column, String)) == '')
    if operator == 'not_empty':
        return and_(column.is_not(None), func.trim(cast(column, String)) != '')
    if operator == 'equals':
        return column_text == value_text
    if operator == 'starts_with':
        return column_text.like(f'{value_text}%')
    if operator == 'ends_with':
        return column_text.like(f'%{value_text}')

    return column_text.like(f'%{value_text}%')


def _normalize_br_date_filter_value(value):
    raw = _normalize_text(value)
    if not raw:
        return raw

    parts = raw.split('/')
    if len(parts) != 3:
        return raw

    day, month, year = parts
    if not (day.isdigit() and month.isdigit() and year.isdigit()):
        return raw

    if len(day) != 2 or len(month) != 2 or len(year) != 4:
        return raw

    try:
        parsed = datetime(int(year), int(month), int(day))
    except ValueError:
        return raw

    return parsed.strftime('%Y-%m-%d')


def _parse_custom_filters(raw_filters):
    if not raw_filters:
        return []

    if isinstance(raw_filters, str):
        try:
            loaded = json.loads(raw_filters)
        except json.JSONDecodeError:
            return []
    elif isinstance(raw_filters, list):
        loaded = raw_filters
    else:
        return []

    valid_filters = []
    for item in loaded:
        if not isinstance(item, dict):
            continue
        field = (item.get('field') or '').strip()
        operator = (item.get('operator') or 'contains').strip()
        value = item.get('value')
        if field not in FILTER_FIELD_MAP:
            continue
        if operator not in {'contains', 'equals', 'starts_with', 'ends_with', 'empty', 'not_empty'}:
            continue
        if field == 'insured_date_of_birth':
            value = _normalize_br_date_filter_value(value)
        if operator in {'contains', 'equals', 'starts_with', 'ends_with'} and not str(value or '').strip():
            continue
        valid_filters.append({'field': field, 'operator': operator, 'value': value})

    return valid_filters


def _parse_int_list(values):
    parsed = []
    for value in values or []:
        try:
            parsed.append(int(value))
        except (TypeError, ValueError):
            continue
    return parsed


def _mark_first_instance_deferred_for_vigencia(law_firm_id, user_id, vigencia, note):
    second_instance_activity_exists = db.session.query(Benefit.id).filter(
        Benefit.law_firm_id == law_firm_id,
        Benefit.fap_vigencia_cnpj_id == vigencia.id,
        func.lower(func.coalesce(cast(Benefit.second_instance_status, String), '')).in_(['deferido', 'indeferido', 'analyzing']),
    ).first() is not None

    if not second_instance_activity_exists:
        return {
            'status': 'missing_second_instance_activity',
            'updated_count': 0,
        }

    eligible_benefits = Benefit.query.filter(
        Benefit.law_firm_id == law_firm_id,
        Benefit.fap_vigencia_cnpj_id == vigencia.id,
        ~func.lower(func.coalesce(cast(Benefit.first_instance_status, String), '')).in_(['deferido', 'indeferido']),
    ).all()

    if not eligible_benefits:
        return {
            'status': 'no_eligible_benefits',
            'updated_count': 0,
        }

    now = datetime.utcnow()
    history_rows = []
    for benefit in eligible_benefits:
        old_status = benefit.first_instance_status
        benefit.first_instance_status = 'deferido'
        benefit.updated_at = now

        history_rows.append(
            BenefitManualHistory(
                law_firm_id=law_firm_id,
                benefit_id=benefit.id,
                vigencia_id=vigencia.id,
                performed_by_user_id=user_id,
                action='mark_first_instance_deferred',
                old_first_instance_status=old_status,
                new_first_instance_status='deferido',
                notes=note,
                created_at=now,
                updated_at=now,
            )
        )

    if history_rows:
        db.session.add_all(history_rows)

    return {
        'status': 'updated',
        'updated_count': len(eligible_benefits),
    }


def _base_benefits_query(law_firm_id):
    return (
        db.session.query(Benefit, Client.name.label('client_name'))
        .outerjoin(Client, Benefit.client_id == Client.id)
        .filter(Benefit.law_firm_id == law_firm_id)
    )


def _base_cats_query(law_firm_id):
    return db.session.query(FapContestationCat).filter(FapContestationCat.law_firm_id == law_firm_id)


def _base_payroll_masses_query(law_firm_id):
    return db.session.query(FapContestationPayrollMass).filter(FapContestationPayrollMass.law_firm_id == law_firm_id)


def _base_employment_links_query(law_firm_id):
    return db.session.query(FapContestationEmploymentLink).filter(FapContestationEmploymentLink.law_firm_id == law_firm_id)


def _parse_cat_custom_filters(raw_filters):
    if not raw_filters:
        return []
    if isinstance(raw_filters, str):
        try:
            loaded = json.loads(raw_filters)
        except json.JSONDecodeError:
            return []
    elif isinstance(raw_filters, list):
        loaded = raw_filters
    else:
        return []
    valid_filters = []
    for item in loaded:
        if not isinstance(item, dict):
            continue
        field = (item.get('field') or '').strip()
        operator = (item.get('operator') or 'contains').strip()
        value = item.get('value')
        if field not in CAT_FILTER_FIELD_MAP:
            continue
        if operator not in {'contains', 'equals', 'starts_with', 'ends_with', 'empty', 'not_empty'}:
            continue
        if operator in {'contains', 'equals', 'starts_with', 'ends_with'} and not str(value or '').strip():
            continue
        valid_filters.append({'field': field, 'operator': operator, 'value': value})
    return valid_filters


def _apply_cats_filters(query, search_value='', custom_filters=None, quick_employer_name='', quick_root='', quick_cnpj='', vigencia_id=None):
    search_text = (search_value or '').strip().lower()
    if search_text:
        like_term = f'%{search_text}%'
        query = query.filter(
            or_(
                func.lower(cast(FapContestationCat.id, String)).like(like_term),
                func.lower(cast(FapContestationCat.cat_number, String)).like(like_term),
                func.lower(cast(FapContestationCat.employer_cnpj, String)).like(like_term),
                func.lower(cast(FapContestationCat.employer_cnpj_assigned, String)).like(like_term),
                func.lower(cast(FapContestationCat.insured_nit, String)).like(like_term),
                func.lower(cast(FapContestationCat.cat_block, String)).like(like_term),
            )
        )

    if quick_employer_name:
        query = query.filter(
            func.lower(cast(FapContestationCat.employer_name, String)) == quick_employer_name.strip().lower()
        )

    if quick_root:
        root = ''.join(ch for ch in quick_root if ch.isdigit())[:8]
        if root:
            sanitized_cnpj = func.replace(
                func.replace(
                    func.replace(func.replace(cast(FapContestationCat.employer_cnpj, String), '.', ''), '/', ''),
                    '-', '',
                ),
                ' ', '',
            )
            query = query.filter(sanitized_cnpj.like(f'{root}%'))

    if quick_cnpj:
        cnpj_digits = ''.join(ch for ch in quick_cnpj if ch.isdigit())[:14]
        if cnpj_digits:
            sanitized_employer_cnpj = func.replace(
                func.replace(
                    func.replace(func.replace(cast(FapContestationCat.employer_cnpj, String), '.', ''), '/', ''),
                    '-', '',
                ),
                ' ', '',
            )
            query = query.filter(sanitized_employer_cnpj == cnpj_digits)

    if vigencia_id:
        try:
            query = query.filter(FapContestationCat.vigencia_id == int(vigencia_id))
        except (TypeError, ValueError):
            pass

    for item in custom_filters or []:
        field = item['field']
        operator = item['operator']
        value = (item.get('value') or '').strip().lower()

        # "Status geral" uses both instance fields with second-instance priority.
        if operator == 'equals' and field == 'status':
            second_expr = func.lower(func.coalesce(cast(FapContestationCat.second_instance_status, String), ''))
            first_expr = func.lower(func.coalesce(cast(FapContestationCat.first_instance_status, String), ''))
            if value == 'pending':
                query = query.filter(
                    ~second_expr.in_(['deferido', 'indeferido', 'analyzing']),
                    ~first_expr.in_(['deferido', 'analyzing']),
                )
            else:
                query = query.filter(
                    or_(
                        second_expr == value,
                        and_(~second_expr.in_(['deferido', 'indeferido', 'analyzing']), first_expr == value),
                    )
                )
            continue

        # "Pending" for status fields must mirror the counter logic, not a simple equality check.
        if operator == 'equals' and value == 'pending':
            if field == 'first_instance_status':
                query = query.filter(
                    ~func.lower(func.coalesce(cast(FapContestationCat.first_instance_status, String), '')).in_(
                        ['deferido', 'indeferido', 'analyzing']
                    )
                )
                continue
            if field == 'second_instance_status':
                query = query.filter(
                    and_(
                        ~func.lower(func.coalesce(cast(FapContestationCat.first_instance_status, String), '')).in_(
                            ['deferido', 'analyzing']
                        ),
                        ~func.lower(func.coalesce(cast(FapContestationCat.second_instance_status, String), '')).in_(
                            ['deferido', 'indeferido', 'analyzing']
                        ),
                    )
                )
                continue

        column = CAT_FILTER_FIELD_MAP.get(field)
        if column is None:
            continue
        query = query.filter(_apply_text_operator(column, item['operator'], item.get('value')))
    return query


def _serialize_cat_row(cat):
    return {
        'id': cat.id,
        'cat_number': cat.cat_number or '',
        'vigencia_year': cat.vigencia_year or '',
        'employer_cnpj': cat.employer_cnpj or '',
        'employer_cnpj_assigned': cat.employer_cnpj_assigned or '',
        'employer_name': cat.employer_name or '',
        'insured_nit': cat.insured_nit or '',
        'insured_date_of_birth': _format_date(cat.insured_date_of_birth),
        'insured_death_date': _format_date(cat.insured_death_date),
        'accident_date': _format_date(cat.accident_date),
        'cat_registration_date': _format_date(cat.cat_registration_date),
        'cat_block': cat.cat_block or '',
        'status': cat.status or '',
        'first_instance_status': cat.first_instance_status or '',
        'first_instance_status_raw': cat.first_instance_status_raw or '',
        'first_instance_justification': cat.first_instance_justification or '',
        'first_instance_opinion': cat.first_instance_opinion or '',
        'second_instance_status': cat.second_instance_status or '',
        'second_instance_status_raw': cat.second_instance_status_raw or '',
        'second_instance_justification': cat.second_instance_justification or '',
        'second_instance_opinion': cat.second_instance_opinion or '',
        'report_id': cat.report_id,
        'report_view_url': url_for('disputes_center.view_fap_contestation_report', report_id=cat.report_id),
        'edit_url': url_for('disputes_center.edit_cat', cat_id=cat.id),
        'timeline_url': url_for('disputes_center.cat_file_timeline', cat_id=cat.id),
    }


def _parse_payroll_mass_custom_filters(raw_filters):
    if not raw_filters:
        return []
    if isinstance(raw_filters, str):
        try:
            loaded = json.loads(raw_filters)
        except json.JSONDecodeError:
            return []
    elif isinstance(raw_filters, list):
        loaded = raw_filters
    else:
        return []
    valid_filters = []
    for item in loaded:
        if not isinstance(item, dict):
            continue
        field = (item.get('field') or '').strip()
        operator = (item.get('operator') or 'contains').strip()
        value = item.get('value')
        if field not in PAYROLL_MASS_FILTER_FIELD_MAP:
            continue
        if operator not in {'contains', 'equals', 'starts_with', 'ends_with', 'empty', 'not_empty'}:
            continue
        if operator in {'contains', 'equals', 'starts_with', 'ends_with'} and not str(value or '').strip():
            continue
        valid_filters.append({'field': field, 'operator': operator, 'value': value})
    return valid_filters


def _apply_payroll_mass_filters(query, search_value='', custom_filters=None, quick_root='', quick_cnpj='', vigencia_id=None):
    search_text = (search_value or '').strip().lower()
    if search_text:
        like_term = f'%{search_text}%'
        query = query.filter(
            or_(
                func.lower(cast(FapContestationPayrollMass.id, String)).like(like_term),
                func.lower(cast(FapContestationPayrollMass.employer_cnpj, String)).like(like_term),
                func.lower(cast(FapContestationPayrollMass.employer_name, String)).like(like_term),
                func.lower(cast(FapContestationPayrollMass.competence, String)).like(like_term),
            )
        )

    if quick_root:
        root = ''.join(ch for ch in quick_root if ch.isdigit())[:8]
        if root:
            sanitized_cnpj = func.replace(
                func.replace(
                    func.replace(func.replace(cast(FapContestationPayrollMass.employer_cnpj, String), '.', ''), '/', ''),
                    '-', '',
                ),
                ' ', '',
            )
            query = query.filter(sanitized_cnpj.like(f'{root}%'))

    if quick_cnpj:
        cnpj_digits = ''.join(ch for ch in quick_cnpj if ch.isdigit())[:14]
        if cnpj_digits:
            sanitized_cnpj = func.replace(
                func.replace(
                    func.replace(func.replace(cast(FapContestationPayrollMass.employer_cnpj, String), '.', ''), '/', ''),
                    '-', '',
                ),
                ' ', '',
            )
            query = query.filter(sanitized_cnpj == cnpj_digits)

    if vigencia_id:
        try:
            query = query.filter(FapContestationPayrollMass.vigencia_id == int(vigencia_id))
        except (TypeError, ValueError):
            pass

    for item in custom_filters or []:
        field = item['field']
        operator = item['operator']
        value = (item.get('value') or '').strip().lower()

        # "Status geral" uses both instance fields with second-instance priority.
        if operator == 'equals' and field == 'status':
            second_expr = func.lower(func.coalesce(cast(FapContestationPayrollMass.second_instance_status, String), ''))
            first_expr = func.lower(func.coalesce(cast(FapContestationPayrollMass.first_instance_status, String), ''))
            if value == 'pending':
                query = query.filter(
                    ~second_expr.in_(['deferido', 'indeferido', 'analyzing']),
                    ~first_expr.in_(['deferido', 'analyzing']),
                )
            else:
                query = query.filter(
                    or_(
                        second_expr == value,
                        and_(~second_expr.in_(['deferido', 'indeferido', 'analyzing']), first_expr == value),
                    )
                )
            continue

        # "Pending" for status fields must mirror the counter logic, not a simple equality check.
        if operator == 'equals' and value == 'pending':
            if field == 'first_instance_status':
                query = query.filter(
                    ~func.lower(func.coalesce(cast(FapContestationPayrollMass.first_instance_status, String), '')).in_(
                        ['deferido', 'indeferido', 'analyzing']
                    )
                )
                continue
            if field == 'second_instance_status':
                query = query.filter(
                    and_(
                        ~func.lower(func.coalesce(cast(FapContestationPayrollMass.first_instance_status, String), '')).in_(
                            ['deferido', 'analyzing']
                        ),
                        ~func.lower(func.coalesce(cast(FapContestationPayrollMass.second_instance_status, String), '')).in_(
                            ['deferido', 'indeferido', 'analyzing']
                        ),
                    )
                )
                continue

        column = PAYROLL_MASS_FILTER_FIELD_MAP.get(field)
        if column is None:
            continue
        query = query.filter(_apply_text_operator(column, item['operator'], item.get('value')))
    return query


def _serialize_payroll_mass_row(pm):
    return {
        'id': pm.id,
        'vigencia_year': pm.vigencia_year or '',
        'employer_cnpj': pm.employer_cnpj or '',
        'employer_name': pm.employer_name or '',
        'competence': pm.competence or '',
        'total_remuneration': _format_decimal(pm.total_remuneration),
        'first_instance_requested_value': _format_decimal(pm.first_instance_requested_value),
        'second_instance_requested_value': _format_decimal(pm.second_instance_requested_value),
        'status': pm.status or '',
        'first_instance_status': pm.first_instance_status or '',
        'first_instance_status_raw': pm.first_instance_status_raw or '',
        'first_instance_justification': pm.first_instance_justification or '',
        'first_instance_opinion': pm.first_instance_opinion or '',
        'second_instance_status': pm.second_instance_status or '',
        'second_instance_status_raw': pm.second_instance_status_raw or '',
        'second_instance_justification': pm.second_instance_justification or '',
        'second_instance_opinion': pm.second_instance_opinion or '',
        'report_id': pm.report_id,
        'report_view_url': url_for('disputes_center.view_fap_contestation_report', report_id=pm.report_id),
        'edit_url': url_for('disputes_center.edit_payroll_mass', payroll_mass_id=pm.id),
        'timeline_url': url_for('disputes_center.payroll_mass_timeline', payroll_mass_id=pm.id),
    }


def _parse_employment_link_custom_filters(raw_filters):
    if not raw_filters:
        return []
    if isinstance(raw_filters, str):
        try:
            loaded = json.loads(raw_filters)
        except json.JSONDecodeError:
            return []
    elif isinstance(raw_filters, list):
        loaded = raw_filters
    else:
        return []
    valid_filters = []
    for item in loaded:
        if not isinstance(item, dict):
            continue
        field = (item.get('field') or '').strip()
        operator = (item.get('operator') or 'contains').strip()
        value = item.get('value')
        if field not in EMPLOYMENT_LINK_FILTER_FIELD_MAP:
            continue
        if operator not in {'contains', 'equals', 'starts_with', 'ends_with', 'empty', 'not_empty'}:
            continue
        if operator in {'contains', 'equals', 'starts_with', 'ends_with'} and not str(value or '').strip():
            continue
        valid_filters.append({'field': field, 'operator': operator, 'value': value})
    return valid_filters


def _apply_employment_link_filters(query, search_value='', custom_filters=None, quick_root='', quick_cnpj='', vigencia_id=None):
    search_text = (search_value or '').strip().lower()
    if search_text:
        like_term = f'%{search_text}%'
        query = query.filter(
            or_(
                func.lower(cast(FapContestationEmploymentLink.id, String)).like(like_term),
                func.lower(cast(FapContestationEmploymentLink.employer_cnpj, String)).like(like_term),
                func.lower(cast(FapContestationEmploymentLink.employer_name, String)).like(like_term),
                func.lower(cast(FapContestationEmploymentLink.competence, String)).like(like_term),
            )
        )

    if quick_root:
        root = ''.join(ch for ch in quick_root if ch.isdigit())[:8]
        if root:
            sanitized_cnpj = func.replace(
                func.replace(
                    func.replace(func.replace(cast(FapContestationEmploymentLink.employer_cnpj, String), '.', ''), '/', ''),
                    '-', '',
                ),
                ' ', '',
            )
            query = query.filter(sanitized_cnpj.like(f'{root}%'))

    if quick_cnpj:
        cnpj_digits = ''.join(ch for ch in quick_cnpj if ch.isdigit())[:14]
        if cnpj_digits:
            sanitized_cnpj = func.replace(
                func.replace(
                    func.replace(func.replace(cast(FapContestationEmploymentLink.employer_cnpj, String), '.', ''), '/', ''),
                    '-', '',
                ),
                ' ', '',
            )
            query = query.filter(sanitized_cnpj == cnpj_digits)

    if vigencia_id:
        try:
            query = query.filter(FapContestationEmploymentLink.vigencia_id == int(vigencia_id))
        except (TypeError, ValueError):
            pass

    for item in custom_filters or []:
        field = item['field']
        operator = item['operator']
        value = (item.get('value') or '').strip().lower()

        # "Status geral" uses both instance fields with second-instance priority.
        if operator == 'equals' and field == 'status':
            second_expr = func.lower(func.coalesce(cast(FapContestationEmploymentLink.second_instance_status, String), ''))
            first_expr = func.lower(func.coalesce(cast(FapContestationEmploymentLink.first_instance_status, String), ''))
            if value == 'pending':
                query = query.filter(
                    ~second_expr.in_(['deferido', 'indeferido', 'analyzing']),
                    ~first_expr.in_(['deferido', 'analyzing']),
                )
            else:
                query = query.filter(
                    or_(
                        second_expr == value,
                        and_(~second_expr.in_(['deferido', 'indeferido', 'analyzing']), first_expr == value),
                    )
                )
            continue

        # "Pending" for status fields must mirror the counter logic, not a simple equality check.
        if operator == 'equals' and value == 'pending':
            if field == 'first_instance_status':
                query = query.filter(
                    ~func.lower(func.coalesce(cast(FapContestationEmploymentLink.first_instance_status, String), '')).in_(
                        ['deferido', 'indeferido', 'analyzing']
                    )
                )
                continue
            if field == 'second_instance_status':
                query = query.filter(
                    and_(
                        ~func.lower(func.coalesce(cast(FapContestationEmploymentLink.first_instance_status, String), '')).in_(
                            ['deferido', 'analyzing']
                        ),
                        ~func.lower(func.coalesce(cast(FapContestationEmploymentLink.second_instance_status, String), '')).in_(
                            ['deferido', 'indeferido', 'analyzing']
                        ),
                    )
                )
                continue

        column = EMPLOYMENT_LINK_FILTER_FIELD_MAP.get(field)
        if column is None:
            continue
        query = query.filter(_apply_text_operator(column, item['operator'], item.get('value')))
    return query


def _serialize_employment_link_row(el):
    return {
        'id': el.id,
        'vigencia_year': el.vigencia_year or '',
        'employer_cnpj': el.employer_cnpj or '',
        'employer_name': el.employer_name or '',
        'competence': el.competence or '',
        'quantity': el.quantity if el.quantity is not None else '',
        'first_instance_requested_quantity': el.first_instance_requested_quantity if el.first_instance_requested_quantity is not None else '',
        'second_instance_requested_quantity': el.second_instance_requested_quantity if el.second_instance_requested_quantity is not None else '',
        'status': el.status or '',
        'first_instance_status': el.first_instance_status or '',
        'first_instance_status_raw': el.first_instance_status_raw or '',
        'first_instance_justification': el.first_instance_justification or '',
        'first_instance_opinion': el.first_instance_opinion or '',
        'second_instance_status': el.second_instance_status or '',
        'second_instance_status_raw': el.second_instance_status_raw or '',
        'second_instance_justification': el.second_instance_justification or '',
        'second_instance_opinion': el.second_instance_opinion or '',
        'report_id': el.report_id,
        'report_view_url': url_for('disputes_center.view_fap_contestation_report', report_id=el.report_id),
        'edit_url': url_for('disputes_center.edit_employment_link', employment_link_id=el.id),
        'timeline_url': url_for('disputes_center.employment_link_timeline', employment_link_id=el.id),
    }


def _base_turnover_rates_query(law_firm_id):
    return db.session.query(FapContestationTurnoverRate).filter(FapContestationTurnoverRate.law_firm_id == law_firm_id)


def _parse_turnover_rate_custom_filters(raw_filters):
    if not raw_filters:
        return []
    if isinstance(raw_filters, str):
        try:
            loaded = json.loads(raw_filters)
        except json.JSONDecodeError:
            return []
    elif isinstance(raw_filters, list):
        loaded = raw_filters
    else:
        return []
    valid_filters = []
    for item in loaded:
        if not isinstance(item, dict):
            continue
        field = (item.get('field') or '').strip()
        operator = (item.get('operator') or 'contains').strip()
        value = item.get('value')
        if field not in TURNOVER_RATE_FILTER_FIELD_MAP:
            continue
        if operator not in {'contains', 'equals', 'starts_with', 'ends_with', 'empty', 'not_empty'}:
            continue
        if operator in {'contains', 'equals', 'starts_with', 'ends_with'} and not str(value or '').strip():
            continue
        valid_filters.append({'field': field, 'operator': operator, 'value': value})
    return valid_filters


def _apply_turnover_rate_filters(query, search_value='', custom_filters=None, quick_root='', quick_cnpj='', vigencia_id=None):
    search_text = (search_value or '').strip().lower()
    if search_text:
        like_term = f'%{search_text}%'
        query = query.filter(
            or_(
                func.lower(cast(FapContestationTurnoverRate.id, String)).like(like_term),
                func.lower(cast(FapContestationTurnoverRate.employer_cnpj, String)).like(like_term),
                func.lower(cast(FapContestationTurnoverRate.employer_name, String)).like(like_term),
                func.lower(cast(FapContestationTurnoverRate.year, String)).like(like_term),
            )
        )

    if quick_root:
        root = ''.join(ch for ch in quick_root if ch.isdigit())[:8]
        if root:
            sanitized_cnpj = func.replace(
                func.replace(
                    func.replace(func.replace(cast(FapContestationTurnoverRate.employer_cnpj, String), '.', ''), '/', ''),
                    '-', '',
                ),
                ' ', '',
            )
            query = query.filter(sanitized_cnpj.like(f'{root}%'))

    if quick_cnpj:
        cnpj_digits = ''.join(ch for ch in quick_cnpj if ch.isdigit())[:14]
        if cnpj_digits:
            sanitized_cnpj = func.replace(
                func.replace(
                    func.replace(func.replace(cast(FapContestationTurnoverRate.employer_cnpj, String), '.', ''), '/', ''),
                    '-', '',
                ),
                ' ', '',
            )
            query = query.filter(sanitized_cnpj == cnpj_digits)

    if vigencia_id:
        try:
            query = query.filter(FapContestationTurnoverRate.vigencia_id == int(vigencia_id))
        except (TypeError, ValueError):
            pass

    for item in custom_filters or []:
        field = item['field']
        operator = item['operator']
        value = (item.get('value') or '').strip().lower()

        # "Status geral" uses both instance fields with second-instance priority.
        if operator == 'equals' and field == 'status':
            second_expr = func.lower(func.coalesce(cast(FapContestationTurnoverRate.second_instance_status, String), ''))
            first_expr = func.lower(func.coalesce(cast(FapContestationTurnoverRate.first_instance_status, String), ''))
            if value == 'pending':
                query = query.filter(
                    ~second_expr.in_(['deferido', 'indeferido', 'analyzing']),
                    ~first_expr.in_(['deferido', 'analyzing']),
                )
            else:
                query = query.filter(
                    or_(
                        second_expr == value,
                        and_(~second_expr.in_(['deferido', 'indeferido', 'analyzing']), first_expr == value),
                    )
                )
            continue

        # "Pending" for status fields must mirror the counter logic, not a simple equality check.
        if operator == 'equals' and value == 'pending':
            if field == 'first_instance_status':
                query = query.filter(
                    ~func.lower(func.coalesce(cast(FapContestationTurnoverRate.first_instance_status, String), '')).in_(
                        ['deferido', 'indeferido', 'analyzing']
                    )
                )
                continue
            if field == 'second_instance_status':
                query = query.filter(
                    and_(
                        ~func.lower(func.coalesce(cast(FapContestationTurnoverRate.first_instance_status, String), '')).in_(
                            ['deferido', 'analyzing']
                        ),
                        ~func.lower(func.coalesce(cast(FapContestationTurnoverRate.second_instance_status, String), '')).in_(
                            ['deferido', 'indeferido', 'analyzing']
                        ),
                    )
                )
                continue

        column = TURNOVER_RATE_FILTER_FIELD_MAP.get(field)
        if column is None:
            continue
        query = query.filter(_apply_text_operator(column, item['operator'], item.get('value')))
    return query


def _serialize_turnover_rate_row(tr):
    return {
        'id': tr.id,
        'vigencia_year': tr.vigencia_year or '',
        'employer_cnpj': tr.employer_cnpj or '',
        'employer_name': tr.employer_name or '',
        'year': tr.year or '',
        'turnover_rate': _format_decimal(tr.turnover_rate),
        'admissions': tr.admissions if tr.admissions is not None else '',
        'dismissals': tr.dismissals if tr.dismissals is not None else '',
        'initial_links_count': tr.initial_links_count if tr.initial_links_count is not None else '',
        'first_instance_requested_admissions': tr.first_instance_requested_admissions if tr.first_instance_requested_admissions is not None else '',
        'first_instance_requested_dismissals': tr.first_instance_requested_dismissals if tr.first_instance_requested_dismissals is not None else '',
        'first_instance_requested_initial_links': tr.first_instance_requested_initial_links if tr.first_instance_requested_initial_links is not None else '',
        'second_instance_requested_admissions': tr.second_instance_requested_admissions if tr.second_instance_requested_admissions is not None else '',
        'second_instance_requested_dismissals': tr.second_instance_requested_dismissals if tr.second_instance_requested_dismissals is not None else '',
        'second_instance_requested_initial_links': tr.second_instance_requested_initial_links if tr.second_instance_requested_initial_links is not None else '',
        'status': tr.status or '',
        'first_instance_status': tr.first_instance_status or '',
        'first_instance_status_raw': tr.first_instance_status_raw or '',
        'first_instance_justification': tr.first_instance_justification or '',
        'first_instance_opinion': tr.first_instance_opinion or '',
        'second_instance_status': tr.second_instance_status or '',
        'second_instance_status_raw': tr.second_instance_status_raw or '',
        'second_instance_justification': tr.second_instance_justification or '',
        'second_instance_opinion': tr.second_instance_opinion or '',
        'report_id': tr.report_id,
        'report_view_url': url_for('disputes_center.view_fap_contestation_report', report_id=tr.report_id),
        'edit_url': url_for('disputes_center.edit_turnover_rate', turnover_rate_id=tr.id),
        'timeline_url': url_for('disputes_center.turnover_rate_timeline', turnover_rate_id=tr.id),
    }


def _group_count_by_status(query, column):
    status_expr = func.lower(func.coalesce(cast(column, String), ''))
    rows = query.with_entities(status_expr.label('status_key'), func.count(Benefit.id)).group_by(status_expr).all()
    return {_normalize_status_key(status): int(count or 0) for status, count in rows}


def _build_instance_stats(total_count, status_counts):
    approved = int(status_counts.get('deferido', 0) or 0)
    rejected = int(status_counts.get('indeferido', 0) or 0)
    analyzing = int(status_counts.get('analyzing', 0) or 0)
    pending = max(int(total_count) - approved - rejected - analyzing, 0)
    return {
        'approved': approved,
        'rejected': rejected,
        'analyzing': analyzing,
        'pending': pending,
    }


def _apply_benefits_filters(query, search_value='', custom_filters=None, quick_client='', quick_root='', quick_cnpj='', vigencia_id=None):
    search_text = (search_value or '').strip().lower()
    if search_text:
        like_term = f'%{search_text}%'
        query = query.filter(
            or_(
                func.lower(cast(Benefit.id, String)).like(like_term),
                func.lower(cast(Benefit.benefit_number, String)).like(like_term),
                func.lower(cast(Client.name, String)).like(like_term),
                func.lower(cast(Benefit.insured_name, String)).like(like_term),
                func.lower(cast(Benefit.benefit_type, String)).like(like_term),
                func.lower(cast(Benefit.insured_cpf, String)).like(like_term),
                func.lower(cast(Benefit.insured_nit, String)).like(like_term),
                func.lower(cast(Benefit.employer_cnpj, String)).like(like_term),
                func.lower(cast(Benefit.employer_name, String)).like(like_term),
                func.lower(cast(Benefit.status, String)).like(like_term),
                func.lower(cast(Benefit.fap_vigencia_years, String)).like(like_term),
            )
        )

    if quick_client:
        query = query.filter(func.lower(cast(Client.name, String)) == quick_client.strip().lower())

    if quick_root:
        root = ''.join(ch for ch in quick_root if ch.isdigit())[:8]
        if root:
            sanitized_cnpj = func.replace(
                func.replace(
                    func.replace(func.replace(cast(Benefit.employer_cnpj, String), '.', ''), '/', ''),
                    '-',
                    '',
                ),
                ' ',
                '',
            )
            query = query.filter(sanitized_cnpj.like(f'{root}%'))

    if quick_cnpj:
        cnpj_digits = ''.join(ch for ch in quick_cnpj if ch.isdigit())[:14]
        if cnpj_digits:
            sanitized_employer_cnpj = func.replace(
                func.replace(
                    func.replace(func.replace(cast(Benefit.employer_cnpj, String), '.', ''), '/', ''),
                    '-',
                    '',
                ),
                ' ',
                '',
            )
            query = query.filter(sanitized_employer_cnpj == cnpj_digits)

    if vigencia_id:
        try:
            query = query.filter(Benefit.fap_vigencia_cnpj_id == int(vigencia_id))
        except (TypeError, ValueError):
            pass

    for item in custom_filters or []:
        field = item['field']
        operator = item['operator']
        value = (item.get('value') or '').strip().lower()

        # "Status geral" uses both instance fields with second-instance priority.
        if operator == 'equals' and field == 'status':
            second_expr = func.lower(func.coalesce(cast(Benefit.second_instance_status, String), ''))
            first_expr = func.lower(func.coalesce(cast(Benefit.first_instance_status, String), ''))
            if value == 'pending':
                query = query.filter(
                    ~second_expr.in_(['deferido', 'indeferido', 'analyzing']),
                    ~first_expr.in_(['deferido', 'analyzing']),
                )
            else:
                query = query.filter(
                    or_(
                        second_expr == value,
                        and_(~second_expr.in_(['deferido', 'indeferido', 'analyzing']), first_expr == value),
                    )
                )
            continue

        # "Pending" for status fields must mirror the counter logic, not a simple equality check.
        if operator == 'equals' and value == 'pending':
            if field == 'first_instance_status':
                query = query.filter(
                    ~func.lower(func.coalesce(cast(Benefit.first_instance_status, String), '')).in_(
                        ['deferido', 'indeferido', 'analyzing']
                    )
                )
                continue
            if field == 'second_instance_status':
                query = query.filter(
                    and_(
                        ~func.lower(func.coalesce(cast(Benefit.first_instance_status, String), '')).in_(
                            ['deferido', 'analyzing']
                        ),
                        ~func.lower(func.coalesce(cast(Benefit.second_instance_status, String), '')).in_(
                            ['deferido', 'indeferido', 'analyzing']
                        ),
                    )
                )
                continue

        column = FILTER_FIELD_MAP.get(field)
        if column is None:
            continue
        query = query.filter(_apply_text_operator(column, item['operator'], item.get('value')))

    return query


def _serialize_benefit_row(benefit, client_name):
    return {
        'id': benefit.id,
        'benefit_number': benefit.benefit_number or '',
        'client_name': client_name or '',
        'insured_name': benefit.insured_name or '',
        'benefit_type': benefit.benefit_type or '',
        'insured_cpf': benefit.insured_cpf or '',
        'insured_nit': benefit.insured_nit or '',
        'insured_date_of_birth': _format_date(benefit.insured_date_of_birth),
        'employer_cnpj': benefit.employer_cnpj or '',
        'employer_name': benefit.employer_name or '',
        'status': benefit.status or '',
        'first_instance_status': benefit.first_instance_status or '',
        'first_instance_status_raw': benefit.first_instance_status_raw or '',
        'first_instance_justification': benefit.first_instance_justification or '',
        'first_instance_opinion': benefit.first_instance_opinion or '',
        'second_instance_status': benefit.second_instance_status or '',
        'second_instance_status_raw': benefit.second_instance_status_raw or '',
        'second_instance_justification': benefit.second_instance_justification or '',
        'second_instance_opinion': benefit.second_instance_opinion or '',
        'fap_vigencia_years': benefit.fap_vigencia_years or '',
        'fap_contestation_topic': benefit.fap_contestation_topic or '',
        'benefit_start_date': _format_date(benefit.benefit_start_date),
        'benefit_end_date': _format_date(benefit.benefit_end_date),
        'accident_date': _format_date(benefit.accident_date),
        'accident_company_name': benefit.accident_company_name or '',
        'cat_number': benefit.cat_number or '',
        'bo_number': benefit.bo_number or '',
        'request_type': benefit.request_type or '',
        'initial_monthly_benefit': _format_decimal(benefit.initial_monthly_benefit),
        'total_paid': _format_decimal(benefit.total_paid),
        'justification': benefit.justification or '',
        'opinion': benefit.opinion or '',
        'notes': benefit.notes or '',
        'timeline_url': url_for('disputes_center.benefit_file_timeline', benefit_id=benefit.id),
        'edit_url': url_for('disputes_center.edit_dispute', benefit_id=benefit.id),
        'delete_url': url_for('disputes_center.delete_dispute', benefit_id=benefit.id),
    }


def _collect_listing_payload(default_length=25):
    if request.is_json:
        payload = request.get_json(silent=True) or {}
        return {
            'draw': int(payload.get('draw', 1) or 1),
            'start': max(int(payload.get('start', 0) or 0), 0),
            'length': max(int(payload.get('length', default_length) or default_length), 1),
            'search': _normalize_text(payload.get('search', '')),
            'order_column': int(payload.get('order_column', 0) or 0),
            'order_dir': 'desc' if str(payload.get('order_dir', 'desc')).lower() == 'desc' else 'asc',
            'filters': _parse_custom_filters(payload.get('filters')),
            'quick_client': _normalize_text(payload.get('quick_client', '')),
            'quick_root': _normalize_text(payload.get('quick_root', '')),
            'quick_cnpj': _normalize_text(payload.get('quick_cnpj', '')),
            'vigencia_id': payload.get('vigencia_id'),
        }

    draw = int(request.args.get('draw', 1) or 1)
    start = max(int(request.args.get('start', 0) or 0), 0)
    length = max(int(request.args.get('length', default_length) or default_length), 1)
    order_column = int(request.args.get('order[0][column]', 0) or 0)
    order_dir = 'desc' if (request.args.get('order[0][dir]', 'desc') or '').lower() == 'desc' else 'asc'
    search_value = _normalize_text(request.args.get('search[value]', ''))

    filters = _parse_custom_filters(request.args.get('custom_filters', '[]'))
    quick_client = _normalize_text(request.args.get('quick_client', ''))
    quick_root = _normalize_text(request.args.get('quick_root', ''))
    quick_cnpj = _normalize_text(request.args.get('quick_cnpj', ''))

    return {
        'draw': draw,
        'start': start,
        'length': length,
        'search': search_value,
        'order_column': order_column,
        'order_dir': order_dir,
        'filters': filters,
        'quick_client': quick_client,
        'quick_root': quick_root,
        'quick_cnpj': quick_cnpj,
        'vigencia_id': request.args.get('vigencia_id'),
    }


def get_current_law_firm_id():
    return session.get('law_firm_id')


def require_law_firm(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not get_current_law_firm_id():
            if request.is_json:
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)

    return decorated_function


@disputes_center_bp.route('/<int:benefit_id>/file-timeline', methods=['GET'])
@require_law_firm
def benefit_file_timeline(benefit_id):
    law_firm_id = get_current_law_firm_id()
    benefit = Benefit.query.filter_by(id=benefit_id, law_firm_id=law_firm_id).first_or_404()

    history_items = (
        BenefitFapSourceHistory.query.filter_by(
            law_firm_id=law_firm_id,
            benefit_id=benefit_id,
        )
        .order_by(
            func.coalesce(
                BenefitFapSourceHistory.publication_datetime,
                BenefitFapSourceHistory.transmission_datetime,
            ).is_(None).asc(),
            func.coalesce(
                BenefitFapSourceHistory.publication_datetime,
                BenefitFapSourceHistory.transmission_datetime,
            ).desc(),
            BenefitFapSourceHistory.created_at.desc(),
        )
        .all()
    )

    manual_history_items = (
        BenefitManualHistory.query.filter_by(
            law_firm_id=law_firm_id,
            benefit_id=benefit_id,
        )
        .order_by(BenefitManualHistory.created_at.desc(), BenefitManualHistory.id.desc())
        .all()
    )

    events = []
    for item in history_items:
        report = item.report
        events.append(
            {
                'event_type': 'fap_file_history',
                'history_id': item.id,
                'report_id': item.report_id,
                'knowledge_base_id': item.knowledge_base_id,
                'action': item.action,
                'transmission_datetime': _format_datetime(item.transmission_datetime),
                'publication_datetime': _format_datetime(item.publication_datetime or item.transmission_datetime),
                'created_at': _format_datetime(item.created_at),
                'report_uploaded_at': _format_datetime(report.uploaded_at if report else None),
                'report_filename': (report.original_filename if report else None) or '-',
                'knowledge_details_url': (
                    url_for('disputes_center.view_fap_contestation_report', report_id=item.report_id)
                    if report else None
                ),
                'sort_datetime': item.publication_datetime or item.transmission_datetime or item.created_at,
            }
        )

    for item in manual_history_items:
        if item.action == 'edit_dispute_first_instance_status':
            manual_action_label = 'Edição manual'
            manual_description = 'Status da 1ª instância alterado na tela de edição.'
        else:
            manual_action_label = 'Marcação manual em lote'
            manual_description = '1ª instância marcada como deferido.'

        performer_name = (item.performed_by_user.name if item.performed_by_user else '') or 'Usuário não identificado'
        events.append(
            {
                'event_type': 'manual_history',
                'history_id': item.id,
                'report_id': None,
                'knowledge_base_id': None,
                'action': item.action,
                'manual_action_label': manual_action_label,
                'manual_description': manual_description,
                'performed_by': performer_name,
                'old_first_instance_status': item.old_first_instance_status,
                'new_first_instance_status': item.new_first_instance_status,
                'old_first_instance_status_label': _status_label_pt(item.old_first_instance_status),
                'new_first_instance_status_label': _status_label_pt(item.new_first_instance_status),
                'notes': item.notes,
                'vigencia_id': item.vigencia_id,
                'transmission_datetime': None,
                'publication_datetime': None,
                'created_at': _format_datetime(item.created_at),
                'report_uploaded_at': None,
                'report_filename': None,
                'knowledge_details_url': None,
                'sort_datetime': item.created_at,
            }
        )

    events.sort(key=lambda item: item.get('sort_datetime') or datetime.min, reverse=True)
    for item in events:
        item.pop('sort_datetime', None)

    return jsonify(
        {
            'benefit_id': benefit.id,
            'benefit_number': benefit.benefit_number,
            'events': events,
        }
    )


@disputes_center_bp.route('/fap-contestation-reports/<int:report_id>/view', methods=['GET'])
@require_law_firm
def view_fap_contestation_report(report_id):
    law_firm_id = get_current_law_firm_id()
    report = FapContestationJudgmentReport.query.filter_by(
        id=report_id,
        law_firm_id=law_firm_id,
    ).first_or_404()

    resolved_path = _resolve_existing_file_path(report.file_path)
    if not resolved_path:
        flash('Arquivo não encontrado no servidor.', 'error')
        return redirect(url_for('disputes_center.fap_contestation_reports'))

    mimetype = 'application/pdf'
    ext = (report.file_type or '').strip().lower()
    if ext in {'doc', 'docx'}:
        mimetype = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    elif ext in {'xls', 'xlsx'}:
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif ext == 'txt':
        mimetype = 'text/plain'

    return send_file(resolved_path, as_attachment=False, mimetype=mimetype)


@disputes_center_bp.route('/')
@require_law_firm
def list_disputes_center():
    law_firm_id = get_current_law_firm_id()
    current_vigencia_filter = None
    total_count = Benefit.query.filter_by(law_firm_id=law_firm_id).count()
    approved_count = Benefit.query.filter_by(law_firm_id=law_firm_id, status='approved').count()
    in_review_count = Benefit.query.filter(
        Benefit.law_firm_id == law_firm_id,
        func.lower(cast(Benefit.status, String)).in_(['in_review', 'analyzing']),
    ).count()
    rejected_count = Benefit.query.filter_by(law_firm_id=law_firm_id, status='rejected').count()
    pending_count = max(total_count - approved_count - in_review_count - rejected_count, 0)
    categorized_count = Benefit.query.filter(
        Benefit.law_firm_id == law_firm_id,
        Benefit.fap_contestation_topic.isnot(None),
        func.trim(cast(Benefit.fap_contestation_topic, String)) != '',
    ).count()
    uncategorized_count = max(total_count - categorized_count, 0)

    general_query = _base_benefits_query(law_firm_id)
    first_instance_status_counts = _group_count_by_status(general_query, Benefit.first_instance_status)
    first_instance_stats = _build_instance_stats(total_count, first_instance_status_counts)
    first_instance_deferred_count = int(first_instance_status_counts.get('deferido', 0) or 0)
    first_instance_analyzing_count = int(first_instance_status_counts.get('analyzing', 0) or 0)
    second_instance_total_base = max(int(total_count) - first_instance_deferred_count - first_instance_analyzing_count, 0)
    second_instance_eligible_query = general_query.filter(
        ~func.lower(func.coalesce(cast(Benefit.first_instance_status, String), '')).in_(['deferido', 'analyzing'])
    )
    second_instance_stats = _build_instance_stats(
        second_instance_total_base,
        _group_count_by_status(second_instance_eligible_query, Benefit.second_instance_status),
    )

    clients_data = (
        Client.query.with_entities(Client.cnpj, Client.name)
        .filter_by(law_firm_id=law_firm_id)
        .all()
    )

    roots_map = {}
    for cnpj, name in clients_data:
        root = _extract_cnpj_root(cnpj)
        if not root:
            continue

        branch = _extract_cnpj_branch(cnpj)
        clean_name = (name or '').strip()

        if root not in roots_map:
            roots_map[root] = {'root': root, 'company_name': '', 'is_main': False}

        current = roots_map[root]
        if branch == '0001' and clean_name:
            current['company_name'] = clean_name
            current['is_main'] = True
            continue

        if not current['is_main'] and clean_name and not current['company_name']:
            current['company_name'] = clean_name

    cnpj_roots = [
        {'root': item['root'], 'company_name': item['company_name']}
        for _, item in sorted(roots_map.items(), key=lambda entry: entry[0])
    ]

    cnpj_by_root = {}
    for cnpj, name in clients_data:
        root = _extract_cnpj_root(cnpj)
        digits = _normalize_cnpj_digits(cnpj)
        if not root or len(digits) < 14:
            continue
        clean_name = (name or '').strip()
        formatted = _format_cnpj(cnpj)
        if root not in cnpj_by_root:
            cnpj_by_root[root] = []
        if not any(item['digits'] == digits for item in cnpj_by_root[root]):
            cnpj_by_root[root].append({'cnpj': formatted, 'digits': digits, 'company_name': clean_name})
    for root_key in cnpj_by_root:
        cnpj_by_root[root_key].sort(key=lambda x: x['digits'])
        for item in cnpj_by_root[root_key]:
            del item['digits']

    client_options = sorted({(name or '').strip() for _, name in clients_data if (name or '').strip()})

    current_vigencia_id = _normalize_text(request.args.get('vigencia_id', ''))
    if current_vigencia_id:
        try:
            vigencia = FapVigenciaCnpj.query.filter_by(
                id=int(current_vigencia_id),
                law_firm_id=law_firm_id,
            ).first()
        except (TypeError, ValueError):
            vigencia = None

        if vigencia is not None:
            clients = Client.query.filter_by(law_firm_id=law_firm_id).all()
            clients_by_exact, clients_by_root = _build_client_cnpj_lookup(clients)
            resolved_client = _resolve_client_for_vigencia(
                vigencia.employer_cnpj,
                clients_by_exact,
                clients_by_root,
            )
            company_name = ''
            if resolved_client is not None:
                company_name = (resolved_client.name or '').strip()
            if not company_name:
                company_name = (
                    db.session.query(Benefit.employer_name)
                    .filter(
                        Benefit.law_firm_id == law_firm_id,
                        Benefit.fap_vigencia_cnpj_id == vigencia.id,
                        Benefit.employer_name.is_not(None),
                        func.trim(Benefit.employer_name) != '',
                    )
                    .order_by(Benefit.updated_at.desc(), Benefit.id.desc())
                    .scalar()
                    or ''
                ).strip()

            current_vigencia_filter = {
                'id': vigencia.id,
                'year': (vigencia.vigencia_year or '').strip(),
                'company_name': company_name,
                'company_cnpj': _format_cnpj(vigencia.employer_cnpj),
            }

    return render_template(
        'disputes_center/list.html',
        total_count=total_count,
        approved_count=approved_count,
        in_review_count=in_review_count,
        rejected_count=rejected_count,
        pending_count=pending_count,
        categorized_count=categorized_count,
        uncategorized_count=uncategorized_count,
        first_instance_stats=first_instance_stats,
        second_instance_stats=second_instance_stats,
        cnpj_roots=cnpj_roots,
        cnpj_by_root=cnpj_by_root,
        client_options=client_options,
        current_vigencia_filter=current_vigencia_filter,
    )


@disputes_center_bp.route('/vigencias', methods=['GET'])
@require_law_firm
def list_fap_vigencias():
    law_firm_id = get_current_law_firm_id()
    first_instance_key = func.lower(func.coalesce(cast(Benefit.first_instance_status, String), ''))
    second_instance_key = func.lower(func.coalesce(cast(Benefit.second_instance_status, String), ''))
    selected_client_id_raw = _normalize_text(request.args.get('client_id', ''))
    selected_client_cnpj = _normalize_cnpj_digits(request.args.get('client_cnpj', ''))[:14]
    selected_client_root = _normalize_cnpj_digits(request.args.get('client_root', ''))[:8]
    selected_show_deferivel = request.args.get('show_deferivel', '') == '1'

    try:
        selected_client_id = int(selected_client_id_raw) if selected_client_id_raw else None
    except ValueError:
        selected_client_id = None
        selected_client_id_raw = ''

    clients = (
        Client.query.filter_by(law_firm_id=law_firm_id)
        .order_by(Client.name.asc())
        .all()
    )
    clients_by_exact, clients_by_root = _build_client_cnpj_lookup(clients)
    client_filter_options = [
        {
            'id': client.id,
            'name': (client.name or '').strip() or f'Cliente #{client.id}',
            'cnpj': _format_cnpj(client.cnpj),
            'root': _extract_cnpj_root(client.cnpj),
        }
        for client in clients
    ]

    vigencia_rows = (
        db.session.query(
            FapVigenciaCnpj,
            func.count(Benefit.id).label('benefits_count'),
            func.max(Benefit.updated_at).label('last_benefit_update'),
            func.sum(
                case((and_(Benefit.id.is_not(None), first_instance_key == 'deferido'), 1), else_=0)
            ).label('first_approved_count'),
            func.sum(
                case((and_(Benefit.id.is_not(None), first_instance_key == 'indeferido'), 1), else_=0)
            ).label('first_rejected_count'),
            func.sum(
                case((and_(Benefit.id.is_not(None), first_instance_key == 'analyzing'), 1), else_=0)
            ).label('first_in_review_count'),
            func.sum(
                case(
                    (
                        and_(
                            Benefit.id.is_not(None),
                            ~first_instance_key.in_(['deferido', 'indeferido', 'analyzing']),
                        ),
                        1,
                    ),
                    else_=0,
                )
            ).label('first_pending_count'),
            func.sum(
                case((and_(Benefit.id.is_not(None), second_instance_key == 'deferido'), 1), else_=0)
            ).label('second_approved_count'),
            func.sum(
                case((and_(Benefit.id.is_not(None), second_instance_key == 'indeferido'), 1), else_=0)
            ).label('second_rejected_count'),
            func.sum(
                case((and_(Benefit.id.is_not(None), second_instance_key == 'analyzing'), 1), else_=0)
            ).label('second_in_review_count'),
            func.sum(
                case(
                    (
                        and_(
                            Benefit.id.is_not(None),
                            ~first_instance_key.in_(['deferido', 'analyzing']),
                            ~second_instance_key.in_(['deferido', 'indeferido', 'analyzing']),
                        ),
                        1,
                    ),
                    else_=0,
                )
            ).label('second_pending_count'),
            func.sum(
                case(
                    (
                        and_(
                            Benefit.id.is_not(None),
                            second_instance_key.in_(['deferido', 'indeferido', 'analyzing']),
                        ),
                        1,
                    ),
                    else_=0,
                )
            ).label('second_instance_activity_count'),
            func.sum(
                case(
                    (
                        and_(
                            Benefit.id.is_not(None),
                            ~first_instance_key.in_(['deferido', 'indeferido']),
                        ),
                        1,
                    ),
                    else_=0,
                )
            ).label('first_instance_eligible_count'),
        )
        .outerjoin(Benefit, Benefit.fap_vigencia_cnpj_id == FapVigenciaCnpj.id)
        .filter(FapVigenciaCnpj.law_firm_id == law_firm_id)
        .group_by(FapVigenciaCnpj.id)
        .order_by(FapVigenciaCnpj.vigencia_year.desc(), FapVigenciaCnpj.employer_cnpj.asc())
        .all()
    )
    root_filter_options = _build_unique_root_options(clients, vigencia_rows)

    # CAT stats grouped by normalized employer_cnpj (14 digits)
    _cat_agg_rows = (
        db.session.query(
            FapContestationCat.employer_cnpj,
            func.count(FapContestationCat.id).label('total'),
            func.sum(case((func.lower(func.coalesce(cast(FapContestationCat.first_instance_status, String), '')) == 'deferido', 1), else_=0)).label('f_appr'),
            func.sum(case((func.lower(func.coalesce(cast(FapContestationCat.first_instance_status, String), '')) == 'indeferido', 1), else_=0)).label('f_rej'),
            func.sum(case((func.lower(func.coalesce(cast(FapContestationCat.first_instance_status, String), '')) == 'analyzing', 1), else_=0)).label('f_rev'),
            func.sum(case((func.lower(func.coalesce(cast(FapContestationCat.second_instance_status, String), '')) == 'deferido', 1), else_=0)).label('s_appr'),
            func.sum(case((func.lower(func.coalesce(cast(FapContestationCat.second_instance_status, String), '')) == 'indeferido', 1), else_=0)).label('s_rej'),
            func.sum(case((func.lower(func.coalesce(cast(FapContestationCat.second_instance_status, String), '')) == 'analyzing', 1), else_=0)).label('s_rev'),
        )
        .filter(
            FapContestationCat.law_firm_id == law_firm_id,
            FapContestationCat.employer_cnpj.isnot(None),
            func.trim(cast(FapContestationCat.employer_cnpj, String)) != '',
        )
        .group_by(FapContestationCat.employer_cnpj)
        .all()
    )
    cat_stats_by_cnpj = {}
    for _cr in _cat_agg_rows:
        _digits = _normalize_cnpj_digits(_cr.employer_cnpj)
        if not _digits:
            continue
        _total = int(_cr.total or 0)
        _fa = int(_cr.f_appr or 0)
        _fr = int(_cr.f_rej or 0)
        _fv = int(_cr.f_rev or 0)
        _sa = int(_cr.s_appr or 0)
        _sr = int(_cr.s_rej or 0)
        _sv = int(_cr.s_rev or 0)
        if _digits not in cat_stats_by_cnpj:
            cat_stats_by_cnpj[_digits] = {'total': 0, 'fa': 0, 'fr': 0, 'fv': 0, 'fp': 0, 'sa': 0, 'sr': 0, 'sv': 0, 'sp': 0}
        _e = cat_stats_by_cnpj[_digits]
        _e['total'] += _total
        _e['fa'] += _fa
        _e['fr'] += _fr
        _e['fv'] += _fv
        _e['fp'] += max(_total - _fa - _fr - _fv, 0)
        _e['sa'] += _sa
        _e['sr'] += _sr
        _e['sv'] += _sv
        _e['sp'] += max(_total - _sa - _sr - _sv, 0)

    # Payroll-mass stats grouped by normalized employer_cnpj (14 digits)
    _pm_agg_rows = (
        db.session.query(
            FapContestationPayrollMass.employer_cnpj,
            func.count(FapContestationPayrollMass.id).label('total'),
        )
        .filter(
            FapContestationPayrollMass.law_firm_id == law_firm_id,
            FapContestationPayrollMass.employer_cnpj.isnot(None),
            func.trim(cast(FapContestationPayrollMass.employer_cnpj, String)) != '',
        )
        .group_by(FapContestationPayrollMass.employer_cnpj)
        .all()
    )
    pm_stats_by_cnpj: dict[str, int] = {}
    for _pr in _pm_agg_rows:
        _digits = _normalize_cnpj_digits(_pr.employer_cnpj)
        if not _digits:
            continue
        pm_stats_by_cnpj[_digits] = pm_stats_by_cnpj.get(_digits, 0) + int(_pr.total or 0)

    # Employment-link stats grouped by normalized employer_cnpj (14 digits)
    _el_agg_rows = (
        db.session.query(
            FapContestationEmploymentLink.employer_cnpj,
            func.count(FapContestationEmploymentLink.id).label('total'),
        )
        .filter(
            FapContestationEmploymentLink.law_firm_id == law_firm_id,
            FapContestationEmploymentLink.employer_cnpj.isnot(None),
            func.trim(cast(FapContestationEmploymentLink.employer_cnpj, String)) != '',
        )
        .group_by(FapContestationEmploymentLink.employer_cnpj)
        .all()
    )
    el_stats_by_cnpj: dict[str, int] = {}
    for _er in _el_agg_rows:
        _digits = _normalize_cnpj_digits(_er.employer_cnpj)
        if not _digits:
            continue
        el_stats_by_cnpj[_digits] = el_stats_by_cnpj.get(_digits, 0) + int(_er.total or 0)

    # Turnover-rate stats grouped by normalized employer_cnpj (14 digits)
    _tr_agg_rows = (
        db.session.query(
            FapContestationTurnoverRate.employer_cnpj,
            func.count(FapContestationTurnoverRate.id).label('total'),
        )
        .filter(
            FapContestationTurnoverRate.law_firm_id == law_firm_id,
            FapContestationTurnoverRate.employer_cnpj.isnot(None),
            func.trim(cast(FapContestationTurnoverRate.employer_cnpj, String)) != '',
        )
        .group_by(FapContestationTurnoverRate.employer_cnpj)
        .all()
    )
    tr_stats_by_cnpj: dict[str, int] = {}
    for _tr in _tr_agg_rows:
        _digits = _normalize_cnpj_digits(_tr.employer_cnpj)
        if not _digits:
            continue
        tr_stats_by_cnpj[_digits] = tr_stats_by_cnpj.get(_digits, 0) + int(_tr.total or 0)

    grouped_clients = {}
    total_benefits_linked = 0
    total_filtered_vigencias = 0

    for (
        vigencia,
        benefits_count,
        last_benefit_update,
        first_approved_count,
        first_rejected_count,
        first_in_review_count,
        first_pending_count,
        second_approved_count,
        second_rejected_count,
        second_in_review_count,
        second_pending_count,
        second_instance_activity_count,
        first_instance_eligible_count,
    ) in vigencia_rows:
        resolved_client = _resolve_client_for_vigencia(
            vigencia.employer_cnpj,
            clients_by_exact,
            clients_by_root,
        )

        if not _matches_vigencia_filters(
            vigencia,
            resolved_client,
            client_id=selected_client_id,
            client_cnpj=selected_client_cnpj,
            client_root=selected_client_root,
        ):
            continue

        if selected_show_deferivel:
            can_mark = (
                int(second_instance_activity_count or 0) > 0
                and int(first_instance_eligible_count or 0) > 0
            )
            if not can_mark:
                continue

        total_filtered_vigencias += 1

        if resolved_client is not None:
            group_key = f'client:{resolved_client.id}'
            client_name = resolved_client.name or 'Cliente sem nome'
            client_cnpj = resolved_client.cnpj or _format_cnpj(vigencia.employer_cnpj)
        else:
            group_key = f'unlinked:{_normalize_cnpj_digits(vigencia.employer_cnpj) or vigencia.id}'
            client_name = 'Sem cliente vinculado'
            client_cnpj = _format_cnpj(vigencia.employer_cnpj)

        if group_key not in grouped_clients:
            grouped_clients[group_key] = {
                'client_id': resolved_client.id if resolved_client is not None else None,
                'client_name': client_name,
                'client_cnpj': client_cnpj,
                'client_establishment_type': _get_cnpj_establishment_type(
                    resolved_client.cnpj if resolved_client is not None else vigencia.employer_cnpj
                ),
                'is_unlinked': resolved_client is None,
                'total_vigencias': 0,
                'total_benefits': 0,
                'total_cats': 0,
                'total_payroll_masses': 0,
                'total_employment_links': 0,
                'total_turnover_rates': 0,
                'vigencias': [],
            }

        benefits_count = int(benefits_count or 0)
        first_approved_count = int(first_approved_count or 0)
        first_rejected_count = int(first_rejected_count or 0)
        first_in_review_count = int(first_in_review_count or 0)
        first_pending_count = int(first_pending_count or 0)
        second_approved_count = int(second_approved_count or 0)
        second_rejected_count = int(second_rejected_count or 0)
        second_in_review_count = int(second_in_review_count or 0)
        second_pending_count = int(second_pending_count or 0)
        second_instance_activity_count = int(second_instance_activity_count or 0)
        first_instance_eligible_count = int(first_instance_eligible_count or 0)
        _vdig = _normalize_cnpj_digits(vigencia.employer_cnpj)
        _cs = cat_stats_by_cnpj.get(_vdig, {})
        cats_count = _cs.get('total', 0)
        pm_count = pm_stats_by_cnpj.get(_vdig, 0)
        el_count = el_stats_by_cnpj.get(_vdig, 0)
        tr_count = tr_stats_by_cnpj.get(_vdig, 0)
        grouped_clients[group_key]['total_vigencias'] += 1
        grouped_clients[group_key]['total_benefits'] += benefits_count
        grouped_clients[group_key]['total_cats'] += cats_count
        grouped_clients[group_key]['total_payroll_masses'] += pm_count
        grouped_clients[group_key]['total_employment_links'] += el_count
        grouped_clients[group_key]['total_turnover_rates'] += tr_count
        grouped_clients[group_key]['vigencias'].append(
            {
                'id': vigencia.id,
                'vigencia_year': vigencia.vigencia_year,
                'employer_cnpj': _format_cnpj(vigencia.employer_cnpj),
                'benefits_count': benefits_count,
                'cats_count': cats_count,
                'payroll_masses_count': pm_count,
                'employment_links_count': el_count,
                'turnover_rates_count': tr_count,
                'first_approved_count': first_approved_count + _cs.get('fa', 0),
                'first_rejected_count': first_rejected_count + _cs.get('fr', 0),
                'first_in_review_count': first_in_review_count + _cs.get('fv', 0),
                'first_pending_count': first_pending_count + _cs.get('fp', 0),
                'second_approved_count': second_approved_count + _cs.get('sa', 0),
                'second_rejected_count': second_rejected_count + _cs.get('sr', 0),
                'second_in_review_count': second_in_review_count + _cs.get('sv', 0),
                'second_pending_count': second_pending_count + _cs.get('sp', 0),
                'can_mark_first_instance_deferred': (
                    second_instance_activity_count > 0 and first_instance_eligible_count > 0
                ),
                'first_instance_eligible_count': first_instance_eligible_count,
                'benefits_view_url': url_for('disputes_center.list_disputes_center', vigencia_id=vigencia.id),
                'cats_view_url': url_for('disputes_center.list_cats', vigencia_id=vigencia.id),
                'payroll_masses_view_url': url_for('disputes_center.list_payroll_masses', vigencia_id=vigencia.id),
                'employment_links_view_url': url_for('disputes_center.list_employment_links', vigencia_id=vigencia.id),
                'turnover_rates_view_url': url_for('disputes_center.list_turnover_rates', vigencia_id=vigencia.id),
            }
        )
        total_benefits_linked += benefits_count

    grouped_client_list = sorted(
        grouped_clients.values(),
        key=lambda item: (
            item['is_unlinked'],
            (item['client_name'] or '').lower(),
            item['client_cnpj'] or '',
        ),
    )

    for group in grouped_client_list:
        group['vigencias'].sort(
            key=lambda item: (
                str(item['vigencia_year'] or ''),
                item['employer_cnpj'] or '',
            ),
            reverse=True,
        )

    markable_vigencia_ids = []
    markable_benefits_count = 0
    for group in grouped_client_list:
        for vigencia in group['vigencias']:
            if not vigencia['can_mark_first_instance_deferred']:
                continue
            markable_vigencia_ids.append(vigencia['id'])
            markable_benefits_count += int(vigencia['first_instance_eligible_count'] or 0)

    linked_clients_count = sum(1 for item in grouped_client_list if not item['is_unlinked'])
    unlinked_groups_count = sum(1 for item in grouped_client_list if item['is_unlinked'])
    active_filter_count = sum(
        1 for value in (selected_client_id_raw, selected_client_cnpj, selected_client_root) if value
    ) + (1 if selected_show_deferivel else 0)

    return render_template(
        'disputes_center/vigencias.html',
        grouped_clients=grouped_client_list,
        total_vigencias=total_filtered_vigencias,
        total_benefits_linked=total_benefits_linked,
        linked_clients_count=linked_clients_count,
        unlinked_groups_count=unlinked_groups_count,
        client_filter_options=client_filter_options,
        root_filter_options=root_filter_options,
        selected_client_id=selected_client_id_raw,
        selected_client_cnpj=_format_cnpj(selected_client_cnpj) if selected_client_cnpj else '',
        selected_client_root=selected_client_root,
        selected_show_deferivel=selected_show_deferivel,
        active_filter_count=active_filter_count,
        markable_vigencia_ids=markable_vigencia_ids,
        markable_benefits_count=markable_benefits_count,
    )


@disputes_center_bp.route('/vigencias/<int:vigencia_id>/mark-first-instance-deferred', methods=['POST'])
@require_law_firm
def mark_vigencia_first_instance_deferred(vigencia_id):
    law_firm_id = get_current_law_firm_id()
    user_id = session.get('user_id')

    vigencia = FapVigenciaCnpj.query.filter_by(
        id=vigencia_id,
        law_firm_id=law_firm_id,
    ).first_or_404()

    result = _mark_first_instance_deferred_for_vigencia(
        law_firm_id=law_firm_id,
        user_id=user_id,
        vigencia=vigencia,
        note='Ação em lote aplicada na tela de vigências.',
    )

    if result['status'] == 'missing_second_instance_activity':
        flash(
            'A ação em lote só pode ser aplicada quando houver decisão em 2ª instância ou benefício em análise.',
            'warning',
        )
        return redirect(url_for('disputes_center.list_fap_vigencias'))

    if result['status'] == 'no_eligible_benefits':
        flash('Não há benefícios elegíveis para marcar como deferido na 1ª instância.', 'info')
        return redirect(url_for('disputes_center.list_fap_vigencias'))

    try:
        db.session.commit()
        flash(
            f'{result["updated_count"]} benefício(s) da vigência {vigencia.vigencia_year or "-"} marcado(s) como deferido na 1ª instância.',
            'success',
        )
    except Exception as exc:
        db.session.rollback()
        flash(f'Erro ao aplicar atualização em lote: {str(exc)}', 'danger')

    return redirect(url_for('disputes_center.list_fap_vigencias'))


@disputes_center_bp.route('/vigencias/mark-first-instance-deferred-batch', methods=['POST'])
@require_law_firm
def mark_vigencias_first_instance_deferred_batch():
    law_firm_id = get_current_law_firm_id()
    user_id = session.get('user_id')
    vigencia_ids = list(dict.fromkeys(_parse_int_list(request.form.getlist('vigencia_ids'))))

    selected_client_id = _normalize_text(request.form.get('client_id', ''))
    selected_client_cnpj = _normalize_text(request.form.get('client_cnpj', ''))
    selected_client_root = _normalize_text(request.form.get('client_root', ''))
    selected_show_deferivel = request.form.get('show_deferivel', '') == '1'

    redirect_kwargs = {
        'client_id': selected_client_id or None,
        'client_cnpj': selected_client_cnpj or None,
        'client_root': selected_client_root or None,
        'show_deferivel': '1' if selected_show_deferivel else None,
    }

    if not vigencia_ids:
        flash('Nenhuma vigência elegível foi enviada para atualização em massa.', 'info')
        return redirect(url_for('disputes_center.list_fap_vigencias', **redirect_kwargs))

    vigencias = FapVigenciaCnpj.query.filter(
        FapVigenciaCnpj.law_firm_id == law_firm_id,
        FapVigenciaCnpj.id.in_(vigencia_ids),
    ).all()
    vigencias_by_id = {vigencia.id: vigencia for vigencia in vigencias}

    updated_vigencias = 0
    updated_benefits = 0
    skipped_no_activity = 0
    skipped_no_eligible = 0

    try:
        for vigencia_id in vigencia_ids:
            vigencia = vigencias_by_id.get(vigencia_id)
            if vigencia is None:
                continue

            result = _mark_first_instance_deferred_for_vigencia(
                law_firm_id=law_firm_id,
                user_id=user_id,
                vigencia=vigencia,
                note='Ação em massa aplicada na tela de vigências.',
            )

            if result['status'] == 'updated':
                updated_vigencias += 1
                updated_benefits += int(result['updated_count'] or 0)
                continue

            if result['status'] == 'missing_second_instance_activity':
                skipped_no_activity += 1
                continue

            if result['status'] == 'no_eligible_benefits':
                skipped_no_eligible += 1

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        flash(f'Erro ao aplicar atualização em massa: {str(exc)}', 'danger')
        return redirect(url_for('disputes_center.list_fap_vigencias', **redirect_kwargs))

    if updated_benefits:
        flash(
            f'Ação em massa concluída: {updated_benefits} benefício(s) atualizado(s) em {updated_vigencias} vigência(s).',
            'success',
        )
    else:
        flash('Nenhum benefício foi atualizado na ação em massa.', 'info')

    if skipped_no_activity:
        flash(
            f'{skipped_no_activity} vigência(s) foram ignoradas por não terem atividade na 2ª instância.',
            'warning',
        )

    if skipped_no_eligible:
        flash(
            f'{skipped_no_eligible} vigência(s) foram ignoradas por não terem benefícios elegíveis.',
            'warning',
        )

    return redirect(url_for('disputes_center.list_fap_vigencias', **redirect_kwargs))


@disputes_center_bp.route('/api/list', methods=['GET'])
@require_law_firm
def list_disputes_center_api():
    law_firm_id = get_current_law_firm_id()
    payload = _collect_listing_payload(default_length=25)

    total_query = db.session.query(func.count(Benefit.id)).filter(Benefit.law_firm_id == law_firm_id)
    records_total = total_query.scalar() or 0

    filtered_query = _apply_benefits_filters(
        _base_benefits_query(law_firm_id),
        search_value=payload['search'],
        custom_filters=payload['filters'],
        quick_client=payload['quick_client'],
        quick_root=payload['quick_root'],
        quick_cnpj=payload.get('quick_cnpj', ''),
        vigencia_id=payload.get('vigencia_id'),
    )

    records_filtered = filtered_query.with_entities(func.count(Benefit.id)).scalar() or 0

    status_counts = _group_count_by_status(filtered_query, Benefit.status)
    approved_filtered = int(status_counts.get('approved', 0) or 0)
    in_review_filtered = int(status_counts.get('in_review', 0) or 0) + int(status_counts.get('analyzing', 0) or 0)
    rejected_filtered = int(status_counts.get('rejected', 0) or 0)
    pending_filtered = max(int(records_filtered) - approved_filtered - in_review_filtered - rejected_filtered, 0)
    categorized_filtered = (
        filtered_query.filter(
            Benefit.fap_contestation_topic.isnot(None),
            func.trim(cast(Benefit.fap_contestation_topic, String)) != '',
        )
        .with_entities(func.count(Benefit.id))
        .scalar()
        or 0
    )
    uncategorized_filtered = max(int(records_filtered) - int(categorized_filtered), 0)
    filtered_first_instance_status_counts = _group_count_by_status(filtered_query, Benefit.first_instance_status)
    filtered_first_instance_stats = _build_instance_stats(
        records_filtered,
        filtered_first_instance_status_counts,
    )
    filtered_first_instance_deferred_count = int(filtered_first_instance_status_counts.get('deferido', 0) or 0)
    filtered_first_instance_analyzing_count = int(filtered_first_instance_status_counts.get('analyzing', 0) or 0)
    filtered_second_instance_total_base = max(int(records_filtered) - filtered_first_instance_deferred_count - filtered_first_instance_analyzing_count, 0)
    filtered_second_instance_eligible_query = filtered_query.filter(
        ~func.lower(func.coalesce(cast(Benefit.first_instance_status, String), '')).in_(['deferido', 'analyzing'])
    )
    filtered_second_instance_stats = _build_instance_stats(
        filtered_second_instance_total_base,
        _group_count_by_status(filtered_second_instance_eligible_query, Benefit.second_instance_status),
    )

    order_column = ORDER_COLUMN_MAP.get(payload['order_column'], Benefit.id)
    if payload['order_dir'] == 'asc':
        filtered_query = filtered_query.order_by(order_column.asc(), Benefit.id.asc())
    else:
        filtered_query = filtered_query.order_by(order_column.desc(), Benefit.id.desc())

    paged_results = (
        filtered_query
        .offset(payload['start'])
        .limit(payload['length'])
        .all()
    )

    data = [_serialize_benefit_row(benefit, client_name) for benefit, client_name in paged_results]

    return jsonify(
        {
            'draw': payload['draw'],
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'filtered_stats': {
                'total': int(records_filtered),
                'approved': approved_filtered,
                'rejected': rejected_filtered,
                'in_review': in_review_filtered,
                'pending': pending_filtered,
                'categorized': int(categorized_filtered),
                'uncategorized': uncategorized_filtered,
                'first_instance': filtered_first_instance_stats,
                'second_instance': filtered_second_instance_stats,
            },
            'data': data,
        }
    )


@disputes_center_bp.route('/export-excel', methods=['POST'])
@require_law_firm
def export_disputes_center_excel():
    law_firm_id = get_current_law_firm_id()
    payload = _collect_listing_payload(default_length=1000)

    filtered_query = _apply_benefits_filters(
        _base_benefits_query(law_firm_id),
        search_value=payload['search'],
        custom_filters=payload['filters'],
        quick_client=payload['quick_client'],
        quick_root=payload['quick_root'],
        quick_cnpj=payload.get('quick_cnpj', ''),
        vigencia_id=payload.get('vigencia_id'),
    )

    order_column = ORDER_COLUMN_MAP.get(payload['order_column'], Benefit.id)
    if payload['order_dir'] == 'asc':
        filtered_query = filtered_query.order_by(order_column.asc(), Benefit.id.asc())
    else:
        filtered_query = filtered_query.order_by(order_column.desc(), Benefit.id.desc())

    benefits = filtered_query.all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Beneficios'

    headers = [
        'ID',
        'Número do benefício',
        'Cliente',
        'Segurado',
        'Tipo benefício',
        'CPF segurado',
        'NIT segurado',
        'Data nascimento segurado',
        'CNPJ empregador',
        'Nome empregador',
        'Status geral',
        'Status 1ª instância',
        'Texto status 1ª instância',
        'Justificativa 1ª instância',
        'Parecer 1ª instância',
        'Status 2ª instância',
        'Texto status 2ª instância',
        'Justificativa 2ª instância',
        'Parecer 2ª instância',
        'Vigência FAP',
        'Categoria FAP',
        'DIB',
        'DCB',
        'Data acidente',
        'Empresa acidente',
        'CAT',
        'BO',
        'Tipo solicitação',
        'RMI',
        'Total pago',
        'Justificativa geral',
        'Parecer geral',
        'Observações',
    ]
    sheet.append(headers)

    for benefit, client_name in benefits:
        general_status_value = _resolve_general_status_excel_value(
            benefit.first_instance_status,
            benefit.second_instance_status,
            benefit.first_instance_status_raw,
            benefit.second_instance_status_raw,
            benefit.status,
        )

        sheet.append(
            [
                benefit.id,
                benefit.benefit_number or '',
                client_name or '',
                benefit.insured_name or '',
                benefit.benefit_type or '',
                benefit.insured_cpf or '',
                benefit.insured_nit or '',
                _format_date(benefit.insured_date_of_birth),
                benefit.employer_cnpj or '',
                benefit.employer_name or '',
                general_status_value,
                _status_label_pt(benefit.first_instance_status),
                benefit.first_instance_status_raw or '',
                benefit.first_instance_justification or '',
                benefit.first_instance_opinion or '',
                _status_label_pt(benefit.second_instance_status),
                benefit.second_instance_status_raw or '',
                benefit.second_instance_justification or '',
                benefit.second_instance_opinion or '',
                benefit.fap_vigencia_years or '',
                benefit.fap_contestation_topic or '',
                _format_date(benefit.benefit_start_date),
                _format_date(benefit.benefit_end_date),
                _format_date(benefit.accident_date),
                benefit.accident_company_name or '',
                benefit.cat_number or '',
                benefit.bo_number or '',
                benefit.request_type or '',
                _format_decimal(benefit.initial_monthly_benefit),
                _format_decimal(benefit.total_paid),
                benefit.justification or '',
                benefit.opinion or '',
                benefit.notes or '',
            ]
        )

    for idx, _ in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = 22

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'disputes_center_{timestamp}.xlsx'

    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@disputes_center_bp.route('/fap-contestation-reports', methods=['GET', 'POST'])
@require_law_firm
def fap_contestation_reports():
    from app.form import FapContestationJudgmentReportForm

    law_firm_id = get_current_law_firm_id()
    current_user_id = session.get('user_id')
    page = request.args.get('page', 1, type=int)
    form = FapContestationJudgmentReportForm()
    allowed_extensions = {'pdf', 'doc', 'docx', 'txt', 'xlsx', 'xls'}
    wants_json = request.method == 'POST' and _normalize_text(request.form.get('response_format')) == 'json'

    if request.method == 'POST':
        if not form.validate_on_submit():
            if wants_json:
                form_errors = []
                for field_errors in form.errors.values():
                    form_errors.extend(field_errors)

                return jsonify(
                    {
                        'ok': False,
                        'message': 'Falha ao validar o envio.',
                        'uploaded_count': 0,
                        'invalid_files': [],
                        'errors': form_errors,
                    }
                ), 400
        else:
            files = request.files.getlist('file') if 'file' in request.files else []
            files = [f for f in files if f and f.filename and f.filename.strip()]

            if not files:
                if wants_json:
                    return jsonify(
                        {
                            'ok': False,
                            'message': 'Selecione ao menos um arquivo para envio.',
                            'uploaded_count': 0,
                            'invalid_files': [],
                            'errors': [],
                        }
                    ), 400
                flash('Selecione ao menos um arquivo para envio.', 'warning')
                return redirect(url_for('disputes_center.fap_contestation_reports'))

            invalid_files = []
            success_count = 0
            upload_errors = []
            upload_dir = os.path.abspath(
                os.path.join(current_app.root_path, '..', 'uploads', 'fap_contestation_reports')
            )
            os.makedirs(upload_dir, exist_ok=True)

            for file in files:
                filename = secure_filename(file.filename)
                extension = os.path.splitext(filename)[1].lower().replace('.', '')
                if extension not in allowed_extensions:
                    invalid_files.append(filename)
                    continue

                try:
                    timestamp = now_sp().strftime('%Y%m%d_%H%M%S_%f')
                    unique_filename = f'{timestamp}_{filename}'
                    file_path = os.path.join(upload_dir, unique_filename)
                    file.save(file_path)

                    file_size = os.path.getsize(file_path)
                    file_type = extension.upper()

                    report = FapContestationJudgmentReport(
                        user_id=current_user_id,
                        law_firm_id=law_firm_id,
                        original_filename=filename,
                        file_path=file_path,
                        file_size=file_size,
                        file_type=file_type,
                        status='pending',
                    )

                    db.session.add(report)
                    db.session.flush()
                    knowledge_file = _ensure_fap_report_in_knowledge_base(
                        law_firm_id=law_firm_id,
                        user_id=session.get('user_id'),
                        filename=filename,
                        file_path=file_path,
                        file_size=file_size,
                        file_type=file_type,
                    )
                    report.knowledge_base_id = knowledge_file.id
                    db.session.commit()
                    success_count += 1
                except Exception as e:
                    db.session.rollback()
                    upload_errors.append(f'Erro ao enviar {filename}: {str(e)}')
                    if wants_json:
                        current_app.logger.exception('Erro ao enviar relatório FAP %s', filename)
                    else:
                        flash(f'Erro ao enviar {filename}: {str(e)}', 'danger')

            if wants_json:
                messages = []
                if success_count:
                    messages.append(f'{success_count} relatório(s) enviado(s) com sucesso.')
                if invalid_files:
                    messages.append(f'{len(invalid_files)} arquivo(s) ignorado(s) por extensão inválida.')
                if upload_errors:
                    messages.append(f'{len(upload_errors)} arquivo(s) falharam no envio.')

                return jsonify(
                    {
                        'ok': success_count > 0 and not upload_errors,
                        'message': ' '.join(messages) if messages else 'Nenhum arquivo foi enviado.',
                        'uploaded_count': success_count,
                        'invalid_files': invalid_files,
                        'errors': upload_errors,
                    }
                )

            if success_count:
                flash(
                    f'{success_count} relatório(s) enviado(s) com sucesso.',
                    'success',
                )

            if invalid_files:
                flash(
                'Arquivos ignorados por extensão inválida: ' + ', '.join(invalid_files),
                'warning'
                )

            return redirect(url_for('disputes_center.fap_contestation_reports'))

    reports = (
        FapContestationJudgmentReport.query.filter_by(law_firm_id=law_firm_id)
        .order_by(
            FapContestationJudgmentReport.uploaded_at.desc(),
            FapContestationJudgmentReport.id.desc(),
        )
        .paginate(page=page, per_page=50, error_out=False)
    )

    # Build a set of report IDs that were auto-imported
    report_ids_on_page = [r.id for r in reports.items]
    auto_imported_report_ids = set()
    if report_ids_on_page:
        auto_imported_report_ids = {
            rec.report_id
            for rec in FapAutoImportedContestacao.query.filter(
                FapAutoImportedContestacao.report_id.in_(report_ids_on_page)
            ).all()
        }

    processed_count = (
        FapContestationJudgmentReport.query.filter_by(
            law_firm_id=law_firm_id,
            status='completed',
        ).count()
    )
    processing_count = (
        FapContestationJudgmentReport.query.filter_by(
            law_firm_id=law_firm_id,
            status='processing',
        ).count()
    )
    not_processed_count = (
        FapContestationJudgmentReport.query.filter(
            FapContestationJudgmentReport.law_firm_id == law_firm_id,
            FapContestationJudgmentReport.status.in_(['pending', 'queued', 'error']),
        ).count()
    )

    return render_template(
        'disputes_center/fap_contestation_reports.html',
        form=form,
        reports=reports,
        processed_count=processed_count,
        not_processed_count=not_processed_count,
        processing_count=processing_count,
        auto_imported_report_ids=auto_imported_report_ids,
    )


@disputes_center_bp.route('/fap-contestation-reports/<int:report_id>/delete', methods=['POST'])
@require_law_firm
def delete_fap_contestation_report(report_id):
    law_firm_id = get_current_law_firm_id()
    report = FapContestationJudgmentReport.query.filter_by(id=report_id, law_firm_id=law_firm_id).first_or_404()

    try:
        if report.file_path and os.path.exists(report.file_path):
            os.remove(report.file_path)
        db.session.delete(report)
        db.session.commit()
        flash('Relatório excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir relatório: {str(e)}', 'danger')

    return redirect(url_for('disputes_center.fap_contestation_reports'))


@disputes_center_bp.route('/fap-auto-import', methods=['GET'])
@require_law_firm
def fap_auto_import():
    from datetime import date
    law_firm_id = get_current_law_firm_id()
    current_year = date.today().year
    years = list(range(current_year, 2009, -1))
    companies = (
        FapCompany.query.filter_by(law_firm_id=law_firm_id)
        .order_by(FapCompany.nome.asc())
        .all()
    )
    saved_auth = session.get('fap_auto_import_auth', '')
    return render_template(
        'disputes_center/fap_auto_import.html',
        years=years,
        companies=companies,
        saved_auth=saved_auth,
    )


@disputes_center_bp.route('/fap-auto-import/save-auth', methods=['POST'])
@require_law_firm
def fap_auto_import_save_auth():
    data = request.get_json(silent=True) or {}
    auth_json = data.get('auth', '').strip()
    if not auth_json:
        return jsonify({'ok': False, 'message': 'Nenhum dado enviado.'}), 400
    session['fap_auto_import_auth'] = auth_json
    session.modified = True
    return jsonify({'ok': True})


@disputes_center_bp.route('/fap-auto-import/check-session', methods=['GET'])
@require_law_firm
def fap_auto_import_check_session():
    saved_auth = session.get('fap_auto_import_auth', '')
    if not saved_auth:
        return jsonify({'ok': False, 'message': 'Sem sessão salva.'}), 400

    try:
        auth = FapWebAuthPayload.from_json(saved_auth)
    except Exception:
        return jsonify({'ok': False, 'message': 'Dados de autenticação inválidos.'}), 400

    result = FapWebService(auth).check_session()
    if result.ok:
        return jsonify({'ok': True, 'status': result.status_code})
    if result.expired:
        return jsonify({'ok': False, 'expired': True, 'status': result.status_code})
    return jsonify({'ok': False, 'status': result.status_code, 'message': result.message}), 502


@disputes_center_bp.route('/fap-auto-import/fetch-reports', methods=['POST'])
@require_law_firm
def fap_auto_import_fetch_reports():
    data = request.get_json(silent=True) or {}
    cnpj_raiz = str(data.get('cnpj') or '').strip()
    year = str(data.get('year') or '').strip()

    if not cnpj_raiz or not year:
        return jsonify({'ok': False, 'message': 'Informe o CNPJ e o ano de vigência.'}), 400

    saved_auth = session.get('fap_auto_import_auth', '')
    if not saved_auth:
        return jsonify({
            'ok': False,
            'message': 'Dados de autenticação não encontrados na sessão. Salve os dados de autenticação antes de buscar.',
        }), 400

    try:
        auth = FapWebAuthPayload.from_json(saved_auth)
    except Exception:
        return jsonify({'ok': False, 'message': 'Dados de autenticação inválidos na sessão. Atualize e salve novamente.'}), 400

    result = FapWebService(auth).fetch_contestacoes(cnpj=cnpj_raiz, year=year)
    if not result.ok:
        detail = (result.data or {}).get('detail', '') if result.data else ''
        payload = {'ok': False, 'message': result.message}
        if detail:
            payload['detail'] = detail
        return jsonify(payload), 502

    items = result.data

    # Anotações de status de importação
    law_firm_id = get_current_law_firm_id()
    item_ids = [item.get('id') for item in items if item.get('id')]
    imported_map = {}
    if item_ids:
        imported_records = FapAutoImportedContestacao.query.filter(
            FapAutoImportedContestacao.law_firm_id == law_firm_id,
            FapAutoImportedContestacao.contestacao_id.in_(item_ids),
        ).all()
        for rec in imported_records:
            imported_map[rec.contestacao_id] = rec.report_id
    for item in items:
        item_id = item.get('id')
        item['_imported_report_id'] = imported_map.get(item_id) if item_id else None

    return jsonify({'ok': True, 'items': items, 'total': len(items)})


@disputes_center_bp.route('/fap-auto-import/import-contestacao', methods=['POST'])
@require_law_firm
def fap_auto_import_import_contestacao():
    law_firm_id = get_current_law_firm_id()
    current_user_id = session.get('user_id')

    data = request.get_json(silent=True) or {}
    year = data.get('year')
    cnpj = str(data.get('cnpj') or '').strip()
    contestacao_id = data.get('contestacao_id')

    if not year or not cnpj or not contestacao_id:
        return jsonify({'ok': False, 'message': 'Parâmetros inválidos.'}), 400

    # Verificar se já foi importada
    existing = FapAutoImportedContestacao.query.filter_by(
        law_firm_id=law_firm_id,
        contestacao_id=int(contestacao_id),
        cnpj=cnpj,
    ).first()
    if existing:
        return jsonify({
            'ok': False,
            'already_imported': True,
            'report_id': existing.report_id,
            'message': 'Esta contestação já foi importada anteriormente.',
        })

    # Tenta usar arquivo local (fap_web_contestacoes) antes de buscar no portal FAP
    pdf_bytes   = None
    filename    = None
    local_path  = None   # abs_path do arquivo já salvo — se preenchido, não duplica no disco
    rec_id = data.get('rec_id')
    if rec_id:
        fap_rec = FapWebContestacao.query.filter_by(id=int(rec_id), law_firm_id=law_firm_id).first()
        if fap_rec and fap_rec.file_path:
            abs_path = os.path.abspath(os.path.join(current_app.root_path, fap_rec.file_path))
            if os.path.isfile(abs_path):
                local_path = abs_path
                base = os.path.basename(abs_path)
                parts = base.split('_', 1)
                filename = parts[1] if len(parts) == 2 and parts[0].isdigit() else base

    if local_path is None:
        saved_auth = session.get('fap_auto_import_auth', '')
        if not saved_auth:
            return jsonify({'ok': False, 'message': 'Dados de autenticação não encontrados na sessão.'}), 400

        try:
            auth = FapWebAuthPayload.from_json(saved_auth)
        except Exception:
            return jsonify({'ok': False, 'message': 'Dados de autenticação inválidos na sessão.'}), 400

        result = FapWebService(auth).download_contestacao(year=year, cnpj=cnpj, contestacao_id=contestacao_id)
        if not result.ok:
            status_code = 401 if result.expired else 502
            return jsonify({'ok': False, 'message': result.message}), status_code

        pdf_bytes = result.data['pdf_bytes']
        filename  = result.data['filename']

    upload_dir = os.path.abspath(
        os.path.join(current_app.root_path, '..', 'uploads', 'fap_contestation_reports')
    )
    os.makedirs(upload_dir, exist_ok=True)

    try:
        from app.utils.timezone import now_sp
        timestamp = now_sp().strftime('%Y%m%d_%H%M%S_%f')
        safe_filename = secure_filename(filename)

        if local_path:
            # Arquivo já existe localmente — aponta direto, sem copiar
            file_path = local_path
            file_size = os.path.getsize(file_path)
        else:
            # Arquivo veio da API — salva em fap_contestation_reports
            unique_filename = f'{timestamp}_{safe_filename}'
            file_path = os.path.join(upload_dir, unique_filename)
            with open(file_path, 'wb') as fh:
                fh.write(pdf_bytes)
            file_size = len(pdf_bytes)
        file_type = 'PDF'

        report = FapContestationJudgmentReport(
            user_id=current_user_id,
            law_firm_id=law_firm_id,
            original_filename=filename,
            file_path=file_path,
            file_size=file_size,
            file_type=file_type,
            status='pending',
        )
        db.session.add(report)
        db.session.flush()

        knowledge_file = _ensure_fap_report_in_knowledge_base(
            law_firm_id=law_firm_id,
            user_id=current_user_id,
            filename=filename,
            file_path=file_path,
            file_size=file_size,
            file_type=file_type,
        )
        report.knowledge_base_id = knowledge_file.id
        db.session.flush()

        imported = FapAutoImportedContestacao(
            law_firm_id=law_firm_id,
            report_id=report.id,
            contestacao_id=int(contestacao_id),
            cnpj=cnpj,
            year=int(year),
            original_filename=filename,
        )
        db.session.add(imported)
        db.session.commit()

        return jsonify({
            'ok': True,
            'report_id': report.id,
            'message': f'Contestação importada com sucesso como "{filename}".',
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Erro ao importar contestação FAP %s', contestacao_id)
        return jsonify({'ok': False, 'message': f'Erro ao salvar importação: {str(e)}'}), 500


@disputes_center_bp.route('/fap-auto-import/download-contestacao/<int:year>/<cnpj>/<int:contestacao_id>', methods=['GET'])
@require_law_firm
def fap_auto_import_download_contestacao(year, cnpj, contestacao_id):
    saved_auth = session.get('fap_auto_import_auth', '')
    if not saved_auth:
        return 'Dados de autenticação não encontrados na sessão.', 400

    try:
        auth = FapWebAuthPayload.from_json(saved_auth)
    except Exception:
        return 'Dados de autenticação inválidos na sessão.', 400

    result = FapWebService(auth).download_contestacao(year=year, cnpj=cnpj, contestacao_id=contestacao_id)
    if not result.ok:
        if result.expired:
            return 'Sessão expirada ou não autorizada. Atualize os dados de autenticação e tente novamente.', 401
        return result.message, result.status_code or 502

    pdf_bytes = result.data['pdf_bytes']
    filename  = result.data['filename']

    inline = request.args.get('inline') == '1'
    disposition = f'inline; filename="{filename}"' if inline else f'attachment; filename="{filename}"'
    from flask import Response
    return Response(
        pdf_bytes,
        status=200,
        headers={
            'Content-Type': 'application/pdf',
            'Content-Disposition': disposition,
        },
    )


@disputes_center_bp.route('/fap-auto-import/fetch-companies', methods=['POST'])
@require_law_firm
def fap_auto_import_fetch_companies():
    law_firm_id = get_current_law_firm_id()
    data = request.get_json(silent=True) or {}
    cookies_dict = data.get('cookies') or {}

    if not cookies_dict or not isinstance(cookies_dict, dict):
        return jsonify({'ok': False, 'message': 'Informe os dados de autenticação no formato JSON com o campo "cookies".'}), 400

    if not any(v for v in cookies_dict.values()):
        return jsonify({'ok': False, 'message': 'O objeto "cookies" está vazio.'}), 400

    auth = FapWebAuthPayload.from_dict(data)
    result = FapWebService(auth).fetch_companies()
    if not result.ok:
        detail = (result.data or {}).get('detail', '') if result.data else ''
        payload = {'ok': False, 'message': result.message}
        if detail:
            payload['detail'] = detail
        return jsonify(payload), 502

    companies = result.data

    # Upsert companies into DB
    try:
        from datetime import datetime as _dt
        now = _dt.utcnow()
        seen_cnpjs = set()
        for item in (companies if isinstance(companies, list) else []):
            cnpj = str(item.get('cnpj') or '').strip()
            if not cnpj:
                continue
            seen_cnpjs.add(cnpj)
            tipo = item.get('tipoProcuracao') or {}
            nome = (item.get('nome') or '').strip()
            rec = FapCompany.query.filter_by(law_firm_id=law_firm_id, cnpj=cnpj).first()
            if rec:
                rec.nome = nome
                rec.tipo_procuracao_codigo = tipo.get('codigo')
                rec.tipo_procuracao_descricao = tipo.get('descricao')
                rec.synced_at = now
            else:
                rec = FapCompany(
                    law_firm_id=law_firm_id,
                    cnpj=cnpj,
                    nome=nome,
                    tipo_procuracao_codigo=tipo.get('codigo'),
                    tipo_procuracao_descricao=tipo.get('descricao'),
                    synced_at=now,
                )
                db.session.add(rec)
        # Remove companies no longer returned by the API
        if seen_cnpjs:
            FapCompany.query.filter(
                FapCompany.law_firm_id == law_firm_id,
                FapCompany.cnpj.notin_(seen_cnpjs),
            ).delete(synchronize_session='fetch')
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Erro ao salvar empresas FAP no banco')
        return jsonify({'ok': False, 'message': f'Empresas recebidas mas falha ao salvar no banco: {str(e)}'}), 500

    return jsonify({'ok': True, 'companies': companies, 'saved_count': len(seen_cnpjs)})


@disputes_center_bp.route('/new', methods=['GET', 'POST'])
@require_law_firm
def new_dispute():
    from app.form import CentralBenefitForm

    law_firm_id = get_current_law_firm_id()
    form = CentralBenefitForm()

    clients = (
        Client.query.filter_by(law_firm_id=law_firm_id)
        .order_by(Client.name.asc())
        .all()
    )
    form.client_id.choices = [('', 'Sem cliente vinculado')] + [(c.id, c.name) for c in clients]

    if form.validate_on_submit():
        benefit = Benefit(
            law_firm_id=law_firm_id,
            client_id=form.client_id.data,
            benefit_number=form.benefit_number.data,
            benefit_type=form.benefit_type.data,
            insured_name=form.insured_name.data,
            insured_nit=form.insured_nit.data,
            insured_cpf=form.insured_cpf.data,
            insured_date_of_birth=form.insured_date_of_birth.data,
            employer_cnpj=form.employer_cnpj.data,
            employer_name=form.employer_name.data,
            benefit_start_date=form.benefit_start_date.data,
            benefit_end_date=form.benefit_end_date.data,
            initial_monthly_benefit=form.initial_monthly_benefit.data,
            total_paid=form.total_paid.data,
            accident_date=form.accident_date.data,
            accident_company_name=form.accident_company_name.data,
            accident_summary=form.accident_summary.data,
            cat_number=form.cat_number.data,
            bo_number=form.bo_number.data,
            fap_vigencia_years=form.fap_vigencia_years.data,
            request_type=form.request_type.data or None,
            status=form.status.data,
            first_instance_status=form.first_instance_status.data or None,
            first_instance_justification=form.first_instance_justification.data,
            first_instance_opinion=form.first_instance_opinion.data,
            second_instance_status=form.second_instance_status.data or None,
            second_instance_justification=form.second_instance_justification.data,
            second_instance_opinion=form.second_instance_opinion.data,
            justification=form.justification.data,
            opinion=form.opinion.data,
            notes=form.notes.data,
        )

        db.session.add(benefit)
        try:
            db.session.commit()
            flash('Registro de disputa cadastrado com sucesso!', 'success')
            return redirect(url_for('disputes_center.list_disputes_center'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar benefício: {str(e)}', 'danger')

    return render_template('disputes_center/form.html', form=form, title='Novo Registro de Disputa')


@disputes_center_bp.route('/<int:benefit_id>/edit', methods=['GET', 'POST'])
@require_law_firm
def edit_dispute(benefit_id):
    from app.form import CentralBenefitForm

    law_firm_id = get_current_law_firm_id()
    benefit = Benefit.query.filter_by(id=benefit_id, law_firm_id=law_firm_id).first_or_404()
    form = CentralBenefitForm(obj=benefit)

    clients = (
        Client.query.filter_by(law_firm_id=law_firm_id)
        .order_by(Client.name.asc())
        .all()
    )
    form.client_id.choices = [('', 'Sem cliente vinculado')] + [(c.id, c.name) for c in clients]

    if request.method == 'GET':
        form.client_id.data = benefit.client_id

    if form.validate_on_submit():
        user_id = session.get('user_id')
        old_first_instance_status = _normalize_optional_status(benefit.first_instance_status)
        new_first_instance_status = _normalize_optional_status(form.first_instance_status.data)

        benefit.client_id = form.client_id.data
        benefit.benefit_number = form.benefit_number.data
        benefit.benefit_type = form.benefit_type.data
        benefit.insured_name = form.insured_name.data
        benefit.insured_nit = form.insured_nit.data
        benefit.insured_cpf = form.insured_cpf.data
        benefit.insured_date_of_birth = form.insured_date_of_birth.data
        benefit.employer_cnpj = form.employer_cnpj.data
        benefit.employer_name = form.employer_name.data
        benefit.benefit_start_date = form.benefit_start_date.data
        benefit.benefit_end_date = form.benefit_end_date.data
        benefit.initial_monthly_benefit = form.initial_monthly_benefit.data
        benefit.total_paid = form.total_paid.data
        benefit.accident_date = form.accident_date.data
        benefit.accident_company_name = form.accident_company_name.data
        benefit.accident_summary = form.accident_summary.data
        benefit.cat_number = form.cat_number.data
        benefit.bo_number = form.bo_number.data
        benefit.fap_vigencia_years = form.fap_vigencia_years.data
        benefit.request_type = form.request_type.data or None
        benefit.status = form.status.data
        benefit.first_instance_status = form.first_instance_status.data or None
        benefit.first_instance_justification = form.first_instance_justification.data
        benefit.first_instance_opinion = form.first_instance_opinion.data
        benefit.second_instance_status = form.second_instance_status.data or None
        benefit.second_instance_justification = form.second_instance_justification.data
        benefit.second_instance_opinion = form.second_instance_opinion.data
        benefit.justification = form.justification.data
        benefit.opinion = form.opinion.data
        benefit.notes = form.notes.data
        benefit.updated_at = datetime.utcnow()

        if old_first_instance_status != new_first_instance_status:
            db.session.add(
                BenefitManualHistory(
                    law_firm_id=law_firm_id,
                    benefit_id=benefit.id,
                    vigencia_id=benefit.fap_vigencia_cnpj_id,
                    performed_by_user_id=user_id,
                    action='edit_dispute_first_instance_status',
                    old_first_instance_status=old_first_instance_status,
                    new_first_instance_status=new_first_instance_status or '',
                    notes='Status da 1ª instância alterado na tela de edição.',
                )
            )

        try:
            db.session.commit()
            flash('Registro de disputa atualizado com sucesso!', 'success')
            return redirect(url_for('disputes_center.list_disputes_center'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar benefício: {str(e)}', 'danger')

    import json as _json
    clients_data = {str(c.id): {'name': c.name, 'cnpj': c.cnpj or ''} for c in clients}

    return render_template(
        'disputes_center/form.html',
        form=form,
        title='Editar Registro de Disputa',
        benefit_id=benefit_id,
        clients_data=_json.dumps(clients_data),
    )


@disputes_center_bp.route('/<int:benefit_id>/delete', methods=['POST'])
@require_law_firm
def delete_dispute(benefit_id):
    law_firm_id = get_current_law_firm_id()
    benefit = Benefit.query.filter_by(id=benefit_id, law_firm_id=law_firm_id).first_or_404()

    try:
        db.session.delete(benefit)
        db.session.commit()
        flash('Registro de disputa excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir benefício: {str(e)}', 'danger')

    return redirect(url_for('disputes_center.list_disputes_center'))


# ---------------------------------------------------------------------------
# CATs (Comunicação de Acidente de Trabalho)
# ---------------------------------------------------------------------------

@disputes_center_bp.route('/cats')
@require_law_firm
def list_cats():
    law_firm_id = get_current_law_firm_id()
    base_query = _base_cats_query(law_firm_id)
    total_count = base_query.with_entities(func.count(FapContestationCat.id)).scalar() or 0

    first_status_counts = {
        _normalize_status_key(status): int(count or 0)
        for status, count in base_query.with_entities(
            func.lower(func.coalesce(cast(FapContestationCat.first_instance_status, String), '')),
            func.count(FapContestationCat.id),
        ).group_by(func.lower(func.coalesce(cast(FapContestationCat.first_instance_status, String), ''))).all()
    }
    first_instance_stats = _build_instance_stats(total_count, first_status_counts)

    second_status_counts = {
        _normalize_status_key(status): int(count or 0)
        for status, count in base_query.filter(
            ~func.lower(func.coalesce(cast(FapContestationCat.first_instance_status, String), '')).in_(['deferido', 'analyzing'])
        ).with_entities(
            func.lower(func.coalesce(cast(FapContestationCat.second_instance_status, String), '')),
            func.count(FapContestationCat.id),
        ).group_by(func.lower(func.coalesce(cast(FapContestationCat.second_instance_status, String), ''))).all()
    }
    first_deferred_count = int(first_status_counts.get('deferido', 0))
    first_analyzing_count = int(first_status_counts.get('analyzing', 0))
    second_total_base = max(total_count - first_deferred_count - first_analyzing_count, 0)
    second_instance_stats = _build_instance_stats(second_total_base, second_status_counts)

    # Quick filter data: client names, CNPJ roots and breakdown (from Client model, same as benefits)
    cnpj_entries = (
        Client.query.with_entities(Client.cnpj, Client.name)
        .filter_by(law_firm_id=law_firm_id)
        .all()
    )

    cats_roots_map = {}
    for cnpj, name in cnpj_entries:
        root = _extract_cnpj_root(cnpj)
        if not root:
            continue
        branch = _extract_cnpj_branch(cnpj)
        clean_name = (name or '').strip()
        if root not in cats_roots_map:
            cats_roots_map[root] = {'root': root, 'company_name': '', 'is_main': False}
        current = cats_roots_map[root]
        if branch == '0001' and clean_name:
            current['company_name'] = clean_name
            current['is_main'] = True
            continue
        if not current['is_main'] and clean_name and not current['company_name']:
            current['company_name'] = clean_name

    cnpj_roots = [
        {'root': item['root'], 'company_name': item['company_name']}
        for _, item in sorted(cats_roots_map.items(), key=lambda entry: entry[0])
    ]

    cnpj_by_root = {}
    for cnpj, name in cnpj_entries:
        root = _extract_cnpj_root(cnpj)
        digits = _normalize_cnpj_digits(cnpj)
        if not root or len(digits) < 14:
            continue
        clean_name = (name or '').strip()
        formatted = _format_cnpj(cnpj)
        if root not in cnpj_by_root:
            cnpj_by_root[root] = []
        if not any(item['digits'] == digits for item in cnpj_by_root[root]):
            cnpj_by_root[root].append({'cnpj': formatted, 'digits': digits, 'company_name': clean_name})
    for root_key in cnpj_by_root:
        cnpj_by_root[root_key].sort(key=lambda x: x['digits'])
        for item in cnpj_by_root[root_key]:
            del item['digits']

    employer_name_options = sorted({(name or '').strip() for _, name in cnpj_entries if (name or '').strip()})

    initial_cnpj = _normalize_cnpj_digits(request.args.get('quick_cnpj', ''))

    # Resolve vigência filter label (from vigencia_id URL param, same pattern as list_disputes_center)
    current_vigencia_filter = None
    current_vigencia_id_raw = _normalize_text(request.args.get('vigencia_id', ''))
    if current_vigencia_id_raw:
        try:
            vigencia_obj = FapVigenciaCnpj.query.filter_by(
                id=int(current_vigencia_id_raw),
                law_firm_id=law_firm_id,
            ).first()
        except (TypeError, ValueError):
            vigencia_obj = None

        if vigencia_obj is not None:
            # Try to resolve company name from CATs of this vigência
            company_name = (
                db.session.query(FapContestationCat.employer_name)
                .filter(
                    FapContestationCat.law_firm_id == law_firm_id,
                    FapContestationCat.vigencia_id == vigencia_obj.id,
                    FapContestationCat.employer_name.is_not(None),
                    func.trim(FapContestationCat.employer_name) != '',
                )
                .order_by(FapContestationCat.updated_at.desc(), FapContestationCat.id.desc())
                .scalar()
                or ''
            ).strip()
            current_vigencia_filter = {
                'id': vigencia_obj.id,
                'year': (vigencia_obj.vigencia_year or '').strip(),
                'company_name': company_name,
                'company_cnpj': _format_cnpj(vigencia_obj.employer_cnpj),
            }

    return render_template(
        'disputes_center/cats.html',
        total_count=total_count,
        first_instance_stats=first_instance_stats,
        second_instance_stats=second_instance_stats,
        cnpj_roots=cnpj_roots,
        cnpj_by_root=cnpj_by_root,
        employer_name_options=employer_name_options,
        initial_cnpj=initial_cnpj,
        current_vigencia_filter=current_vigencia_filter,
    )


@disputes_center_bp.route('/api/cats', methods=['GET'])
@require_law_firm
def list_cats_api():
    law_firm_id = get_current_law_firm_id()
    payload = _collect_listing_payload(default_length=25)
    # Override filter parsing to use CAT fields
    if request.is_json:
        raw = (request.get_json(silent=True) or {}).get('filters')
    else:
        raw = request.args.get('custom_filters', '[]')
    cat_filters = _parse_cat_custom_filters(raw)

    total_count = (
        _base_cats_query(law_firm_id)
        .with_entities(func.count(FapContestationCat.id))
        .scalar() or 0
    )

    filtered_query = _apply_cats_filters(
        _base_cats_query(law_firm_id),
        search_value=payload['search'],
        custom_filters=cat_filters,
        quick_employer_name=payload['quick_client'],
        quick_root=payload['quick_root'],
        quick_cnpj=payload['quick_cnpj'],
        vigencia_id=payload['vigencia_id'],
    )
    records_filtered = filtered_query.with_entities(func.count(FapContestationCat.id)).scalar() or 0

    first_counts = {
        _normalize_status_key(s): int(c or 0)
        for s, c in filtered_query.with_entities(
            func.lower(func.coalesce(cast(FapContestationCat.first_instance_status, String), '')),
            func.count(FapContestationCat.id),
        ).group_by(func.lower(func.coalesce(cast(FapContestationCat.first_instance_status, String), ''))).all()
    }
    approved_filtered = first_counts.get('deferido', 0)
    rejected_filtered = first_counts.get('indeferido', 0)
    in_review_filtered = first_counts.get('analyzing', 0)
    pending_filtered = max(int(records_filtered) - approved_filtered - rejected_filtered - in_review_filtered, 0)

    order_column = CAT_ORDER_COLUMN_MAP.get(payload['order_column'], FapContestationCat.id)
    if payload['order_dir'] == 'asc':
        filtered_query = filtered_query.order_by(order_column.asc(), FapContestationCat.id.asc())
    else:
        filtered_query = filtered_query.order_by(order_column.desc(), FapContestationCat.id.desc())

    paged_results = filtered_query.offset(payload['start']).limit(payload['length']).all()
    data = [_serialize_cat_row(cat) for cat in paged_results]

    return jsonify({
        'draw': payload['draw'],
        'recordsTotal': total_count,
        'recordsFiltered': records_filtered,
        'filtered_stats': {
            'total': int(records_filtered),
            'approved': approved_filtered,
            'rejected': rejected_filtered,
            'in_review': in_review_filtered,
            'pending': pending_filtered,
        },
        'data': data,
    })


@disputes_center_bp.route('/cats/<int:cat_id>/edit', methods=['GET', 'POST'])
@require_law_firm
def edit_cat(cat_id):
    law_firm_id = get_current_law_firm_id()
    cat = FapContestationCat.query.filter_by(id=cat_id, law_firm_id=law_firm_id).first_or_404()

    if request.method == 'POST':
        old_first_instance_status = cat.first_instance_status or ''

        cat.employer_cnpj = request.form.get('employer_cnpj') or cat.employer_cnpj
        cat.employer_cnpj_assigned = request.form.get('employer_cnpj_assigned') or None
        cat.insured_nit = request.form.get('insured_nit') or cat.insured_nit
        cat.cat_block = request.form.get('cat_block') or cat.cat_block

        def parse_date(val):
            if not val:
                return None
            try:
                return datetime.strptime(val.strip(), '%Y-%m-%d').date()
            except ValueError:
                return None

        cat.insured_date_of_birth = parse_date(request.form.get('insured_date_of_birth')) or cat.insured_date_of_birth
        cat.insured_death_date = parse_date(request.form.get('insured_death_date'))
        cat.accident_date = parse_date(request.form.get('accident_date')) or cat.accident_date
        cat.cat_registration_date = parse_date(request.form.get('cat_registration_date')) or cat.cat_registration_date

        cat.first_instance_status = request.form.get('first_instance_status') or None
        cat.first_instance_status_raw = request.form.get('first_instance_status_raw') or None
        cat.first_instance_justification = request.form.get('first_instance_justification') or None
        cat.first_instance_opinion = request.form.get('first_instance_opinion') or None

        cat.second_instance_status = request.form.get('second_instance_status') or None
        cat.second_instance_status_raw = request.form.get('second_instance_status_raw') or None
        cat.second_instance_justification = request.form.get('second_instance_justification') or None
        cat.second_instance_opinion = request.form.get('second_instance_opinion') or None

        # Consolidado: prioriza 2a instância
        raw_consolidated = (
            request.form.get('second_instance_status') or request.form.get('first_instance_status') or ''
        ).strip().lower()
        if raw_consolidated == 'deferido':
            cat.status = 'approved'
        elif raw_consolidated == 'indeferido':
            cat.status = 'rejected'
        elif raw_consolidated == 'analyzing':
            cat.status = 'analyzing'
        else:
            cat.status = 'pending'
        cat.justification = (
            request.form.get('second_instance_justification')
            or request.form.get('first_instance_justification')
            or None
        )
        cat.opinion = (
            request.form.get('second_instance_opinion')
            or request.form.get('first_instance_opinion')
            or None
        )
        cat.notes = request.form.get('notes') or None
        cat.updated_at = datetime.utcnow()

        new_first_instance_status = cat.first_instance_status or ''
        if new_first_instance_status != old_first_instance_status:
            db.session.add(FapContestationCatManualHistory(
                law_firm_id=law_firm_id,
                cat_id=cat.id,
                performed_by_user_id=session.get('user_id'),
                action='edit_cat_first_instance_status',
                old_first_instance_status=old_first_instance_status or None,
                new_first_instance_status=new_first_instance_status,
                notes='Status da 1ª instância alterado na tela de edição.',
            ))

        try:
            db.session.commit()
            flash('CAT atualizada com sucesso!', 'success')
            return redirect(url_for('disputes_center.list_cats'))
        except Exception as exc:
            db.session.rollback()
            flash(f'Erro ao atualizar CAT: {exc}', 'danger')

    return render_template('disputes_center/cat_edit.html', cat=cat)


@disputes_center_bp.route('/cats/<int:cat_id>/timeline', methods=['GET'])
@require_law_firm
def cat_file_timeline(cat_id):
    law_firm_id = get_current_law_firm_id()
    cat = FapContestationCat.query.filter_by(id=cat_id, law_firm_id=law_firm_id).first_or_404()

    history_items = (
        FapContestationCatSourceHistory.query.filter_by(
            law_firm_id=law_firm_id,
            cat_id=cat_id,
        )
        .order_by(
            func.coalesce(
                FapContestationCatSourceHistory.publication_datetime,
                FapContestationCatSourceHistory.transmission_datetime,
            ).is_(None).asc(),
            func.coalesce(
                FapContestationCatSourceHistory.publication_datetime,
                FapContestationCatSourceHistory.transmission_datetime,
            ).desc(),
            FapContestationCatSourceHistory.created_at.desc(),
        )
        .all()
    )

    manual_history_items = (
        FapContestationCatManualHistory.query.filter_by(
            law_firm_id=law_firm_id,
            cat_id=cat_id,
        )
        .order_by(FapContestationCatManualHistory.created_at.desc(), FapContestationCatManualHistory.id.desc())
        .all()
    )

    events = []
    for item in history_items:
        report = item.report
        events.append(
            {
                'event_type': 'fap_file_history',
                'history_id': item.id,
                'report_id': item.report_id,
                'knowledge_base_id': item.knowledge_base_id,
                'action': item.action,
                'transmission_datetime': _format_datetime(item.transmission_datetime),
                'publication_datetime': _format_datetime(item.publication_datetime or item.transmission_datetime),
                'created_at': _format_datetime(item.created_at),
                'report_uploaded_at': _format_datetime(report.uploaded_at if report else None),
                'report_filename': (report.original_filename if report else None) or '-',
                'knowledge_details_url': (
                    url_for('disputes_center.view_fap_contestation_report', report_id=item.report_id)
                    if report else None
                ),
                'sort_datetime': item.publication_datetime or item.transmission_datetime or item.created_at,
            }
        )

    for item in manual_history_items:
        performer_name = (item.performed_by_user.name if item.performed_by_user else '') or 'Usuário não identificado'
        events.append(
            {
                'event_type': 'manual_history',
                'history_id': item.id,
                'report_id': None,
                'knowledge_base_id': None,
                'action': item.action,
                'manual_action_label': 'Edição manual',
                'manual_description': 'Status da 1ª instância alterado na tela de edição.',
                'performed_by': performer_name,
                'old_first_instance_status': item.old_first_instance_status,
                'new_first_instance_status': item.new_first_instance_status,
                'old_first_instance_status_label': _status_label_pt(item.old_first_instance_status),
                'new_first_instance_status_label': _status_label_pt(item.new_first_instance_status),
                'notes': item.notes,
                'transmission_datetime': None,
                'publication_datetime': None,
                'created_at': _format_datetime(item.created_at),
                'report_uploaded_at': None,
                'report_filename': None,
                'knowledge_details_url': None,
                'sort_datetime': item.created_at,
            }
        )

    events.sort(key=lambda e: e.get('sort_datetime') or datetime.min, reverse=True)
    for e in events:
        e.pop('sort_datetime', None)

    return jsonify(
        {
            'cat_id': cat.id,
            'cat_number': cat.cat_number,
            'events': events,
        }
    )


# ---------------------------------------------------------------------------
# Payroll Masses (Massa Salarial)
# ---------------------------------------------------------------------------

@disputes_center_bp.route('/payroll-masses')
@require_law_firm
def list_payroll_masses():
    law_firm_id = get_current_law_firm_id()
    base_query = _base_payroll_masses_query(law_firm_id)
    total_count = base_query.with_entities(func.count(FapContestationPayrollMass.id)).scalar() or 0

    first_status_counts = {
        _normalize_status_key(status): int(count or 0)
        for status, count in base_query.with_entities(
            func.lower(func.coalesce(cast(FapContestationPayrollMass.first_instance_status, String), '')),
            func.count(FapContestationPayrollMass.id),
        ).group_by(func.lower(func.coalesce(cast(FapContestationPayrollMass.first_instance_status, String), ''))).all()
    }
    first_instance_stats = _build_instance_stats(total_count, first_status_counts)

    second_status_counts = {
        _normalize_status_key(status): int(count or 0)
        for status, count in base_query.filter(
            ~func.lower(func.coalesce(cast(FapContestationPayrollMass.first_instance_status, String), '')).in_(['deferido', 'analyzing'])
        ).with_entities(
            func.lower(func.coalesce(cast(FapContestationPayrollMass.second_instance_status, String), '')),
            func.count(FapContestationPayrollMass.id),
        ).group_by(func.lower(func.coalesce(cast(FapContestationPayrollMass.second_instance_status, String), ''))).all()
    }
    first_deferred_count = int(first_status_counts.get('deferido', 0))
    first_analyzing_count = int(first_status_counts.get('analyzing', 0))
    second_total_base = max(total_count - first_deferred_count - first_analyzing_count, 0)
    second_instance_stats = _build_instance_stats(second_total_base, second_status_counts)

    cnpj_entries = (
        Client.query.with_entities(Client.cnpj, Client.name)
        .filter_by(law_firm_id=law_firm_id)
        .all()
    )

    roots_map = {}
    for cnpj, name in cnpj_entries:
        root = _extract_cnpj_root(cnpj)
        if not root:
            continue
        branch = _extract_cnpj_branch(cnpj)
        clean_name = (name or '').strip()
        if root not in roots_map:
            roots_map[root] = {'root': root, 'company_name': '', 'is_main': False}
        current = roots_map[root]
        if branch == '0001' and clean_name:
            current['company_name'] = clean_name
            current['is_main'] = True
            continue
        if not current['is_main'] and clean_name and not current['company_name']:
            current['company_name'] = clean_name

    cnpj_roots = [
        {'root': item['root'], 'company_name': item['company_name']}
        for _, item in sorted(roots_map.items(), key=lambda entry: entry[0])
    ]

    cnpj_by_root = {}
    for cnpj, name in cnpj_entries:
        root = _extract_cnpj_root(cnpj)
        digits = _normalize_cnpj_digits(cnpj)
        if not root or len(digits) < 14:
            continue
        clean_name = (name or '').strip()
        formatted = _format_cnpj(cnpj)
        if root not in cnpj_by_root:
            cnpj_by_root[root] = []
        if not any(item['digits'] == digits for item in cnpj_by_root[root]):
            cnpj_by_root[root].append({'cnpj': formatted, 'digits': digits, 'company_name': clean_name})
    for root_key in cnpj_by_root:
        cnpj_by_root[root_key].sort(key=lambda x: x['digits'])
        for item in cnpj_by_root[root_key]:
            del item['digits']

    initial_cnpj = _normalize_cnpj_digits(request.args.get('quick_cnpj', ''))

    current_vigencia_filter = None
    current_vigencia_id_raw = _normalize_text(request.args.get('vigencia_id', ''))
    if current_vigencia_id_raw:
        try:
            vigencia_obj = FapVigenciaCnpj.query.filter_by(
                id=int(current_vigencia_id_raw),
                law_firm_id=law_firm_id,
            ).first()
        except (TypeError, ValueError):
            vigencia_obj = None

        if vigencia_obj is not None:
            company_name = (
                db.session.query(FapContestationPayrollMass.employer_name)
                .filter(
                    FapContestationPayrollMass.law_firm_id == law_firm_id,
                    FapContestationPayrollMass.vigencia_id == vigencia_obj.id,
                    FapContestationPayrollMass.employer_name.is_not(None),
                    func.trim(FapContestationPayrollMass.employer_name) != '',
                )
                .order_by(FapContestationPayrollMass.updated_at.desc(), FapContestationPayrollMass.id.desc())
                .scalar()
                or ''
            ).strip()
            current_vigencia_filter = {
                'id': vigencia_obj.id,
                'year': (vigencia_obj.vigencia_year or '').strip(),
                'company_name': company_name,
                'company_cnpj': _format_cnpj(vigencia_obj.employer_cnpj),
            }

    return render_template(
        'disputes_center/payroll_masses.html',
        total_count=total_count,
        first_instance_stats=first_instance_stats,
        second_instance_stats=second_instance_stats,
        cnpj_roots=cnpj_roots,
        cnpj_by_root=cnpj_by_root,
        initial_cnpj=initial_cnpj,
        current_vigencia_filter=current_vigencia_filter,
    )


@disputes_center_bp.route('/api/payroll-masses', methods=['GET'])
@require_law_firm
def list_payroll_masses_api():
    law_firm_id = get_current_law_firm_id()
    payload = _collect_listing_payload(default_length=25)
    if request.is_json:
        raw = (request.get_json(silent=True) or {}).get('filters')
    else:
        raw = request.args.get('custom_filters', '[]')
    pm_filters = _parse_payroll_mass_custom_filters(raw)

    total_count = (
        _base_payroll_masses_query(law_firm_id)
        .with_entities(func.count(FapContestationPayrollMass.id))
        .scalar() or 0
    )

    filtered_query = _apply_payroll_mass_filters(
        _base_payroll_masses_query(law_firm_id),
        search_value=payload['search'],
        custom_filters=pm_filters,
        quick_root=payload['quick_root'],
        quick_cnpj=payload['quick_cnpj'],
        vigencia_id=payload['vigencia_id'],
    )
    records_filtered = filtered_query.with_entities(func.count(FapContestationPayrollMass.id)).scalar() or 0

    first_counts = {
        _normalize_status_key(s): int(c or 0)
        for s, c in filtered_query.with_entities(
            func.lower(func.coalesce(cast(FapContestationPayrollMass.first_instance_status, String), '')),
            func.count(FapContestationPayrollMass.id),
        ).group_by(func.lower(func.coalesce(cast(FapContestationPayrollMass.first_instance_status, String), ''))).all()
    }
    approved_filtered = first_counts.get('deferido', 0)
    rejected_filtered = first_counts.get('indeferido', 0)
    in_review_filtered = first_counts.get('analyzing', 0)
    pending_filtered = max(int(records_filtered) - approved_filtered - rejected_filtered - in_review_filtered, 0)

    order_column = PAYROLL_MASS_ORDER_COLUMN_MAP.get(payload['order_column'], FapContestationPayrollMass.id)
    if payload['order_dir'] == 'asc':
        filtered_query = filtered_query.order_by(order_column.asc(), FapContestationPayrollMass.id.asc())
    else:
        filtered_query = filtered_query.order_by(order_column.desc(), FapContestationPayrollMass.id.desc())

    paged_results = filtered_query.offset(payload['start']).limit(payload['length']).all()
    data = [_serialize_payroll_mass_row(pm) for pm in paged_results]

    return jsonify({
        'draw': payload['draw'],
        'recordsTotal': total_count,
        'recordsFiltered': records_filtered,
        'filtered_stats': {
            'total': int(records_filtered),
            'approved': approved_filtered,
            'rejected': rejected_filtered,
            'in_review': in_review_filtered,
            'pending': pending_filtered,
        },
        'data': data,
    })


@disputes_center_bp.route('/payroll-masses/<int:payroll_mass_id>/edit', methods=['GET', 'POST'])
@require_law_firm
def edit_payroll_mass(payroll_mass_id):
    law_firm_id = get_current_law_firm_id()
    pm = FapContestationPayrollMass.query.filter_by(id=payroll_mass_id, law_firm_id=law_firm_id).first_or_404()

    if request.method == 'POST':
        old_first_instance_status = pm.first_instance_status or ''

        def parse_date(val):
            if not val:
                return None
            try:
                return datetime.strptime(val.strip(), '%Y-%m-%d').date()
            except ValueError:
                return None

        def parse_decimal(val):
            from decimal import Decimal, InvalidOperation
            if not val:
                return None
            try:
                return Decimal(val.strip().replace(',', '.'))
            except (InvalidOperation, ValueError):
                return None

        pm.employer_cnpj = request.form.get('employer_cnpj') or pm.employer_cnpj
        pm.employer_name = request.form.get('employer_name') or pm.employer_name
        pm.competence = request.form.get('competence') or pm.competence
        pm.total_remuneration = parse_decimal(request.form.get('total_remuneration'))
        pm.first_instance_requested_value = parse_decimal(request.form.get('first_instance_requested_value'))
        pm.second_instance_requested_value = parse_decimal(request.form.get('second_instance_requested_value'))

        pm.first_instance_status = request.form.get('first_instance_status') or None
        pm.first_instance_status_raw = request.form.get('first_instance_status_raw') or None
        pm.first_instance_justification = request.form.get('first_instance_justification') or None
        pm.first_instance_opinion = request.form.get('first_instance_opinion') or None

        pm.second_instance_status = request.form.get('second_instance_status') or None
        pm.second_instance_status_raw = request.form.get('second_instance_status_raw') or None
        pm.second_instance_justification = request.form.get('second_instance_justification') or None
        pm.second_instance_opinion = request.form.get('second_instance_opinion') or None

        raw_consolidated = (
            request.form.get('second_instance_status') or request.form.get('first_instance_status') or ''
        ).strip().lower()
        if raw_consolidated == 'deferido':
            pm.status = 'approved'
        elif raw_consolidated == 'indeferido':
            pm.status = 'rejected'
        elif raw_consolidated == 'analyzing':
            pm.status = 'analyzing'
        else:
            pm.status = 'pending'
        pm.justification = (
            request.form.get('second_instance_justification')
            or request.form.get('first_instance_justification')
            or None
        )
        pm.opinion = (
            request.form.get('second_instance_opinion')
            or request.form.get('first_instance_opinion')
            or None
        )
        pm.notes = request.form.get('notes') or None
        pm.updated_at = datetime.utcnow()

        new_first_instance_status = pm.first_instance_status or ''
        if new_first_instance_status != old_first_instance_status:
            db.session.add(FapContestationPayrollMassManualHistory(
                law_firm_id=law_firm_id,
                payroll_mass_id=pm.id,
                performed_by_user_id=session.get('user_id'),
                action='edit_payroll_mass_first_instance_status',
                old_first_instance_status=old_first_instance_status or None,
                new_first_instance_status=new_first_instance_status,
                notes='Status da 1ª instância alterado na tela de edição.',
            ))

        try:
            db.session.commit()
            flash('Massa Salarial atualizada com sucesso!', 'success')
            return redirect(url_for('disputes_center.list_payroll_masses'))
        except Exception as exc:
            db.session.rollback()
            flash(f'Erro ao atualizar Massa Salarial: {exc}', 'danger')

    return render_template('disputes_center/payroll_mass_edit.html', pm=pm)


@disputes_center_bp.route('/payroll-masses/<int:payroll_mass_id>/timeline', methods=['GET'])
@require_law_firm
def payroll_mass_timeline(payroll_mass_id):
    law_firm_id = get_current_law_firm_id()
    pm = FapContestationPayrollMass.query.filter_by(id=payroll_mass_id, law_firm_id=law_firm_id).first_or_404()

    history_items = (
        FapContestationPayrollMassSourceHistory.query.filter_by(
            law_firm_id=law_firm_id,
            payroll_mass_id=payroll_mass_id,
        )
        .order_by(
            func.coalesce(
                FapContestationPayrollMassSourceHistory.publication_datetime,
                FapContestationPayrollMassSourceHistory.transmission_datetime,
            ).is_(None).asc(),
            func.coalesce(
                FapContestationPayrollMassSourceHistory.publication_datetime,
                FapContestationPayrollMassSourceHistory.transmission_datetime,
            ).desc(),
            FapContestationPayrollMassSourceHistory.created_at.desc(),
        )
        .all()
    )

    manual_history_items = (
        FapContestationPayrollMassManualHistory.query.filter_by(
            law_firm_id=law_firm_id,
            payroll_mass_id=payroll_mass_id,
        )
        .order_by(FapContestationPayrollMassManualHistory.created_at.desc(), FapContestationPayrollMassManualHistory.id.desc())
        .all()
    )

    events = []
    for item in history_items:
        report = item.report
        events.append({
            'event_type': 'fap_file_history',
            'history_id': item.id,
            'report_id': item.report_id,
            'knowledge_base_id': item.knowledge_base_id,
            'action': item.action,
            'transmission_datetime': _format_datetime(item.transmission_datetime),
            'publication_datetime': _format_datetime(item.publication_datetime or item.transmission_datetime),
            'created_at': _format_datetime(item.created_at),
            'report_uploaded_at': _format_datetime(report.uploaded_at if report else None),
            'report_filename': (report.original_filename if report else None) or '-',
            'knowledge_details_url': (
                url_for('disputes_center.view_fap_contestation_report', report_id=item.report_id)
                if report else None
            ),
            'sort_datetime': item.publication_datetime or item.transmission_datetime or item.created_at,
        })

    for item in manual_history_items:
        performer_name = (item.performed_by_user.name if item.performed_by_user else '') or 'Usuário não identificado'
        events.append({
            'event_type': 'manual_history',
            'history_id': item.id,
            'report_id': None,
            'knowledge_base_id': None,
            'action': item.action,
            'manual_action_label': 'Edição manual',
            'manual_description': 'Status da 1ª instância alterado na tela de edição.',
            'performed_by': performer_name,
            'old_first_instance_status': item.old_first_instance_status,
            'new_first_instance_status': item.new_first_instance_status,
            'old_first_instance_status_label': _status_label_pt(item.old_first_instance_status),
            'new_first_instance_status_label': _status_label_pt(item.new_first_instance_status),
            'notes': item.notes,
            'transmission_datetime': None,
            'publication_datetime': None,
            'created_at': _format_datetime(item.created_at),
            'report_uploaded_at': None,
            'report_filename': None,
            'knowledge_details_url': None,
            'sort_datetime': item.created_at,
        })

    events.sort(key=lambda e: e.get('sort_datetime') or datetime.min, reverse=True)
    for e in events:
        e.pop('sort_datetime', None)

    return jsonify({
        'payroll_mass_id': pm.id,
        'employer_cnpj': pm.employer_cnpj,
        'competence': pm.competence,
        'events': events,
    })


# ──────────────────────────────────────────────────────────────────────────────
# Número Médio de Vínculos routes
# ──────────────────────────────────────────────────────────────────────────────

@disputes_center_bp.route('/employment-links', methods=['GET'])
@require_law_firm
def list_employment_links():
    law_firm_id = get_current_law_firm_id()
    base_query = _base_employment_links_query(law_firm_id)
    total_count = base_query.with_entities(func.count(FapContestationEmploymentLink.id)).scalar() or 0

    first_status_counts = {
        _normalize_status_key(status): int(count or 0)
        for status, count in base_query.with_entities(
            func.lower(func.coalesce(cast(FapContestationEmploymentLink.first_instance_status, String), '')),
            func.count(FapContestationEmploymentLink.id),
        ).group_by(func.lower(func.coalesce(cast(FapContestationEmploymentLink.first_instance_status, String), ''))).all()
    }
    first_instance_stats = _build_instance_stats(total_count, first_status_counts)

    second_status_counts = {
        _normalize_status_key(status): int(count or 0)
        for status, count in base_query.filter(
            ~func.lower(func.coalesce(cast(FapContestationEmploymentLink.first_instance_status, String), '')).in_(['deferido', 'analyzing'])
        ).with_entities(
            func.lower(func.coalesce(cast(FapContestationEmploymentLink.second_instance_status, String), '')),
            func.count(FapContestationEmploymentLink.id),
        ).group_by(func.lower(func.coalesce(cast(FapContestationEmploymentLink.second_instance_status, String), ''))).all()
    }
    first_deferred_count = int(first_status_counts.get('deferido', 0))
    first_analyzing_count = int(first_status_counts.get('analyzing', 0))
    second_total_base = max(total_count - first_deferred_count - first_analyzing_count, 0)
    second_instance_stats = _build_instance_stats(second_total_base, second_status_counts)

    cnpj_entries = (
        Client.query.with_entities(Client.cnpj, Client.name)
        .filter_by(law_firm_id=law_firm_id)
        .all()
    )

    roots_map = {}
    for cnpj, name in cnpj_entries:
        root = _extract_cnpj_root(cnpj)
        if not root:
            continue
        branch = _extract_cnpj_branch(cnpj)
        clean_name = (name or '').strip()
        if root not in roots_map:
            roots_map[root] = {'root': root, 'company_name': '', 'is_main': False}
        current = roots_map[root]
        if branch == '0001' and clean_name:
            current['company_name'] = clean_name
            current['is_main'] = True
            continue
        if not current['is_main'] and clean_name and not current['company_name']:
            current['company_name'] = clean_name

    cnpj_roots = [
        {'root': item['root'], 'company_name': item['company_name']}
        for _, item in sorted(roots_map.items(), key=lambda entry: entry[0])
    ]

    cnpj_by_root = {}
    for cnpj, name in cnpj_entries:
        root = _extract_cnpj_root(cnpj)
        digits = _normalize_cnpj_digits(cnpj)
        if not root or len(digits) < 14:
            continue
        clean_name = (name or '').strip()
        formatted = _format_cnpj(cnpj)
        if root not in cnpj_by_root:
            cnpj_by_root[root] = []
        if not any(item['digits'] == digits for item in cnpj_by_root[root]):
            cnpj_by_root[root].append({'cnpj': formatted, 'digits': digits, 'company_name': clean_name})
    for root_key in cnpj_by_root:
        cnpj_by_root[root_key].sort(key=lambda x: x['digits'])
        for item in cnpj_by_root[root_key]:
            del item['digits']

    initial_cnpj = _normalize_cnpj_digits(request.args.get('quick_cnpj', ''))

    current_vigencia_filter = None
    current_vigencia_id_raw = _normalize_text(request.args.get('vigencia_id', ''))
    if current_vigencia_id_raw:
        try:
            vigencia_obj = FapVigenciaCnpj.query.filter_by(
                id=int(current_vigencia_id_raw),
                law_firm_id=law_firm_id,
            ).first()
        except (TypeError, ValueError):
            vigencia_obj = None

        if vigencia_obj is not None:
            company_name = (
                db.session.query(FapContestationEmploymentLink.employer_name)
                .filter(
                    FapContestationEmploymentLink.law_firm_id == law_firm_id,
                    FapContestationEmploymentLink.vigencia_id == vigencia_obj.id,
                    FapContestationEmploymentLink.employer_name.is_not(None),
                    func.trim(FapContestationEmploymentLink.employer_name) != '',
                )
                .order_by(FapContestationEmploymentLink.updated_at.desc(), FapContestationEmploymentLink.id.desc())
                .scalar()
                or ''
            ).strip()
            current_vigencia_filter = {
                'id': vigencia_obj.id,
                'year': (vigencia_obj.vigencia_year or '').strip(),
                'company_name': company_name,
                'company_cnpj': _format_cnpj(vigencia_obj.employer_cnpj),
            }

    return render_template(
        'disputes_center/employment_links.html',
        total_count=total_count,
        first_instance_stats=first_instance_stats,
        second_instance_stats=second_instance_stats,
        cnpj_roots=cnpj_roots,
        cnpj_by_root=cnpj_by_root,
        initial_cnpj=initial_cnpj,
        current_vigencia_filter=current_vigencia_filter,
    )


@disputes_center_bp.route('/api/employment-links', methods=['GET'])
@require_law_firm
def list_employment_links_api():
    law_firm_id = get_current_law_firm_id()
    payload = _collect_listing_payload(default_length=25)
    if request.is_json:
        raw = (request.get_json(silent=True) or {}).get('filters')
    else:
        raw = request.args.get('custom_filters', '[]')
    el_filters = _parse_employment_link_custom_filters(raw)

    total_count = (
        _base_employment_links_query(law_firm_id)
        .with_entities(func.count(FapContestationEmploymentLink.id))
        .scalar() or 0
    )

    filtered_query = _apply_employment_link_filters(
        _base_employment_links_query(law_firm_id),
        search_value=payload['search'],
        custom_filters=el_filters,
        quick_root=payload['quick_root'],
        quick_cnpj=payload['quick_cnpj'],
        vigencia_id=payload['vigencia_id'],
    )
    records_filtered = filtered_query.with_entities(func.count(FapContestationEmploymentLink.id)).scalar() or 0

    first_counts = {
        _normalize_status_key(s): int(c or 0)
        for s, c in filtered_query.with_entities(
            func.lower(func.coalesce(cast(FapContestationEmploymentLink.first_instance_status, String), '')),
            func.count(FapContestationEmploymentLink.id),
        ).group_by(func.lower(func.coalesce(cast(FapContestationEmploymentLink.first_instance_status, String), ''))).all()
    }
    approved_filtered = first_counts.get('deferido', 0)
    rejected_filtered = first_counts.get('indeferido', 0)
    in_review_filtered = first_counts.get('analyzing', 0)
    pending_filtered = max(int(records_filtered) - approved_filtered - rejected_filtered - in_review_filtered, 0)

    order_column = EMPLOYMENT_LINK_ORDER_COLUMN_MAP.get(payload['order_column'], FapContestationEmploymentLink.id)
    if payload['order_dir'] == 'asc':
        filtered_query = filtered_query.order_by(order_column.asc(), FapContestationEmploymentLink.id.asc())
    else:
        filtered_query = filtered_query.order_by(order_column.desc(), FapContestationEmploymentLink.id.desc())

    paged_results = filtered_query.offset(payload['start']).limit(payload['length']).all()
    data = [_serialize_employment_link_row(el) for el in paged_results]

    return jsonify({
        'draw': payload['draw'],
        'recordsTotal': total_count,
        'recordsFiltered': records_filtered,
        'filtered_stats': {
            'total': int(records_filtered),
            'approved': approved_filtered,
            'rejected': rejected_filtered,
            'in_review': in_review_filtered,
            'pending': pending_filtered,
        },
        'data': data,
    })


@disputes_center_bp.route('/employment-links/<int:employment_link_id>/edit', methods=['GET', 'POST'])
@require_law_firm
def edit_employment_link(employment_link_id):
    law_firm_id = get_current_law_firm_id()
    el = FapContestationEmploymentLink.query.filter_by(id=employment_link_id, law_firm_id=law_firm_id).first_or_404()

    if request.method == 'POST':
        old_first_instance_status = el.first_instance_status or ''

        def parse_int(val):
            if not val:
                return None
            try:
                return int(str(val).strip())
            except (ValueError, TypeError):
                return None

        el.employer_cnpj = request.form.get('employer_cnpj') or el.employer_cnpj
        el.employer_name = request.form.get('employer_name') or el.employer_name
        el.competence = request.form.get('competence') or el.competence
        el.quantity = parse_int(request.form.get('quantity'))
        el.first_instance_requested_quantity = parse_int(request.form.get('first_instance_requested_quantity'))
        el.second_instance_requested_quantity = parse_int(request.form.get('second_instance_requested_quantity'))

        el.first_instance_status = request.form.get('first_instance_status') or None
        el.first_instance_status_raw = request.form.get('first_instance_status_raw') or None
        el.first_instance_justification = request.form.get('first_instance_justification') or None
        el.first_instance_opinion = request.form.get('first_instance_opinion') or None

        el.second_instance_status = request.form.get('second_instance_status') or None
        el.second_instance_status_raw = request.form.get('second_instance_status_raw') or None
        el.second_instance_justification = request.form.get('second_instance_justification') or None
        el.second_instance_opinion = request.form.get('second_instance_opinion') or None

        raw_consolidated = (
            request.form.get('second_instance_status') or request.form.get('first_instance_status') or ''
        ).strip().lower()
        if raw_consolidated == 'deferido':
            el.status = 'approved'
        elif raw_consolidated == 'indeferido':
            el.status = 'rejected'
        elif raw_consolidated == 'analyzing':
            el.status = 'analyzing'
        else:
            el.status = 'pending'
        el.justification = (
            request.form.get('second_instance_justification')
            or request.form.get('first_instance_justification')
            or None
        )
        el.opinion = (
            request.form.get('second_instance_opinion')
            or request.form.get('first_instance_opinion')
            or None
        )
        el.notes = request.form.get('notes') or None
        el.updated_at = datetime.utcnow()

        new_first_instance_status = el.first_instance_status or ''
        if new_first_instance_status != old_first_instance_status:
            db.session.add(FapContestationEmploymentLinkManualHistory(
                law_firm_id=law_firm_id,
                employment_link_id=el.id,
                performed_by_user_id=session.get('user_id'),
                action='edit_employment_link_first_instance_status',
                old_first_instance_status=old_first_instance_status or None,
                new_first_instance_status=new_first_instance_status,
                notes='Status da 1ª instância alterado na tela de edição.',
            ))

        try:
            db.session.commit()
            flash('Vínculo atualizado com sucesso!', 'success')
            return redirect(url_for('disputes_center.list_employment_links'))
        except Exception as exc:
            db.session.rollback()
            flash(f'Erro ao atualizar Vínculo: {exc}', 'danger')

    return render_template('disputes_center/employment_link_edit.html', el=el)


@disputes_center_bp.route('/employment-links/<int:employment_link_id>/timeline', methods=['GET'])
@require_law_firm
def employment_link_timeline(employment_link_id):
    law_firm_id = get_current_law_firm_id()
    el = FapContestationEmploymentLink.query.filter_by(id=employment_link_id, law_firm_id=law_firm_id).first_or_404()

    history_items = (
        FapContestationEmploymentLinkSourceHistory.query.filter_by(
            law_firm_id=law_firm_id,
            employment_link_id=employment_link_id,
        )
        .order_by(
            func.coalesce(
                FapContestationEmploymentLinkSourceHistory.publication_datetime,
                FapContestationEmploymentLinkSourceHistory.transmission_datetime,
            ).is_(None).asc(),
            func.coalesce(
                FapContestationEmploymentLinkSourceHistory.publication_datetime,
                FapContestationEmploymentLinkSourceHistory.transmission_datetime,
            ).desc(),
            FapContestationEmploymentLinkSourceHistory.created_at.desc(),
        )
        .all()
    )

    manual_history_items = (
        FapContestationEmploymentLinkManualHistory.query.filter_by(
            law_firm_id=law_firm_id,
            employment_link_id=employment_link_id,
        )
        .order_by(FapContestationEmploymentLinkManualHistory.created_at.desc(), FapContestationEmploymentLinkManualHistory.id.desc())
        .all()
    )

    events = []
    for item in history_items:
        report = item.report
        events.append({
            'event_type': 'fap_file_history',
            'history_id': item.id,
            'report_id': item.report_id,
            'knowledge_base_id': item.knowledge_base_id,
            'action': item.action,
            'transmission_datetime': _format_datetime(item.transmission_datetime),
            'publication_datetime': _format_datetime(item.publication_datetime or item.transmission_datetime),
            'created_at': _format_datetime(item.created_at),
            'report_uploaded_at': _format_datetime(report.uploaded_at if report else None),
            'report_filename': (report.original_filename if report else None) or '-',
            'knowledge_details_url': (
                url_for('disputes_center.view_fap_contestation_report', report_id=item.report_id)
                if report else None
            ),
            'sort_datetime': item.publication_datetime or item.transmission_datetime or item.created_at,
        })

    for item in manual_history_items:
        performer_name = (item.performed_by_user.name if item.performed_by_user else '') or 'Usuário não identificado'
        events.append({
            'event_type': 'manual_history',
            'history_id': item.id,
            'report_id': None,
            'knowledge_base_id': None,
            'action': item.action,
            'manual_action_label': 'Edição manual',
            'manual_description': 'Status da 1ª instância alterado na tela de edição.',
            'performed_by': performer_name,
            'old_first_instance_status': item.old_first_instance_status,
            'new_first_instance_status': item.new_first_instance_status,
            'old_first_instance_status_label': _status_label_pt(item.old_first_instance_status),
            'new_first_instance_status_label': _status_label_pt(item.new_first_instance_status),
            'notes': item.notes,
            'transmission_datetime': None,
            'publication_datetime': None,
            'created_at': _format_datetime(item.created_at),
            'report_uploaded_at': None,
            'report_filename': None,
            'knowledge_details_url': None,
            'sort_datetime': item.created_at,
        })

    events.sort(key=lambda e: e.get('sort_datetime') or datetime.min, reverse=True)
    for e in events:
        e.pop('sort_datetime', None)

    return jsonify({
        'employment_link_id': el.id,
        'employer_cnpj': el.employer_cnpj,
        'competence': el.competence,
        'events': events,
    })


@disputes_center_bp.route('/turnover-rates', methods=['GET'])
@require_law_firm
def list_turnover_rates():
    law_firm_id = get_current_law_firm_id()
    base_query = _base_turnover_rates_query(law_firm_id)
    total_count = base_query.with_entities(func.count(FapContestationTurnoverRate.id)).scalar() or 0

    first_status_counts = {
        _normalize_status_key(status): int(count or 0)
        for status, count in base_query.with_entities(
            func.lower(func.coalesce(cast(FapContestationTurnoverRate.first_instance_status, String), '')),
            func.count(FapContestationTurnoverRate.id),
        ).group_by(func.lower(func.coalesce(cast(FapContestationTurnoverRate.first_instance_status, String), ''))).all()
    }
    first_instance_stats = _build_instance_stats(total_count, first_status_counts)

    second_status_counts = {
        _normalize_status_key(status): int(count or 0)
        for status, count in base_query.filter(
            ~func.lower(func.coalesce(cast(FapContestationTurnoverRate.first_instance_status, String), '')).in_(['deferido', 'analyzing'])
        ).with_entities(
            func.lower(func.coalesce(cast(FapContestationTurnoverRate.second_instance_status, String), '')),
            func.count(FapContestationTurnoverRate.id),
        ).group_by(func.lower(func.coalesce(cast(FapContestationTurnoverRate.second_instance_status, String), ''))).all()
    }
    first_deferred_count = int(first_status_counts.get('deferido', 0))
    first_analyzing_count = int(first_status_counts.get('analyzing', 0))
    second_total_base = max(total_count - first_deferred_count - first_analyzing_count, 0)
    second_instance_stats = _build_instance_stats(second_total_base, second_status_counts)

    cnpj_entries = (
        Client.query.with_entities(Client.cnpj, Client.name)
        .filter_by(law_firm_id=law_firm_id)
        .all()
    )

    roots_map = {}
    for cnpj, name in cnpj_entries:
        root = _extract_cnpj_root(cnpj)
        if not root:
            continue
        branch = _extract_cnpj_branch(cnpj)
        clean_name = (name or '').strip()
        if root not in roots_map:
            roots_map[root] = {'root': root, 'company_name': '', 'is_main': False}
        current = roots_map[root]
        if branch == '0001' and clean_name:
            current['company_name'] = clean_name
            current['is_main'] = True
            continue
        if not current['is_main'] and clean_name and not current['company_name']:
            current['company_name'] = clean_name

    cnpj_roots = [
        {'root': item['root'], 'company_name': item['company_name']}
        for _, item in sorted(roots_map.items(), key=lambda entry: entry[0])
    ]

    cnpj_by_root = {}
    for cnpj, name in cnpj_entries:
        root = _extract_cnpj_root(cnpj)
        digits = _normalize_cnpj_digits(cnpj)
        if not root or len(digits) < 14:
            continue
        clean_name = (name or '').strip()
        formatted = _format_cnpj(cnpj)
        if root not in cnpj_by_root:
            cnpj_by_root[root] = []
        if not any(item['digits'] == digits for item in cnpj_by_root[root]):
            cnpj_by_root[root].append({'cnpj': formatted, 'digits': digits, 'company_name': clean_name})
    for root_key in cnpj_by_root:
        cnpj_by_root[root_key].sort(key=lambda x: x['digits'])
        for item in cnpj_by_root[root_key]:
            del item['digits']

    initial_cnpj = _normalize_cnpj_digits(request.args.get('quick_cnpj', ''))

    current_vigencia_filter = None
    current_vigencia_id_raw = _normalize_text(request.args.get('vigencia_id', ''))
    if current_vigencia_id_raw:
        try:
            vigencia_obj = FapVigenciaCnpj.query.filter_by(
                id=int(current_vigencia_id_raw),
                law_firm_id=law_firm_id,
            ).first()
        except (TypeError, ValueError):
            vigencia_obj = None

        if vigencia_obj is not None:
            company_name = (
                db.session.query(FapContestationTurnoverRate.employer_name)
                .filter(
                    FapContestationTurnoverRate.law_firm_id == law_firm_id,
                    FapContestationTurnoverRate.vigencia_id == vigencia_obj.id,
                    FapContestationTurnoverRate.employer_name.is_not(None),
                    func.trim(FapContestationTurnoverRate.employer_name) != '',
                )
                .order_by(FapContestationTurnoverRate.updated_at.desc(), FapContestationTurnoverRate.id.desc())
                .scalar()
                or ''
            ).strip()
            current_vigencia_filter = {
                'id': vigencia_obj.id,
                'year': (vigencia_obj.vigencia_year or '').strip(),
                'company_name': company_name,
                'company_cnpj': _format_cnpj(vigencia_obj.employer_cnpj),
            }

    return render_template(
        'disputes_center/turnover_rates.html',
        total_count=total_count,
        first_instance_stats=first_instance_stats,
        second_instance_stats=second_instance_stats,
        cnpj_roots=cnpj_roots,
        cnpj_by_root=cnpj_by_root,
        initial_cnpj=initial_cnpj,
        current_vigencia_filter=current_vigencia_filter,
    )


@disputes_center_bp.route('/api/turnover-rates', methods=['GET'])
@require_law_firm
def list_turnover_rates_api():
    law_firm_id = get_current_law_firm_id()
    payload = _collect_listing_payload(default_length=25)
    if request.is_json:
        raw = (request.get_json(silent=True) or {}).get('filters')
    else:
        raw = request.args.get('custom_filters', '[]')
    tr_filters = _parse_turnover_rate_custom_filters(raw)

    total_count = (
        _base_turnover_rates_query(law_firm_id)
        .with_entities(func.count(FapContestationTurnoverRate.id))
        .scalar() or 0
    )

    filtered_query = _apply_turnover_rate_filters(
        _base_turnover_rates_query(law_firm_id),
        search_value=payload['search'],
        custom_filters=tr_filters,
        quick_root=payload['quick_root'],
        quick_cnpj=payload['quick_cnpj'],
        vigencia_id=payload['vigencia_id'],
    )
    records_filtered = filtered_query.with_entities(func.count(FapContestationTurnoverRate.id)).scalar() or 0

    first_counts = {
        _normalize_status_key(s): int(c or 0)
        for s, c in filtered_query.with_entities(
            func.lower(func.coalesce(cast(FapContestationTurnoverRate.first_instance_status, String), '')),
            func.count(FapContestationTurnoverRate.id),
        ).group_by(func.lower(func.coalesce(cast(FapContestationTurnoverRate.first_instance_status, String), ''))).all()
    }
    approved_filtered = first_counts.get('deferido', 0)
    rejected_filtered = first_counts.get('indeferido', 0)
    in_review_filtered = first_counts.get('analyzing', 0)
    pending_filtered = max(int(records_filtered) - approved_filtered - rejected_filtered - in_review_filtered, 0)

    order_column = TURNOVER_RATE_ORDER_COLUMN_MAP.get(payload['order_column'], FapContestationTurnoverRate.id)
    if payload['order_dir'] == 'asc':
        filtered_query = filtered_query.order_by(order_column.asc(), FapContestationTurnoverRate.id.asc())
    else:
        filtered_query = filtered_query.order_by(order_column.desc(), FapContestationTurnoverRate.id.desc())

    paged_results = filtered_query.offset(payload['start']).limit(payload['length']).all()
    data = [_serialize_turnover_rate_row(tr) for tr in paged_results]

    return jsonify({
        'draw': payload['draw'],
        'recordsTotal': total_count,
        'recordsFiltered': records_filtered,
        'filtered_stats': {
            'total': int(records_filtered),
            'approved': approved_filtered,
            'rejected': rejected_filtered,
            'in_review': in_review_filtered,
            'pending': pending_filtered,
        },
        'data': data,
    })


@disputes_center_bp.route('/turnover-rates/<int:turnover_rate_id>/edit', methods=['GET', 'POST'])
@require_law_firm
def edit_turnover_rate(turnover_rate_id):
    law_firm_id = get_current_law_firm_id()
    tr = FapContestationTurnoverRate.query.filter_by(id=turnover_rate_id, law_firm_id=law_firm_id).first_or_404()

    if request.method == 'POST':
        old_first_instance_status = tr.first_instance_status or ''

        def parse_int(val):
            if not val:
                return None
            try:
                return int(str(val).strip())
            except (ValueError, TypeError):
                return None

        def parse_decimal(val):
            if not val:
                return None
            from decimal import Decimal, InvalidOperation
            try:
                cleaned = str(val).strip().replace(',', '.')
                return Decimal(cleaned)
            except (InvalidOperation, ValueError):
                return None

        tr.employer_cnpj = request.form.get('employer_cnpj') or tr.employer_cnpj
        tr.employer_name = request.form.get('employer_name') or tr.employer_name
        tr.year = request.form.get('year') or tr.year
        tr.turnover_rate = parse_decimal(request.form.get('turnover_rate'))
        tr.admissions = parse_int(request.form.get('admissions'))
        tr.dismissals = parse_int(request.form.get('dismissals'))
        tr.initial_links_count = parse_int(request.form.get('initial_links_count'))
        tr.first_instance_requested_admissions = parse_int(request.form.get('first_instance_requested_admissions'))
        tr.first_instance_requested_dismissals = parse_int(request.form.get('first_instance_requested_dismissals'))
        tr.first_instance_requested_initial_links = parse_int(request.form.get('first_instance_requested_initial_links'))
        tr.second_instance_requested_admissions = parse_int(request.form.get('second_instance_requested_admissions'))
        tr.second_instance_requested_dismissals = parse_int(request.form.get('second_instance_requested_dismissals'))
        tr.second_instance_requested_initial_links = parse_int(request.form.get('second_instance_requested_initial_links'))

        tr.first_instance_status = request.form.get('first_instance_status') or None
        tr.first_instance_status_raw = request.form.get('first_instance_status_raw') or None
        tr.first_instance_justification = request.form.get('first_instance_justification') or None
        tr.first_instance_opinion = request.form.get('first_instance_opinion') or None

        tr.second_instance_status = request.form.get('second_instance_status') or None
        tr.second_instance_status_raw = request.form.get('second_instance_status_raw') or None
        tr.second_instance_justification = request.form.get('second_instance_justification') or None
        tr.second_instance_opinion = request.form.get('second_instance_opinion') or None

        raw_consolidated = (
            request.form.get('second_instance_status') or request.form.get('first_instance_status') or ''
        ).strip().lower()
        if raw_consolidated == 'deferido':
            tr.status = 'approved'
        elif raw_consolidated == 'indeferido':
            tr.status = 'rejected'
        elif raw_consolidated == 'analyzing':
            tr.status = 'analyzing'
        else:
            tr.status = 'pending'
        tr.justification = (
            request.form.get('second_instance_justification')
            or request.form.get('first_instance_justification')
            or None
        )
        tr.opinion = (
            request.form.get('second_instance_opinion')
            or request.form.get('first_instance_opinion')
            or None
        )
        tr.notes = request.form.get('notes') or None
        tr.updated_at = datetime.utcnow()

        new_first_instance_status = tr.first_instance_status or ''
        if new_first_instance_status != old_first_instance_status:
            db.session.add(FapContestationTurnoverRateManualHistory(
                law_firm_id=law_firm_id,
                turnover_rate_id=tr.id,
                performed_by_user_id=session.get('user_id'),
                action='edit_turnover_rate_first_instance_status',
                old_first_instance_status=old_first_instance_status or None,
                new_first_instance_status=new_first_instance_status,
                notes='Status da 1ª instância alterado na tela de edição.',
            ))

        try:
            db.session.commit()
            flash('Taxa de Rotatividade atualizada com sucesso!', 'success')
            return redirect(url_for('disputes_center.list_turnover_rates'))
        except Exception as exc:
            db.session.rollback()
            flash(f'Erro ao atualizar Taxa de Rotatividade: {exc}', 'danger')

    return render_template('disputes_center/turnover_rate_edit.html', tr=tr)


@disputes_center_bp.route('/turnover-rates/<int:turnover_rate_id>/timeline', methods=['GET'])
@require_law_firm
def turnover_rate_timeline(turnover_rate_id):
    law_firm_id = get_current_law_firm_id()
    tr = FapContestationTurnoverRate.query.filter_by(id=turnover_rate_id, law_firm_id=law_firm_id).first_or_404()

    history_items = (
        FapContestationTurnoverRateSourceHistory.query.filter_by(
            law_firm_id=law_firm_id,
            turnover_rate_id=turnover_rate_id,
        )
        .order_by(
            func.coalesce(
                FapContestationTurnoverRateSourceHistory.publication_datetime,
                FapContestationTurnoverRateSourceHistory.transmission_datetime,
            ).is_(None).asc(),
            func.coalesce(
                FapContestationTurnoverRateSourceHistory.publication_datetime,
                FapContestationTurnoverRateSourceHistory.transmission_datetime,
            ).desc(),
            FapContestationTurnoverRateSourceHistory.created_at.desc(),
        )
        .all()
    )

    manual_history_items = (
        FapContestationTurnoverRateManualHistory.query.filter_by(
            law_firm_id=law_firm_id,
            turnover_rate_id=turnover_rate_id,
        )
        .order_by(FapContestationTurnoverRateManualHistory.created_at.desc(), FapContestationTurnoverRateManualHistory.id.desc())
        .all()
    )

    events = []
    for item in history_items:
        report = item.report
        events.append({
            'event_type': 'fap_file_history',
            'history_id': item.id,
            'report_id': item.report_id,
            'knowledge_base_id': item.knowledge_base_id,
            'action': item.action,
            'transmission_datetime': _format_datetime(item.transmission_datetime),
            'publication_datetime': _format_datetime(item.publication_datetime or item.transmission_datetime),
            'created_at': _format_datetime(item.created_at),
            'report_uploaded_at': _format_datetime(report.uploaded_at if report else None),
            'report_filename': (report.original_filename if report else None) or '-',
            'knowledge_details_url': (
                url_for('disputes_center.view_fap_contestation_report', report_id=item.report_id)
                if report else None
            ),
            'sort_datetime': item.publication_datetime or item.transmission_datetime or item.created_at,
        })

    for item in manual_history_items:
        performer_name = (item.performed_by_user.name if item.performed_by_user else '') or 'Usuário não identificado'
        events.append({
            'event_type': 'manual_history',
            'history_id': item.id,
            'report_id': None,
            'knowledge_base_id': None,
            'action': item.action,
            'manual_action_label': 'Edição manual',
            'manual_description': 'Status da 1ª instância alterado na tela de edição.',
            'performed_by': performer_name,
            'old_first_instance_status': item.old_first_instance_status,
            'new_first_instance_status': item.new_first_instance_status,
            'old_first_instance_status_label': _status_label_pt(item.old_first_instance_status),
            'new_first_instance_status_label': _status_label_pt(item.new_first_instance_status),
            'notes': item.notes,
            'transmission_datetime': None,
            'publication_datetime': None,
            'created_at': _format_datetime(item.created_at),
            'report_uploaded_at': None,
            'report_filename': None,
            'knowledge_details_url': None,
            'sort_datetime': item.created_at,
        })

    events.sort(key=lambda e: e.get('sort_datetime') or datetime.min, reverse=True)
    for e in events:
        e.pop('sort_datetime', None)

    return jsonify({
        'turnover_rate_id': tr.id,
        'employer_cnpj': tr.employer_cnpj,
        'year': tr.year,
        'events': events,
    })
