"""
Blueprint: Painel FAP
Prefixo: /fap-panel

Rotas:
    GET  /fap-panel/sync            — Página de sincronização de contestações
    POST /fap-panel/sync/run-year   — AJAX: sincroniza uma empresa + ano específico
    GET  /fap-panel/import          — Redireciona para a página de importação automática existente
"""

from __future__ import annotations

import json
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from functools import wraps
from io import BytesIO

from flask import (
    Blueprint,
    current_app,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.models import FapAutoImportedContestacao, FapCompany, FapWebContestacao, FapWebProcuracao, db
from app.services.fap_web_service import FapWebAuthPayload, FapWebService

fap_panel_bp = Blueprint('fap_panel', __name__, url_prefix='/fap-panel')

# ---------------------------------------------------------------------------
# Auth helpers (mesmos do disputes_center)
# ---------------------------------------------------------------------------

def get_current_law_firm_id():
    return session.get('law_firm_id')


def require_law_firm(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not get_current_law_firm_id():
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Anos de vigência disponíveis (mesmos da importação: 2010 até ano atual)
# ---------------------------------------------------------------------------

FAP_AVAILABLE_YEARS = list(range(datetime.now().year, 2009, -1))


# ---------------------------------------------------------------------------
# Rotas
# ---------------------------------------------------------------------------

@fap_panel_bp.route('/')
@require_law_firm
def index():
    return redirect(url_for('fap_panel.sync_page'))


@fap_panel_bp.route('/import')
@require_law_firm
def import_page():
    """Redireciona para a importação automática existente no disputes_center."""
    return redirect(url_for('disputes_center.fap_auto_import'))


@fap_panel_bp.route('/sync')
@require_law_firm
def sync_page():
    law_firm_id = get_current_law_firm_id()
    companies = (
        FapCompany.query
        .filter_by(law_firm_id=law_firm_id)
        .order_by(FapCompany.nome)
        .all()
    )

    saved_auth = session.get('fap_auto_import_auth', '')

    # Resumo rápido de contestações já sincronizadas
    total_synced = FapWebContestacao.query.filter_by(law_firm_id=law_firm_id).count()
    total_procuracoes = FapWebProcuracao.query.filter_by(law_firm_id=law_firm_id).count()

    return render_template(
        'fap_panel/sync.html',
        companies=companies,
        years=FAP_AVAILABLE_YEARS,
        saved_auth=saved_auth,
        total_synced=total_synced,
        total_procuracoes=total_procuracoes,
    )


@fap_panel_bp.route('/sync/run-year', methods=['POST'])
@require_law_firm
def sync_run_year():
    """AJAX — Busca contestações de uma empresa em um ano e persiste no banco.

    Body JSON:
        { "cnpj": "12345678", "year": 2023 }

    Response JSON:
        {
            "ok": true,
            "year": 2023,
            "total": 10,
            "created": 7,
            "updated": 3
        }
    """
    law_firm_id = get_current_law_firm_id()

    data = request.get_json(silent=True) or {}
    cnpj_raiz = str(data.get('cnpj') or '').strip()
    year = data.get('year')

    if not cnpj_raiz or not year:
        return jsonify({'ok': False, 'message': 'Informe o CNPJ e o ano de vigência.'}), 400

    saved_auth = session.get('fap_auto_import_auth', '')
    if not saved_auth:
        return jsonify({
            'ok': False,
            'message': 'Dados de autenticação não encontrados. Salve a sessão FAP primeiro.',
        }), 400

    try:
        auth = FapWebAuthPayload.from_json(saved_auth)
    except Exception:
        return jsonify({'ok': False, 'message': 'Dados de autenticação inválidos na sessão.'}), 400

    # Busca na API FAP
    result = FapWebService(auth).fetch_contestacoes(cnpj=cnpj_raiz, year=year)
    if not result.ok:
        detail = (result.data or {}).get('detail', '') if result.data else ''
        payload = {'ok': False, 'message': result.message}
        if detail:
            payload['detail'] = detail
        return jsonify(payload), 502

    items = result.data
    year_int = int(year)

    # Resolve fap_company_id pelo cnpj_raiz (pode ser nulo se não sincronizou empresas ainda)
    fap_company = FapCompany.query.filter_by(law_firm_id=law_firm_id, cnpj=cnpj_raiz).first()
    fap_company_id = fap_company.id if fap_company else None

    created = 0
    updated = 0
    now = datetime.utcnow()

    try:
        for item in items:
            cid = item.get('id')
            if not cid:
                continue

            cnpj_full = str(item.get('cnpj') or '').strip()
            if not cnpj_full:
                cnpj_full = cnpj_raiz

            # Garante string com 14 dígitos
            cnpj_digits = ''.join(ch for ch in cnpj_full if ch.isdigit())
            cnpj_full_14 = cnpj_digits.zfill(14) if len(cnpj_digits) <= 14 else cnpj_digits

            instancia   = item.get('instancia') or {}
            situacao    = item.get('situacao') or {}

            # Normaliza data de transmissão
            raw_dt = item.get('dataTransmissao')
            data_transmissao = None
            if raw_dt:
                try:
                    # Formato ISO: "2024-11-13T00:00:00" ou com timezone
                    data_transmissao = datetime.fromisoformat(
                        raw_dt.replace('Z', '+00:00').split('+')[0]
                    )
                except Exception:
                    pass

            existing = FapWebContestacao.query.filter_by(
                law_firm_id=law_firm_id,
                contestacao_id=int(cid),
                cnpj_raiz=cnpj_digits[:8],
            ).first()

            if existing:
                existing.cnpj                = cnpj_full_14
                existing.cnpj_raiz           = cnpj_digits[:8]
                existing.ano_vigencia        = year_int
                existing.fap_company_id      = fap_company_id
                existing.instancia_codigo    = instancia.get('codigo')
                existing.instancia_descricao = instancia.get('descricao')
                existing.situacao_codigo     = situacao.get('codigo')
                existing.situacao_descricao  = situacao.get('descricao')
                existing.protocolo           = item.get('protocolo')
                existing.data_transmissao    = data_transmissao
                existing.raw_data            = json.dumps(item, ensure_ascii=False)
                existing.last_synced_at      = now
                updated += 1
            else:
                rec = FapWebContestacao(
                    law_firm_id         = law_firm_id,
                    fap_company_id      = fap_company_id,
                    contestacao_id      = int(cid),
                    cnpj                = cnpj_full_14,
                    cnpj_raiz           = cnpj_digits[:8],
                    ano_vigencia        = year_int,
                    instancia_codigo    = instancia.get('codigo'),
                    instancia_descricao = instancia.get('descricao'),
                    situacao_codigo     = situacao.get('codigo'),
                    situacao_descricao  = situacao.get('descricao'),
                    protocolo           = item.get('protocolo'),
                    data_transmissao    = data_transmissao,
                    raw_data            = json.dumps(item, ensure_ascii=False),
                    last_synced_at      = now,
                )
                db.session.add(rec)
                created += 1

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Erro ao salvar no banco: {str(e)}'}), 500

    return jsonify({
        'ok': True,
        'year': year_int,
        'total': len(items),
        'created': created,
        'updated': updated,
    })


@fap_panel_bp.route('/sync/download-batch', methods=['POST'])
@require_law_firm
def sync_download_batch():
    """AJAX — Consulta o banco e baixa o próximo lote de PDFs sem file_path para um CNPJ.

    Body JSON:
        { "cnpj": "12345678" }

    Response JSON:
        { "ok": true, "total": 50, "remaining": 20, "processed": 30,
          "results": [{rec_id, ok, filename?, error?}, ...] }

    Chamar em loop até remaining == 0 ou processed == 0.
    """
    law_firm_id = get_current_law_firm_id()

    data = request.get_json(silent=True) or {}
    cnpj_raiz = str(data.get('cnpj') or '').strip()

    if not cnpj_raiz:
        return jsonify({'ok': False, 'message': 'Informe o CNPJ.'}), 400

    saved_auth = session.get('fap_auto_import_auth', '')
    if not saved_auth:
        return jsonify({'ok': False, 'message': 'Dados de autenticação não encontrados.'}), 400

    try:
        auth = FapWebAuthPayload.from_json(saved_auth)
    except Exception:
        return jsonify({'ok': False, 'message': 'Dados de autenticação inválidos.'}), 400

    pending_q = (
        FapWebContestacao.query
        .filter_by(law_firm_id=law_firm_id)
        .filter(FapWebContestacao.file_path.is_(None))
        .filter(FapWebContestacao.cnpj.like(cnpj_raiz + '%'))
    )

    total_before = pending_q.count()

    if total_before == 0:
        return jsonify({'ok': True, 'total': 0, 'remaining': 0, 'processed': 0, 'results': []})

    recs = (
        pending_q
        .with_entities(
            FapWebContestacao.id,
            FapWebContestacao.contestacao_id,
            FapWebContestacao.cnpj,
            FapWebContestacao.ano_vigencia,
        )
        .limit(30)
        .all()
    )

    upload_root = os.path.join(current_app.root_path, 'uploads', 'fap_web_contestacoes', str(law_firm_id))
    flask_app = current_app._get_current_object()

    def _download_one(rec):
        svc = FapWebService(auth)
        try:
            dl = svc.download_contestacao(
                year=rec.ano_vigencia,
                cnpj=rec.cnpj,
                contestacao_id=rec.contestacao_id,
            )
            if not dl.ok:
                return {'rec_id': rec.id, 'ok': False, 'error': dl.message}

            pdf_bytes = dl.data['pdf_bytes']
            filename  = f"{rec.contestacao_id}_{dl.data['filename']}"
            save_dir  = os.path.join(upload_root, str(rec.ano_vigencia), rec.cnpj)
            os.makedirs(save_dir, exist_ok=True)

            with open(os.path.join(save_dir, filename), 'wb') as f:
                f.write(pdf_bytes)

            rel_path = '/'.join([
                'uploads', 'fap_web_contestacoes',
                str(law_firm_id), str(rec.ano_vigencia), rec.cnpj, filename,
            ])

            with flask_app.app_context():
                db_rec = db.session.get(FapWebContestacao, rec.id)
                if db_rec:
                    db_rec.file_path = rel_path
                    db.session.commit()

            return {'rec_id': rec.id, 'ok': True, 'filename': filename}
        except Exception as e:
            try:
                with flask_app.app_context():
                    db.session.rollback()
            except Exception:
                pass
            return {'rec_id': rec.id, 'ok': False, 'error': str(e)}

    results = []
    with ThreadPoolExecutor(max_workers=min(len(recs), 20)) as pool:
        futures = {pool.submit(_download_one, r): r for r in recs}
        for fut in as_completed(futures):
            results.append(fut.result())

    remaining = pending_q.count()

    return jsonify({
        'ok': True,
        'total': total_before,
        'remaining': remaining,
        'processed': len(recs),
        'results': results,
    })


@fap_panel_bp.route('/contestacoes/<int:rec_id>/file')
@require_law_firm
def serve_contestacao_file(rec_id):
    """Serve o arquivo PDF local de uma contestação sincronizada."""
    from flask import abort, send_file as _send_file

    law_firm_id = get_current_law_firm_id()
    rec = FapWebContestacao.query.filter_by(id=rec_id, law_firm_id=law_firm_id).first_or_404()

    if not rec.file_path:
        abort(404)

    abs_path = os.path.abspath(os.path.join(current_app.root_path, rec.file_path))
    if not os.path.isfile(abs_path):
        abort(404)

    inline = request.args.get('inline') == '1'
    filename = os.path.basename(abs_path)
    # Remove o prefixo {contestacao_id}_ para nome amigável no download
    parts = filename.split('_', 1)
    display_name = parts[1] if len(parts) == 2 and parts[0].isdigit() else filename

    return _send_file(
        abs_path,
        mimetype='application/pdf',
        as_attachment=not inline,
        download_name=display_name,
    )


@fap_panel_bp.route('/contestacoes')
@require_law_firm
def contestacoes_page():
    """Página de visualização das contestações sincronizadas — estilo portal FAP."""
    from collections import defaultdict

    law_firm_id = get_current_law_firm_id()

    companies = (
        FapCompany.query
        .filter_by(law_firm_id=law_firm_id)
        .order_by(FapCompany.nome)
        .all()
    )

    # ── Filtros ──────────────────────────────────────────────────────────
    f_year      = request.args.get('ano_vigencia', '').strip()
    f_cnpj_raiz = request.args.get('cnpj_raiz', '').strip()
    f_cnpj      = request.args.get('cnpj', '').strip()          # estabelecimento (14 dígitos)
    f_instancia = request.args.get('instancia', '').strip()
    f_situacao  = request.args.get('situacao', '').strip()
    f_protocolo = request.args.get('protocolo', '').strip()
    f_prazo2    = request.args.get('prazo_2_instancia') == '1'

    query = FapWebContestacao.query.filter_by(law_firm_id=law_firm_id)

    if f_year:
        query = query.filter(FapWebContestacao.ano_vigencia == int(f_year))
    if f_cnpj_raiz:
        query = query.filter(FapWebContestacao.cnpj_raiz == f_cnpj_raiz)
    if f_cnpj:
        query = query.filter(FapWebContestacao.cnpj == f_cnpj)
    if f_instancia:
        query = query.filter(FapWebContestacao.instancia_codigo == f_instancia)
    if f_situacao:
        query = query.filter(FapWebContestacao.situacao_codigo == f_situacao)
    if f_protocolo:
        query = query.filter(FapWebContestacao.protocolo.ilike(f'%{f_protocolo}%'))

    # ── Mapa de CNPJs (14 dígitos) por raiz para o filtro de estabelecimento ──
    cnpj_rows = (
        FapWebContestacao.query
        .filter_by(law_firm_id=law_firm_id)
        .with_entities(FapWebContestacao.cnpj_raiz, FapWebContestacao.cnpj)
        .distinct()
        .all()
    )
    cnpjs_by_raiz: dict[str, list[str]] = {}
    for raiz, cnpj_full in cnpj_rows:
        if raiz and cnpj_full:
            cnpjs_by_raiz.setdefault(raiz, [])
            if cnpj_full not in cnpjs_by_raiz[raiz]:
                cnpjs_by_raiz[raiz].append(cnpj_full)
    for raiz in cnpjs_by_raiz:
        cnpjs_by_raiz[raiz].sort()

    all_rows = query.order_by(
        FapWebContestacao.ano_vigencia.desc(),
        FapWebContestacao.cnpj.asc(),
    ).all()

    # ── Mapa de contestações já importadas (contestacao_id → report_id) ──
    contestacao_ids = [r.contestacao_id for r in all_rows]
    imported_map = {}
    if contestacao_ids:
        imported_rows = (
            FapAutoImportedContestacao.query
            .filter(
                FapAutoImportedContestacao.law_firm_id == law_firm_id,
                FapAutoImportedContestacao.contestacao_id.in_(contestacao_ids),
            )
            .all()
        )
        for imp in imported_rows:
            imported_map[imp.contestacao_id] = imp.report_id

    # ── Valores únicos para os filtros de instância e situação ───────────
    distinct = FapWebContestacao.query.filter_by(law_firm_id=law_firm_id)
    instancias = {(r.instancia_codigo, r.instancia_descricao)
                  for r in distinct.with_entities(
                      FapWebContestacao.instancia_codigo,
                      FapWebContestacao.instancia_descricao).distinct()
                  if r.instancia_codigo}
    situacoes = {(r.situacao_codigo, r.situacao_descricao)
                 for r in distinct.with_entities(
                     FapWebContestacao.situacao_codigo,
                     FapWebContestacao.situacao_descricao).distinct()
                 if r.situacao_codigo}

    # ── Agrupamento por (ano_vigencia, cnpj) para montar a tabela ────────
    # Cada célula contém a lista de contestações naquela categoria.
    # Classificação:
    #   instancia_codigo contém "SEGUNDA" → recurso (cols 5,6); senão → contestação (cols 3,4)
    #   "em andamento" = situacao que contém "ANDAMENTO" ou não contém "TRANSMIT"/"RESULT"/"DIVULG"
    def _is_segunda(cod):
        return cod and 'SEGUNDA' in cod.upper()

    def _is_em_andamento(cod, desc):
        s = ((cod or '') + ' ' + (desc or '')).upper()
        return ('ANDAMENTO' in s or 'PRAZO' in s) and 'TRANSMIT' not in s and 'RESULT' not in s and 'DIVULG' not in s

    grouped = defaultdict(lambda: {'c_and1': [], 'c_tra1': [], 'c_and2': [], 'c_tra2': []})

    for r in all_rows:
        key = (r.ano_vigencia, r.cnpj)
        is2 = _is_segunda(r.instancia_codigo)
        em_and = _is_em_andamento(r.situacao_codigo, r.situacao_descricao)
        if is2:
            grouped[key]['c_and2' if em_and else 'c_tra2'].append(r)
        else:
            grouped[key]['c_and1' if em_and else 'c_tra1'].append(r)

    # ── Formata CNPJ 14 dígitos para exibição ────────────────────────────
    def _fmt_cnpj(digits):
        s = (digits or '').zfill(14)
        if len(s) == 14:
            return f'{s[:2]}.{s[2:5]}.{s[5:8]}/{s[8:12]}-{s[12:]}'
        return digits

    table_rows = []
    for (ano, cnpj_raw), cells in sorted(grouped.items(), key=lambda x: (-x[0][0], x[0][1])):
        table_rows.append({
            'ano_vigencia': ano,
            'cnpj_raw':     cnpj_raw,
            'cnpj_fmt':     _fmt_cnpj(cnpj_raw),
            'c_and1':       cells['c_and1'],
            'c_tra1':       cells['c_tra1'],
            'c_and2':       cells['c_and2'],
            'c_tra2':       cells['c_tra2'],
        })

    return render_template(
        'fap_panel/contestacoes.html',
        companies=companies,
        years=FAP_AVAILABLE_YEARS,
        table_rows=table_rows,
        total=len(all_rows),
        instancias=sorted(instancias),
        situacoes=sorted(situacoes),
        imported_map=imported_map,
        cnpjs_by_raiz=cnpjs_by_raiz,
        # filtros ativos (para repreencher o form)
        f_year=f_year,
        f_cnpj_raiz=f_cnpj_raiz,
        f_cnpj=f_cnpj,
        f_instancia=f_instancia,
        f_situacao=f_situacao,
        f_protocolo=f_protocolo,
        f_prazo2=f_prazo2,
    )


@fap_panel_bp.route('/contestacoes/export-excel')
@require_law_firm
def contestacoes_export_excel():
    """Exporta as contestações filtradas para Excel (sem paginação) — planilha elaborada."""
    law_firm_id = get_current_law_firm_id()

    f_year      = request.args.get('ano_vigencia', '').strip()
    f_cnpj_raiz = request.args.get('cnpj_raiz', '').strip()
    f_cnpj      = request.args.get('cnpj', '').strip()
    f_instancia = request.args.get('instancia', '').strip()
    f_situacao  = request.args.get('situacao', '').strip()
    f_protocolo = request.args.get('protocolo', '').strip()

    query = FapWebContestacao.query.filter_by(law_firm_id=law_firm_id)
    if f_year:
        query = query.filter(FapWebContestacao.ano_vigencia == int(f_year))
    if f_cnpj_raiz:
        query = query.filter(FapWebContestacao.cnpj_raiz == f_cnpj_raiz)
    if f_cnpj:
        query = query.filter(FapWebContestacao.cnpj == f_cnpj)
    if f_instancia:
        query = query.filter(FapWebContestacao.instancia_codigo == f_instancia)
    if f_situacao:
        query = query.filter(FapWebContestacao.situacao_codigo == f_situacao)
    if f_protocolo:
        query = query.filter(FapWebContestacao.protocolo.ilike(f'%{f_protocolo}%'))

    rows = query.order_by(
        FapWebContestacao.ano_vigencia.desc(),
        FapWebContestacao.cnpj.asc(),
        FapWebContestacao.contestacao_id.asc(),
    ).all()

    # Mapa de importados
    contestacao_ids = [r.contestacao_id for r in rows]
    imported_map: dict[int, int] = {}
    if contestacao_ids:
        from app.models import FapAutoImportedContestacao
        for imp in FapAutoImportedContestacao.query.filter(
            FapAutoImportedContestacao.law_firm_id == law_firm_id,
            FapAutoImportedContestacao.contestacao_id.in_(contestacao_ids),
        ).all():
            imported_map[imp.contestacao_id] = imp.report_id

    def _fmt_cnpj(digits):
        s = (digits or '').zfill(14)
        if len(s) == 14:
            return f'{s[:2]}.{s[2:5]}.{s[5:8]}/{s[8:12]}-{s[12:]}'
        return digits

    # ── Estilos ──────────────────────────────────────────────────────────────
    # Título principal
    title_font   = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    title_fill   = PatternFill('solid', fgColor='0D47A1')
    title_align  = Alignment(horizontal='center', vertical='center')

    # Bloco de filtros (label + valor)
    filter_label_font = Font(name='Calibri', size=10, bold=True, color='1A237E')
    filter_val_font   = Font(name='Calibri', size=10, color='1A237E')
    filter_fill       = PatternFill('solid', fgColor='E8EAF6')

    # Cabeçalho de grupo (1ª inst / 2ª inst)
    group1_font  = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    group1_fill  = PatternFill('solid', fgColor='1565C0')
    group2_font  = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    group2_fill  = PatternFill('solid', fgColor='2E7D32')

    # Cabeçalho de coluna
    col_header_font  = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    col_header_fill  = PatternFill('solid', fgColor='1976D2')
    col_header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Bordas finas
    thin = Side(style='thin', color='B0BEC5')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Dados: zebra
    data_fill_even = PatternFill('solid', fgColor='FAFAFA')
    data_fill_odd  = PatternFill('solid', fgColor='EEF2FF')
    data_font      = Font(name='Calibri', size=10)
    data_align_ctr = Alignment(horizontal='center', vertical='center')
    data_align_lft = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # Importado: sim/não
    yes_font  = Font(name='Calibri', size=10, bold=True, color='1B5E20')
    yes_fill  = PatternFill('solid', fgColor='C8E6C9')
    no_font   = Font(name='Calibri', size=10, bold=True, color='B71C1C')
    no_fill   = PatternFill('solid', fgColor='FFCDD2')

    # ── Workbook ─────────────────────────────────────────────────────────────
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Contestações FAP'

    num_cols = 14

    # ── Linha 1: Título ───────────────────────────────────────────────────────
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = sheet.cell(row=1, column=1,
                            value=f'Contestações FAP — Exportado em {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    title_cell.font   = title_font
    title_cell.fill   = title_fill
    title_cell.alignment = title_align
    sheet.row_dimensions[1].height = 28

    # ── Bloco de filtros (linhas 2–5) ─────────────────────────────────────────
    filter_pairs = [
        ('Vigência',    f_year      or 'Todas'),
        ('CNPJ Raiz',   _fmt_cnpj(f_cnpj_raiz.zfill(8)) if f_cnpj_raiz else 'Todos'),
        ('CNPJ',        _fmt_cnpj(f_cnpj) if f_cnpj else 'Todos'),
        ('Instância',   f_instancia or 'Todas'),
        ('Situação',    f_situacao  or 'Todas'),
        ('Protocolo',   f_protocolo or 'Todos'),
        ('Total linhas', str(len(rows))),
    ]

    # Distribui em 2 colunas de filtro por linha (label | valor | label | valor …)
    # Cada par ocupa 2 colunas; 7 pares → 4 linhas (3 pares + 3 pares + 1 par)
    pairs_per_row = 3
    filter_start_row = 2
    for pair_idx, (lbl, val) in enumerate(filter_pairs):
        frow = filter_start_row + (pair_idx // pairs_per_row)
        fcol = 1 + (pair_idx % pairs_per_row) * 2

        lbl_cell = sheet.cell(row=frow, column=fcol, value=lbl + ':')
        lbl_cell.font      = filter_label_font
        lbl_cell.fill      = filter_fill
        lbl_cell.alignment = Alignment(horizontal='right', vertical='center')
        lbl_cell.border    = cell_border

        val_cell = sheet.cell(row=frow, column=fcol + 1, value=val)
        val_cell.font      = filter_val_font
        val_cell.fill      = filter_fill
        val_cell.alignment = Alignment(horizontal='left', vertical='center')
        val_cell.border    = cell_border

    # Preenche células vazias do bloco de filtros com fill
    filter_end_row = filter_start_row + (len(filter_pairs) - 1) // pairs_per_row
    for frow in range(filter_start_row, filter_end_row + 1):
        for fcol in range(pairs_per_row * 2 + 1, num_cols + 1):
            c = sheet.cell(row=frow, column=fcol)
            c.fill   = filter_fill
            c.border = cell_border
        sheet.row_dimensions[frow].height = 18

    # Linha separadora em branco
    sep_row = filter_end_row + 1
    for col in range(1, num_cols + 1):
        c = sheet.cell(row=sep_row, column=col)
        c.fill = PatternFill('solid', fgColor='BBDEFB')
    sheet.row_dimensions[sep_row].height = 6

    # ── Cabeçalhos de coluna ──────────────────────────────────────────────────
    header_row = sep_row + 1
    headers = [
        'ID Contestação',
        'Vigência',
        'CNPJ',
        'CNPJ Formatado',
        'CNPJ Raiz',
        'Instância\n(código)',
        'Instância\n(descrição)',
        'Situação\n(código)',
        'Situação\n(descrição)',
        'Protocolo',
        'Data\nTransmissão',
        'Importado?',
        'Última\nSincronização',
        'Link Visualizar',
    ]
    for col_idx, header in enumerate(headers, start=1):
        c = sheet.cell(row=header_row, column=col_idx, value=header)
        c.font      = col_header_font
        c.fill      = col_header_fill
        c.alignment = col_header_align
        c.border    = cell_border
    sheet.row_dimensions[header_row].height = 32

    # ── Dados ─────────────────────────────────────────────────────────────────
    data_start_row = header_row + 1
    for row_idx, r in enumerate(rows):
        xrow = data_start_row + row_idx
        is_odd = row_idx % 2 == 1
        row_fill = data_fill_odd if is_odd else data_fill_even

        cnpj14 = (r.cnpj or '').zfill(14)
        dl_url = url_for(
            'disputes_center.fap_auto_import_download_contestacao',
            year=r.ano_vigencia,
            cnpj=cnpj14,
            contestacao_id=r.contestacao_id,
            _external=True,
        ) + '?inline=1'

        importado = r.contestacao_id in imported_map
        values = [
            r.contestacao_id,
            r.ano_vigencia,
            r.cnpj or '',
            _fmt_cnpj(r.cnpj),
            r.cnpj_raiz or '',
            r.instancia_codigo or '',
            r.instancia_descricao or '',
            r.situacao_codigo or '',
            r.situacao_descricao or '',
            r.protocolo or '',
            r.data_transmissao.strftime('%d/%m/%Y') if r.data_transmissao else '',
            'Sim' if importado else 'Não',
            r.last_synced_at.strftime('%d/%m/%Y %H:%M') if r.last_synced_at else '',
            dl_url,
        ]

        for col_idx, val in enumerate(values, start=1):
            c = sheet.cell(row=xrow, column=col_idx, value=val)
            c.font   = data_font
            c.fill   = row_fill
            c.border = cell_border

            # Alinhamento por tipo
            if col_idx in (1, 2, 6, 8, 11, 13):
                c.alignment = data_align_ctr
            else:
                c.alignment = data_align_lft

            # Destaque Importado?
            if col_idx == 12:
                c.font  = yes_font if importado else no_font
                c.fill  = yes_fill if importado else no_fill
                c.alignment = data_align_ctr

        sheet.row_dimensions[xrow].height = 18

    # ── Auto-filter nos cabeçalhos ────────────────────────────────────────────
    sheet.auto_filter.ref = (
        f'A{header_row}:{get_column_letter(num_cols)}{header_row + len(rows)}'
    )

    # ── Congelar painéis abaixo do cabeçalho ─────────────────────────────────
    sheet.freeze_panes = f'A{data_start_row}'

    # ── Larguras das colunas ──────────────────────────────────────────────────
    # ID | Vig | CNPJ | CNPJFmt | Raiz | Inst(cod) | Inst(desc) | Sit(cod) | Sit(desc) | Protocolo | DataTrans | Importado | UltSync | Link
    col_widths = [16, 10, 20, 24, 14, 16, 44, 16, 46, 24, 16, 13, 22, 80]
    for idx, width in enumerate(col_widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    # ── Linha de totais ───────────────────────────────────────────────────────
    total_row = data_start_row + len(rows)
    total_fill = PatternFill('solid', fgColor='E3F2FD')
    total_font = Font(name='Calibri', size=10, bold=True, color='0D47A1')
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=num_cols - 1)
    tc = sheet.cell(row=total_row, column=1,
                    value=f'Total: {len(rows)} contestação(ões) exportada(s)')
    tc.font      = total_font
    tc.fill      = total_fill
    tc.alignment = Alignment(horizontal='center', vertical='center')
    tc.border    = cell_border
    for col in range(2, num_cols + 1):
        c = sheet.cell(row=total_row, column=col)
        c.fill   = total_fill
        c.border = cell_border
    sheet.row_dimensions[total_row].height = 18

    # ── Gerar arquivo ─────────────────────────────────────────────────────────
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'contestacoes_fap_{timestamp}.xlsx'

    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@fap_panel_bp.route('/contestacoes/export-excel-agrupado')
@require_law_firm
def contestacoes_export_excel_agrupado():
    """Exporta contestações agrupadas por (vigência, CNPJ) — espelha a listagem em tela."""
    from collections import defaultdict

    law_firm_id = get_current_law_firm_id()

    f_year      = request.args.get('ano_vigencia', '').strip()
    f_cnpj_raiz = request.args.get('cnpj_raiz', '').strip()
    f_cnpj      = request.args.get('cnpj', '').strip()
    f_instancia = request.args.get('instancia', '').strip()
    f_situacao  = request.args.get('situacao', '').strip()
    f_protocolo = request.args.get('protocolo', '').strip()

    query = FapWebContestacao.query.filter_by(law_firm_id=law_firm_id)
    if f_year:
        query = query.filter(FapWebContestacao.ano_vigencia == int(f_year))
    if f_cnpj_raiz:
        query = query.filter(FapWebContestacao.cnpj_raiz == f_cnpj_raiz)
    if f_cnpj:
        query = query.filter(FapWebContestacao.cnpj == f_cnpj)
    if f_instancia:
        query = query.filter(FapWebContestacao.instancia_codigo == f_instancia)
    if f_situacao:
        query = query.filter(FapWebContestacao.situacao_codigo == f_situacao)
    if f_protocolo:
        query = query.filter(FapWebContestacao.protocolo.ilike(f'%{f_protocolo}%'))

    all_rows = query.order_by(
        FapWebContestacao.ano_vigencia.desc(),
        FapWebContestacao.cnpj.asc(),
        FapWebContestacao.contestacao_id.asc(),
    ).all()

    # Mapa de importados
    contestacao_ids = [r.contestacao_id for r in all_rows]
    imported_set: set[int] = set()
    if contestacao_ids:
        from app.models import FapAutoImportedContestacao
        for imp in FapAutoImportedContestacao.query.filter(
            FapAutoImportedContestacao.law_firm_id == law_firm_id,
            FapAutoImportedContestacao.contestacao_id.in_(contestacao_ids),
        ).all():
            imported_set.add(imp.contestacao_id)

    # Mapa de empresas (raiz → nome)
    from app.models import FapCompany
    company_names: dict[str, str] = {
        c.cnpj: (c.nome or '') for c in FapCompany.query.filter_by(law_firm_id=law_firm_id).all()
    }

    def _fmt_cnpj(digits):
        s = (digits or '').zfill(14)
        if len(s) == 14:
            return f'{s[:2]}.{s[2:5]}.{s[5:8]}/{s[8:12]}-{s[12:]}'
        return digits

    def _is_segunda(cod):
        return cod and 'SEGUNDA' in cod.upper()

    def _is_em_andamento(cod, desc):
        s = ((cod or '') + ' ' + (desc or '')).upper()
        return ('ANDAMENTO' in s or 'PRAZO' in s) and 'TRANSMIT' not in s and 'RESULT' not in s and 'DIVULG' not in s

    grouped = defaultdict(lambda: {'c_and1': [], 'c_tra1': [], 'c_and2': [], 'c_tra2': []})
    for r in all_rows:
        key = (r.ano_vigencia, r.cnpj)
        is2 = _is_segunda(r.instancia_codigo)
        em_and = _is_em_andamento(r.situacao_codigo, r.situacao_descricao)
        if is2:
            grouped[key]['c_and2' if em_and else 'c_tra2'].append(r)
        else:
            grouped[key]['c_and1' if em_and else 'c_tra1'].append(r)

    def _cell_summary(contest_list):
        qtd   = len(contest_list)
        protos = ' | '.join(c.protocolo or '—' for c in contest_list)
        sits   = ' | '.join(c.situacao_descricao or '' for c in contest_list)
        return qtd, protos, sits

    # ── Estilos ───────────────────────────────────────────────────────────────
    thin = Side(style='thin', color='B0BEC5')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font  = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    title_fill  = PatternFill('solid', fgColor='0D47A1')
    title_align = Alignment(horizontal='center', vertical='center')

    filter_label_font = Font(name='Calibri', size=10, bold=True, color='1A237E')
    filter_val_font   = Font(name='Calibri', size=10, color='1A237E')
    filter_fill       = PatternFill('solid', fgColor='E8EAF6')

    grp1_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    grp1_fill = PatternFill('solid', fgColor='1565C0')
    grp2_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    grp2_fill = PatternFill('solid', fgColor='2E7D32')

    sub_and_fill = PatternFill('solid', fgColor='F57F17')
    sub_and_font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
    sub_tra_fill = PatternFill('solid', fgColor='388E3C')
    sub_tra_font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')

    base_hdr_fill = PatternFill('solid', fgColor='37474F')
    base_hdr_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    hdr_align     = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_font      = Font(name='Calibri', size=9)
    data_align_ctr = Alignment(horizontal='center', vertical='center')
    data_align_lft = Alignment(horizontal='left', vertical='center', wrap_text=True)
    data_fill_even = PatternFill('solid', fgColor='FAFAFA')
    data_fill_odd  = PatternFill('solid', fgColor='EEF2FF')
    num_fill_and   = PatternFill('solid', fgColor='FFF8E1')
    num_fill_tra   = PatternFill('solid', fgColor='E8F5E9')
    num_font_and   = Font(name='Calibri', size=11, bold=True, color='E65100')
    num_font_tra   = Font(name='Calibri', size=11, bold=True, color='1B5E20')
    zero_font      = Font(name='Calibri', size=11, bold=False, color='B0BEC5')
    total_fill     = PatternFill('solid', fgColor='E3F2FD')
    total_font_s   = Font(name='Calibri', size=10, bold=True, color='0D47A1')

    # ── Workbook ──────────────────────────────────────────────────────────────
    # Colunas:
    #   A  Vigência | B CNPJ | C CNPJ Fmt | D Nome
    #   E  1ªAnd Qtd | F 1ªAnd Protocolos | G 1ªAnd Situações
    #   H  1ªTra Qtd | I 1ªTra Protocolos | J 1ªTra Situações
    #   K  2ªAnd Qtd | L 2ªAnd Protocolos
    #   M  2ªTra Qtd | N 2ªTra Protocolos
    num_cols = 14

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Contestações FAP (Agrupado)'

    # Linha 1: título
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    tc = sheet.cell(row=1, column=1,
                    value=f'Contestações FAP por Estabelecimento — {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    tc.font = title_font; tc.fill = title_fill; tc.alignment = title_align
    sheet.row_dimensions[1].height = 28

    # Linhas 2-4: filtros
    filter_pairs = [
        ('Vigência',          f_year      or 'Todas'),
        ('CNPJ Raiz',         _fmt_cnpj(f_cnpj_raiz.zfill(8)) if f_cnpj_raiz else 'Todos'),
        ('CNPJ',              _fmt_cnpj(f_cnpj) if f_cnpj else 'Todos'),
        ('Instância',         f_instancia or 'Todas'),
        ('Situação',          f_situacao  or 'Todas'),
        ('Protocolo',         f_protocolo or 'Todos'),
        ('Linhas exportadas', str(len(grouped))),
    ]
    pairs_per_row = 3
    filter_start_row = 2
    for pair_idx, (lbl, val) in enumerate(filter_pairs):
        frow = filter_start_row + (pair_idx // pairs_per_row)
        fcol = 1 + (pair_idx % pairs_per_row) * 2
        lbl_c = sheet.cell(row=frow, column=fcol, value=lbl + ':')
        lbl_c.font = filter_label_font; lbl_c.fill = filter_fill
        lbl_c.alignment = Alignment(horizontal='right', vertical='center'); lbl_c.border = cell_border
        val_c = sheet.cell(row=frow, column=fcol + 1, value=val)
        val_c.font = filter_val_font; val_c.fill = filter_fill
        val_c.alignment = Alignment(horizontal='left', vertical='center'); val_c.border = cell_border

    filter_end_row = filter_start_row + (len(filter_pairs) - 1) // pairs_per_row
    for frow in range(filter_start_row, filter_end_row + 1):
        for fcol in range(pairs_per_row * 2 + 1, num_cols + 1):
            c = sheet.cell(row=frow, column=fcol)
            c.fill = filter_fill; c.border = cell_border
        sheet.row_dimensions[frow].height = 18

    sep_row = filter_end_row + 1
    for col in range(1, num_cols + 1):
        sheet.cell(row=sep_row, column=col).fill = PatternFill('solid', fgColor='BBDEFB')
    sheet.row_dimensions[sep_row].height = 6

    # Cabeçalho duplo
    grp_row = sep_row + 1
    sub_row = grp_row + 1

    for col, label in [(1, 'Vigência'), (2, 'CNPJ'), (3, 'CNPJ Formatado'), (4, 'Nome Empresa')]:
        sheet.merge_cells(start_row=grp_row, start_column=col, end_row=sub_row, end_column=col)
        c = sheet.cell(row=grp_row, column=col, value=label)
        c.font = base_hdr_font; c.fill = base_hdr_fill; c.alignment = hdr_align; c.border = cell_border

    sheet.merge_cells(start_row=grp_row, start_column=5, end_row=grp_row, end_column=10)
    g1 = sheet.cell(row=grp_row, column=5, value='1ª Instância — Contestações')
    g1.font = grp1_font; g1.fill = grp1_fill; g1.alignment = hdr_align; g1.border = cell_border

    sheet.merge_cells(start_row=grp_row, start_column=11, end_row=grp_row, end_column=14)
    g2 = sheet.cell(row=grp_row, column=11, value='2ª Instância — Recursos')
    g2.font = grp2_font; g2.fill = grp2_fill; g2.alignment = hdr_align; g2.border = cell_border

    sub_headers = [
        (5,  'Em Andamento\n(Qtd)',        sub_and_fill, sub_and_font),
        (6,  'Em Andamento\n(Protocolos)', sub_and_fill, sub_and_font),
        (7,  'Em Andamento\n(Situações)',  sub_and_fill, sub_and_font),
        (8,  'Transmitidas\n(Qtd)',        sub_tra_fill, sub_tra_font),
        (9,  'Transmitidas\n(Protocolos)', sub_tra_fill, sub_tra_font),
        (10, 'Transmitidas\n(Situações)',  sub_tra_fill, sub_tra_font),
        (11, 'Em Andamento\n(Qtd)',        sub_and_fill, sub_and_font),
        (12, 'Em Andamento\n(Protocolos)', sub_and_fill, sub_and_font),
        (13, 'Transmitidos\n(Qtd)',        sub_tra_fill, sub_tra_font),
        (14, 'Transmitidos\n(Protocolos)', sub_tra_fill, sub_tra_font),
    ]
    for col, label, sfill, sfont in sub_headers:
        c = sheet.cell(row=sub_row, column=col, value=label)
        c.font = sfont; c.fill = sfill; c.alignment = hdr_align; c.border = cell_border

    sheet.row_dimensions[grp_row].height = 22
    sheet.row_dimensions[sub_row].height = 30

    # Dados
    data_start = sub_row + 1
    sorted_groups = sorted(grouped.items(), key=lambda x: (-x[0][0], x[0][1]))

    for row_idx, ((ano, cnpj_raw), cells) in enumerate(sorted_groups):
        xrow = data_start + row_idx
        dfill = data_fill_odd if row_idx % 2 else data_fill_even

        cnpj14 = (cnpj_raw or '').zfill(14)
        raiz8  = cnpj14[:8] if len(cnpj14) == 14 else (cnpj_raw or '')[:8]
        nome   = company_names.get(raiz8, '')

        and1_qtd, and1_pro, and1_sit = _cell_summary(cells['c_and1'])
        tra1_qtd, tra1_pro, tra1_sit = _cell_summary(cells['c_tra1'])
        and2_qtd, and2_pro, _        = _cell_summary(cells['c_and2'])
        tra2_qtd, tra2_pro, _        = _cell_summary(cells['c_tra2'])

        row_values = [
            ano, cnpj_raw or '', _fmt_cnpj(cnpj_raw), nome,
            and1_qtd, and1_pro, and1_sit,
            tra1_qtd, tra1_pro, tra1_sit,
            and2_qtd, and2_pro,
            tra2_qtd, tra2_pro,
        ]

        for col_idx, val in enumerate(row_values, start=1):
            c = sheet.cell(row=xrow, column=col_idx, value=val)
            c.border = cell_border
            if col_idx in (5, 11):
                c.fill = num_fill_and
                c.font = num_font_and if val else zero_font
                c.alignment = data_align_ctr
            elif col_idx in (8, 13):
                c.fill = num_fill_tra
                c.font = num_font_tra if val else zero_font
                c.alignment = data_align_ctr
            elif col_idx in (1, 2):
                c.font = data_font; c.fill = dfill; c.alignment = data_align_ctr
            else:
                c.font = data_font; c.fill = dfill; c.alignment = data_align_lft

        # Altura da linha de dado ajustada dinamicamente pelo conteúdo mais longo
        max_lines = 1
        for val in [and1_pro, and1_sit, tra1_pro, tra1_sit, and2_pro, tra2_pro]:
            if val:
                max_lines = max(max_lines, val.count('|') + 1)
        sheet.row_dimensions[xrow].height = max(18, min(max_lines * 14, 60))

    sheet.auto_filter.ref = (
        f'A{sub_row}:{get_column_letter(num_cols)}{data_start + len(sorted_groups) - 1}'
    )
    sheet.freeze_panes = f'A{data_start}'

    # A=Vigência B=CNPJ C=CNPJFmt D=Nome E=And1Qtd F=And1Proto G=And1Sit H=Tra1Qtd I=Tra1Proto J=Tra1Sit K=And2Qtd L=And2Proto M=Tra2Qtd N=Tra2Proto
    col_widths = [10, 20, 22, 40, 8, 48, 48, 8, 48, 48, 8, 48, 8, 48]
    for idx, width in enumerate(col_widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    # Linha de totais
    total_row = data_start + len(sorted_groups)
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    tlbl = sheet.cell(row=total_row, column=1,
                      value=f'Total: {len(sorted_groups)} estabelecimento(s) · {len(all_rows)} contestação(ões)')
    tlbl.font = total_font_s; tlbl.fill = total_fill
    tlbl.alignment = Alignment(horizontal='center', vertical='center'); tlbl.border = cell_border

    for col_idx in (5, 8, 11, 13):
        vals = [sheet.cell(row=data_start + i, column=col_idx).value or 0 for i in range(len(sorted_groups))]
        sc = sheet.cell(row=total_row, column=col_idx, value=sum(vals))
        sc.font = total_font_s; sc.fill = total_fill
        sc.alignment = data_align_ctr; sc.border = cell_border

    for col_idx in range(1, num_cols + 1):
        c = sheet.cell(row=total_row, column=col_idx)
        if not c.value:
            c.fill = total_fill; c.border = cell_border
    sheet.row_dimensions[total_row].height = 20

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'contestacoes_fap_agrupado_{timestamp}.xlsx'

    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@fap_panel_bp.route('/sync/summary', methods=['GET'])
@require_law_firm
def sync_summary():
    """AJAX — Retorna um resumo das contestações sincronizadas por empresa/vigência."""
    law_firm_id = get_current_law_firm_id()
    cnpj_raiz = request.args.get('cnpj', '').strip()

    query = FapWebContestacao.query.filter_by(law_firm_id=law_firm_id)
    if cnpj_raiz:
        query = query.filter_by(cnpj_raiz=cnpj_raiz)

    rows = query.order_by(
        FapWebContestacao.cnpj,
        FapWebContestacao.ano_vigencia.desc(),
    ).all()

    items = [
        {
            'id':                  r.id,
            'contestacao_id':      r.contestacao_id,
            'cnpj':                r.cnpj,
            'cnpj_raiz':           r.cnpj_raiz,
            'ano_vigencia':        r.ano_vigencia,
            'instancia_codigo':    r.instancia_codigo or '',
            'instancia_descricao': r.instancia_descricao or '',
            'situacao_codigo':     r.situacao_codigo or '',
            'situacao_descricao':  r.situacao_descricao or '',
            'protocolo':           r.protocolo or '',
            'data_transmissao':    r.data_transmissao.strftime('%d/%m/%Y') if r.data_transmissao else '',
            'report_id':           r.report_id,
            'last_synced_at':      r.last_synced_at.strftime('%d/%m/%Y %H:%M') if r.last_synced_at else '',
        }
        for r in rows
    ]

    return jsonify({'ok': True, 'items': items, 'total': len(items)})


@fap_panel_bp.route('/sync/procuracoes', methods=['POST'])
@require_law_firm
def sync_procuracoes():
    """AJAX — Busca as procurações eletrônicas do portal FAP e persiste no banco.

    Response JSON:
        { "ok": true, "total": 42, "created": 10, "updated": 32 }
    """
    law_firm_id = get_current_law_firm_id()

    saved_auth = session.get('fap_auto_import_auth', '')
    if not saved_auth:
        return jsonify({
            'ok': False,
            'message': 'Dados de autenticação não encontrados. Salve a sessão FAP primeiro.',
        }), 400

    try:
        auth = FapWebAuthPayload.from_json(saved_auth)
    except Exception:
        return jsonify({'ok': False, 'message': 'Dados de autenticação inválidos na sessão.'}), 400

    result = FapWebService(auth).fetch_procuracoes()
    if not result.ok:
        expired = getattr(result, 'expired', False)
        payload = {'ok': False, 'message': result.message}
        if expired:
            payload['expired'] = True
        return jsonify(payload), 502

    items = result.data if isinstance(result.data, list) else []
    now = datetime.utcnow()
    created = 0
    updated = 0

    try:
        for item in items:
            protocolo = str(item.get('protocolo') or '').strip()
            if not protocolo:
                continue

            tipo   = item.get('tipoProcuracao') or {}
            sit    = item.get('situacao') or {}

            # Datas
            from datetime import date as _date
            def _parse_date(s):
                if not s:
                    return None
                try:
                    return _date.fromisoformat(s[:10])
                except Exception:
                    return None

            def _parse_datetime(s):
                if not s:
                    return None
                try:
                    return datetime.fromisoformat(s.replace('Z', '+00:00').split('+')[0])
                except Exception:
                    return None

            existing = FapWebProcuracao.query.filter_by(
                law_firm_id=law_firm_id,
                protocolo=protocolo,
            ).first()

            fields = dict(
                tipo_procuracao_codigo    = tipo.get('codigo'),
                tipo_procuracao_descricao = tipo.get('descricao'),
                situacao_codigo           = sit.get('codigo'),
                situacao_descricao        = sit.get('descricao'),
                data_inicio               = _parse_date(item.get('dataInicio')),
                data_fim                  = _parse_date(item.get('dataFim')),
                cnpj_raiz_outorgante      = str(item['cnpjRaizOutorgante']) if item.get('cnpjRaizOutorgante') is not None else None,
                nome_empresa_outorgante   = item.get('nomeEmpresaOutorgante'),
                cpf_outorgado             = str(item['cpfOutorgado']) if item.get('cpfOutorgado') is not None else None,
                cnpj_raiz_outorgado       = str(item['cnpjRaizOutorgado']) if item.get('cnpjRaizOutorgado') is not None else None,
                data_cadastro             = _parse_datetime(item.get('dataCadastro')),
                raw_data                  = json.dumps(item, ensure_ascii=False),
                last_synced_at            = now,
            )

            if existing:
                for k, v in fields.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                rec = FapWebProcuracao(
                    law_firm_id=law_firm_id,
                    protocolo=protocolo,
                    **fields,
                )
                db.session.add(rec)
                created += 1

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Erro ao salvar no banco: {str(e)}'}), 500

    return jsonify({'ok': True, 'total': len(items), 'created': created, 'updated': updated})


@fap_panel_bp.route('/sync/procuracoes/list', methods=['GET'])
@require_law_firm
def sync_procuracoes_list():
    """AJAX — Retorna as procurações salvas no banco para o escritório atual."""
    law_firm_id = get_current_law_firm_id()

    rows = (
        FapWebProcuracao.query
        .filter_by(law_firm_id=law_firm_id)
        .order_by(FapWebProcuracao.data_cadastro.desc())
        .all()
    )

    items = [
        {
            'id':                       r.id,
            'protocolo':                r.protocolo,
            'tipo_procuracao_codigo':   r.tipo_procuracao_codigo or '',
            'tipo_procuracao_descricao':r.tipo_procuracao_descricao or '',
            'situacao_codigo':          r.situacao_codigo or '',
            'situacao_descricao':       r.situacao_descricao or '',
            'data_inicio':              r.data_inicio.strftime('%d/%m/%Y') if r.data_inicio else '',
            'data_fim':                 r.data_fim.strftime('%d/%m/%Y') if r.data_fim else '',
            'cnpj_raiz_outorgante':     r.cnpj_raiz_outorgante or '',
            'nome_empresa_outorgante':  r.nome_empresa_outorgante or '',
            'cpf_outorgado':            r.cpf_outorgado or '',
            'cnpj_raiz_outorgado':      r.cnpj_raiz_outorgado or '',
            'data_cadastro':            r.data_cadastro.strftime('%d/%m/%Y %H:%M') if r.data_cadastro else '',
            'last_synced_at':           r.last_synced_at.strftime('%d/%m/%Y %H:%M') if r.last_synced_at else '',
        }
        for r in rows
    ]

    return jsonify({'ok': True, 'items': items, 'total': len(items)})


# ---------------------------------------------------------------------------
# Página de listagem de procurações
# ---------------------------------------------------------------------------

@fap_panel_bp.route('/procuracoes')
@require_law_firm
def procuracoes_page():
    """Página de visualização das procurações eletrônicas sincronizadas."""
    law_firm_id = get_current_law_firm_id()

    # Filtros
    f_situacao     = request.args.get('situacao', '').strip()
    f_tipo         = request.args.get('tipo', '').strip()
    f_outorgante   = request.args.get('outorgante', '').strip()
    f_protocolo    = request.args.get('protocolo', '').strip()
    f_vigencia_ini = request.args.get('vigencia_ini', '').strip()
    f_vigencia_fim = request.args.get('vigencia_fim', '').strip()
    f_vencendo_em  = request.args.get('vencendo_em', '').strip()   # '7','15','30','60','90'

    query = FapWebProcuracao.query.filter_by(law_firm_id=law_firm_id)

    if f_situacao:
        query = query.filter(FapWebProcuracao.situacao_codigo == f_situacao)
    if f_tipo:
        query = query.filter(FapWebProcuracao.tipo_procuracao_codigo == f_tipo)
    if f_outorgante:
        like = f'%{f_outorgante}%'
        query = query.filter(
            (FapWebProcuracao.nome_empresa_outorgante.ilike(like)) |
            (FapWebProcuracao.cnpj_raiz_outorgante.ilike(like))
        )
    if f_protocolo:
        query = query.filter(FapWebProcuracao.protocolo.ilike(f'%{f_protocolo}%'))
    if f_vigencia_ini:
        try:
            from datetime import date as _date
            di = _date.fromisoformat(f_vigencia_ini)
            query = query.filter(FapWebProcuracao.data_fim >= di)
        except Exception:
            pass
    if f_vigencia_fim:
        try:
            from datetime import date as _date
            df = _date.fromisoformat(f_vigencia_fim)
            query = query.filter(FapWebProcuracao.data_inicio <= df)
        except Exception:
            pass
    if f_vencendo_em:
        try:
            from datetime import date as _date, timedelta
            days = int(f_vencendo_em)
            hoje = _date.today()
            limite = hoje + timedelta(days=days)
            query = query.filter(
                FapWebProcuracao.data_fim >= hoje,
                FapWebProcuracao.data_fim <= limite,
            )
        except Exception:
            pass

    rows = query.order_by(FapWebProcuracao.data_cadastro.desc()).all()

    # Valores únicos para os filtros
    all_situacoes = (
        db.session.query(
            FapWebProcuracao.situacao_codigo,
            FapWebProcuracao.situacao_descricao,
        )
        .filter_by(law_firm_id=law_firm_id)
        .distinct()
        .order_by(FapWebProcuracao.situacao_descricao)
        .all()
    )
    all_tipos = (
        db.session.query(
            FapWebProcuracao.tipo_procuracao_codigo,
            FapWebProcuracao.tipo_procuracao_descricao,
        )
        .filter_by(law_firm_id=law_firm_id)
        .distinct()
        .order_by(FapWebProcuracao.tipo_procuracao_descricao)
        .all()
    )

    total = FapWebProcuracao.query.filter_by(law_firm_id=law_firm_id).count()

    from datetime import date as _date
    return render_template(
        'fap_panel/procuracoes.html',
        rows=rows,
        total=total,
        today=_date.today(),
        f_situacao=f_situacao,
        f_tipo=f_tipo,
        f_outorgante=f_outorgante,
        f_protocolo=f_protocolo,
        f_vigencia_ini=f_vigencia_ini,
        f_vigencia_fim=f_vigencia_fim,
        f_vencendo_em=f_vencendo_em,
        all_situacoes=all_situacoes,
        all_tipos=all_tipos,
    )
