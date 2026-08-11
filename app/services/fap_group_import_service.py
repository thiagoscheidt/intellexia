"""
Importação da planilha de grupos empresariais (CNPJ Raiz | Grupo | Razão Social).

O parsing é função pura — sem Flask, sem banco, sem rede —, então a peça que
muda se a planilha do escritório mudar de formato fica isolada e testável
sozinha. O confronto com o banco e a gravação ficam em ``preview`` e ``apply``.

Fluxo em duas etapas, sem tabela de job: o arquivo salvo em disco é a fonte, o
preview mostra o de-para e o apply reprocessa o mesmo arquivo. A planilha tem
centenas de linhas, não milhares — não compensa o peso de um job persistido.
"""

import unicodedata

from app.services.fap_group_service import (
    ORIGEM_MANUAL,
    ORIGEM_PLANILHA,
    cnpj_root,
    normalize_group_key,
)

# Quantas linhas varrer procurando o cabeçalho antes de desistir. A planilha do
# escritório costuma ter título/logo antes da tabela, então posição fixa falha.
MAX_LINHAS_CABECALHO = 15

COLUNA_CNPJ = 'cnpj'
COLUNA_GRUPO = 'grupo'
COLUNA_RAZAO = 'razao'

# Sinônimos aceitos por coluna, já normalizados (sem acento, maiúsculas).
CABECALHOS = {
    COLUNA_CNPJ: ('CNPJ RAIZ OUTORGANTE', 'CNPJ RAIZ', 'CNPJ'),
    COLUNA_GRUPO: ('GRUPO', 'GRUPO EMPRESARIAL'),
    COLUNA_RAZAO: ('RAZAO SOCIAL', 'RAZAO', 'EMPRESA'),
}


class SpreadsheetFormatError(Exception):
    """Planilha ilegível ou sem as colunas esperadas — mensagem pronta para a tela."""


def _normalize_header(valor):
    texto = str(valor or '').strip()
    sem_acento = ''.join(
        ch for ch in unicodedata.normalize('NFKD', texto)
        if not unicodedata.combining(ch)
    )
    return ' '.join(sem_acento.upper().split())


def _match_column(cabecalho_normalizado):
    """Nome de coluna → chave interna, ou None se não for uma das esperadas."""
    for chave, sinonimos in CABECALHOS.items():
        if cabecalho_normalizado in sinonimos:
            return chave
    return None


def _find_header(linhas):
    """Acha a linha de cabeçalho e o índice de cada coluna.

    Devolve (indice_da_linha, {chave: indice_da_coluna}). Exige pelo menos CNPJ
    e Grupo — a razão social é opcional (só enriquece a conferência).
    """
    for indice, linha in enumerate(linhas[:MAX_LINHAS_CABECALHO]):
        mapa = {}
        for coluna, celula in enumerate(linha):
            chave = _match_column(_normalize_header(celula))
            if chave and chave not in mapa:
                mapa[chave] = coluna
        if COLUNA_CNPJ in mapa and COLUNA_GRUPO in mapa:
            return indice, mapa
    return None, {}


def parse_rows(linhas):
    """Linhas cruas da planilha → (registros, erros). Função pura.

    ``registros``: [{'cnpj_raiz', 'grupo_nome', 'razao_social', 'linha'}]
    ``erros``:     [{'linha', 'mensagem'}]

    O mesmo CNPJ repetido com o mesmo grupo é aceito e contado uma vez; repetido
    com grupos diferentes vira erro apontando as duas linhas, porque adivinhar
    qual vale seria escolher em silêncio pelo usuário.
    """
    indice_cabecalho, colunas = _find_header(linhas)
    if indice_cabecalho is None:
        raise SpreadsheetFormatError(
            'Não encontrei as colunas esperadas na planilha. '
            'É preciso ter uma coluna de CNPJ raiz e uma coluna "Grupo".'
        )

    registros = {}
    erros = []

    for deslocamento, linha in enumerate(linhas[indice_cabecalho + 1:]):
        numero_linha = indice_cabecalho + deslocamento + 2  # 1-based, como no Excel

        def celula(chave):
            indice = colunas.get(chave)
            if indice is None or indice >= len(linha):
                return ''
            return str(linha[indice] or '').strip()

        bruto_cnpj = celula(COLUNA_CNPJ)
        grupo_nome = celula(COLUNA_GRUPO)
        razao = celula(COLUNA_RAZAO)

        if not bruto_cnpj and not grupo_nome and not razao:
            continue  # linha vazia: separador ou sobra da planilha

        raiz = cnpj_root(bruto_cnpj)
        if len(raiz) < 8:
            erros.append({
                'linha': numero_linha,
                'mensagem': f'CNPJ raiz inválido: "{bruto_cnpj}".',
            })
            continue

        if not normalize_group_key(grupo_nome):
            erros.append({
                'linha': numero_linha,
                'mensagem': f'Grupo em branco para o CNPJ {raiz}.',
            })
            continue

        anterior = registros.get(raiz)
        if anterior is None:
            registros[raiz] = {
                'cnpj_raiz': raiz,
                'grupo_nome': grupo_nome,
                'razao_social': razao,
                'linha': numero_linha,
            }
            continue

        if normalize_group_key(anterior['grupo_nome']) != normalize_group_key(grupo_nome):
            erros.append({
                'linha': numero_linha,
                'mensagem': (
                    f'CNPJ {raiz} aparece com grupos diferentes: '
                    f'"{anterior["grupo_nome"]}" na linha {anterior["linha"]} e '
                    f'"{grupo_nome}" nesta.'
                ),
            })

    return list(registros.values()), erros


def read_workbook_rows(caminho):
    """Lê todas as linhas da primeira aba do .xlsx. Levanta SpreadsheetFormatError."""
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(caminho, read_only=True, data_only=True)
    except SpreadsheetFormatError:
        raise
    except Exception:
        raise SpreadsheetFormatError(
            'Não consegui ler a planilha. Envie um arquivo .xlsx válido '
            '(salvo pelo Excel ou equivalente).'
        )

    try:
        aba = workbook.worksheets[0]
        return [list(linha) for linha in aba.iter_rows(values_only=True)]
    finally:
        workbook.close()


def parse_file(caminho):
    """Arquivo .xlsx → (registros, erros)."""
    return parse_rows(read_workbook_rows(caminho))


def build_preview(law_firm_id, caminho):
    """Confronta a planilha com o banco, sem gravar nada.

    Devolve as listas que a tela de conferência mostra. É aqui que a decisão
    "planilha vence, mas avisa antes" se materializa: quem já tinha grupo
    definido à mão aparece marcado em ``alterados``, para o usuário ver o que
    vai perder antes de confirmar.
    """
    from app.models import FapCompany
    from app.services.fap_group_service import groups_by_root

    registros, erros = parse_file(caminho)
    atuais = groups_by_root(law_firm_id)
    conhecidas = {
        c.cnpj for c in
        FapCompany.query.filter_by(law_firm_id=law_firm_id)
        .with_entities(FapCompany.cnpj).all()
    }

    novos, alterados, inalterados, sem_empresa = [], [], [], []

    for registro in registros:
        raiz = registro['cnpj_raiz']
        atual = atuais.get(raiz)
        item = dict(registro)
        # A planilha traz empresas de que o escritório já teve procuração e que
        # não vêm mais da API — aceitas, mas sinalizadas na conferência.
        item['sem_empresa'] = raiz not in conhecidas
        if item['sem_empresa']:
            sem_empresa.append(item)

        if atual is None:
            novos.append(item)
        elif atual['chave'] == normalize_group_key(registro['grupo_nome']):
            inalterados.append(item)
        else:
            item['grupo_atual'] = atual['nome']
            item['era_manual'] = atual['origem'] == ORIGEM_MANUAL
            alterados.append(item)

    return {
        'novos': novos,
        'alterados': alterados,
        'inalterados': inalterados,
        'sem_empresa': sem_empresa,
        'erros': erros,
        'totais': {
            'linhas_validas': len(registros),
            'novos': len(novos),
            'alterados': len(alterados),
            'inalterados': len(inalterados),
            'sem_empresa': len(sem_empresa),
            'erros': len(erros),
        },
    }


def apply_import(law_firm_id, caminho):
    """Aplica a planilha. Faz commit. Devolve os mesmos totais do preview.

    Reprocessa o arquivo em vez de confiar num payload carregado antes: o que
    for gravado é sempre o que está no arquivo que o usuário conferiu.
    """
    from app.models import db
    from app.services.fap_group_service import assign_group

    registros, erros = parse_file(caminho)

    criados = alterados = inalterados = 0
    for registro in registros:
        acao, _ = assign_group(
            law_firm_id,
            registro['cnpj_raiz'],
            registro['grupo_nome'],
            origem=ORIGEM_PLANILHA,
            razao_social_origem=registro.get('razao_social') or None,
        )
        if acao == 'criado':
            criados += 1
        elif acao == 'alterado':
            alterados += 1
        else:
            inalterados += 1

    db.session.commit()

    return {
        'linhas_validas': len(registros),
        'novos': criados,
        'alterados': alterados,
        'inalterados': inalterados,
        'erros': len(erros),
    }
