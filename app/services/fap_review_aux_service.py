"""
Serviço de extração dirigida dos documentos auxiliares do Revisor FAP — fonte única.

Orquestra: âncoras de benefícios (planilha, com fallback por regex na petição),
cache das extrações por SHA-256 do arquivo e montagem do payload da tela e do
bloco de contexto entregue ao agente revisor.

As extrações rodam SEQUENCIALMENTE (não em asyncio.gather): TokenUsageService e o
cache compartilham a sessão SQLAlchemy da thread, e commits intercalados entre
corrotinas corromperiam a transação.
"""

import hashlib
import json
import os
import re
from pathlib import Path

from flask import current_app

from app.agents.fap_review.auxiliary_extractor_agent import FapAuxiliaryDocumentExtractorAgent
from app.models import db, FapReviewAuxExtraction

try:
    from openpyxl import load_workbook
except ImportError:  # openpyxl é dependência do projeto; guarda defensiva
    load_workbook = None

_TEXT_EXTENSIONS = {'.pdf', '.docx', '.txt'}
_SPREADSHEET_EXTENSIONS = {'.xls', '.xlsx'}
_IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg'}
_MAX_DOCS = int(os.environ.get('FAP_REVIEW_AUX_MAX_DOCS', '10'))
_MAX_TEXT_CHARS = int(os.environ.get('FAP_REVIEW_AUX_MAX_TEXT_CHARS', '40000'))


def _normalize_number(value) -> str:
    return re.sub(r'\D+', '', str(value or ''))


def build_benefit_anchors(spreadsheet_rows: list[dict] | None,
                          petition_text: str | None) -> tuple[list[dict], str]:
    """Monta a lista de benefícios-âncora. Planilha tem prioridade; sem ela,
    tenta achar NBs (10 dígitos) no texto da petição."""
    anchors: list[dict] = []
    seen: dict[str, dict] = {}

    if spreadsheet_rows:
        for row in spreadsheet_rows:
            normalized = str(row.get('benefit_number_normalized') or '').strip()
            if not normalized:
                continue
            thesis = str(row.get('thesis') or '').strip()
            entry = seen.get(normalized)
            if entry:
                if thesis and thesis not in entry['theses']:
                    entry['theses'].append(thesis)
                continue
            entry = {
                'benefit_number': str(row.get('benefit_number') or normalized),
                'benefit_number_normalized': normalized,
                'theses': [thesis] if thesis else [],
            }
            seen[normalized] = entry
            anchors.append(entry)
        if anchors:
            return anchors, 'spreadsheet'

    if petition_text:
        for candidate in re.findall(r'\d[\d\.\s\-\/]{8,18}\d', petition_text):
            normalized = _normalize_number(candidate)
            if len(normalized) != 10 or normalized in seen:
                continue
            entry = {
                'benefit_number': ' '.join(candidate.split()),
                'benefit_number_normalized': normalized,
                'theses': [],
            }
            seen[normalized] = entry
            anchors.append(entry)
        if anchors:
            return anchors, 'petition_text'

    return [], 'none'


def anchors_fingerprint(anchors: list[dict]) -> str:
    """Hash estável (independe de ordem) da lista de âncoras — compõe a chave do cache."""
    parts = sorted(
        f"{a.get('benefit_number_normalized', '')}:{'|'.join(sorted(a.get('theses') or []))}"
        for a in anchors
    )
    return hashlib.sha256(';'.join(parts).encode('utf-8')).hexdigest()


def compute_file_sha256(file_path: str) -> str:
    digest = hashlib.sha256()
    with open(file_path, 'rb') as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b''):
            digest.update(chunk)
    return digest.hexdigest()


def get_cached_extraction(law_firm_id: int, file_sha256: str,
                          extractor_model: str, fingerprint: str) -> dict | None:
    row = FapReviewAuxExtraction.query.filter_by(
        law_firm_id=law_firm_id,
        file_sha256=file_sha256,
        extractor_model=extractor_model,
        anchors_fingerprint=fingerprint,
    ).first()
    if not row:
        return None
    try:
        parsed = json.loads(row.extraction_json)
        return parsed if isinstance(parsed, dict) else None
    except (TypeError, json.JSONDecodeError):
        return None


def store_extraction(law_firm_id: int, file_sha256: str, file_name: str,
                     extractor_model: str, fingerprint: str, extraction: dict) -> None:
    """Grava/atualiza o cache. Faz commit — chamar fora de transação aberta."""
    existing = FapReviewAuxExtraction.query.filter_by(
        law_firm_id=law_firm_id,
        file_sha256=file_sha256,
        extractor_model=extractor_model,
        anchors_fingerprint=fingerprint,
    ).first()
    if existing:
        existing.extraction_json = json.dumps(extraction, ensure_ascii=False)
        existing.file_name = file_name
    else:
        db.session.add(FapReviewAuxExtraction(
            law_firm_id=law_firm_id,
            file_sha256=file_sha256,
            file_name=file_name,
            extractor_model=extractor_model,
            anchors_fingerprint=fingerprint,
            extraction_json=json.dumps(extraction, ensure_ascii=False),
        ))
    db.session.commit()


def _spreadsheet_to_text(file_path: str) -> str:
    if not load_workbook:
        raise ImportError('openpyxl não está instalado')
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        lines: list[str] = []
        for worksheet in workbook.worksheets:
            lines.append(f'[ABA: {worksheet.title}]')
            for row in worksheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    lines.append(' | '.join(cells))
        return '\n'.join(lines)
    finally:
        workbook.close()


async def run_auxiliary_extractions(*, law_firm_id: int,
                                    documents: list[dict],
                                    spreadsheet_rows: list[dict] | None,
                                    petition_text: str | None,
                                    extract_text_fn,
                                    openai_api_key: str | None = None) -> tuple[dict, list[dict]]:
    """Extrai todos os documentos auxiliares e retorna (payload da tela, docs p/ o revisor)."""
    anchors, anchor_source = build_benefit_anchors(spreadsheet_rows, petition_text)
    fingerprint = anchors_fingerprint(anchors)

    valid_docs = [d for d in documents if isinstance(d, dict) and d.get('path')]
    skipped = [str(d.get('name') or Path(str(d.get('path'))).name) for d in valid_docs[_MAX_DOCS:]]
    if skipped:
        current_app.logger.warning(
            'FAP aux: %s documentos acima do limite FAP_REVIEW_AUX_MAX_DOCS=%s foram pulados: %s',
            len(skipped), _MAX_DOCS, ', '.join(skipped))
    valid_docs = valid_docs[:_MAX_DOCS]

    agent = FapAuxiliaryDocumentExtractorAgent(openai_api_key=openai_api_key)
    results: list[dict] = []

    for doc in valid_docs:
        path = str(doc['path'])
        name = str(doc.get('name') or Path(path).name)
        try:
            if not Path(path).exists():
                raise FileNotFoundError(f'Arquivo não encontrado: {path}')

            sha = compute_file_sha256(path)
            cached = get_cached_extraction(law_firm_id, sha, agent.model_name, fingerprint)
            if cached is not None:
                results.append({'file_name': name, 'from_cache': True, 'extraction': cached, 'error': None})
                continue

            extension = Path(path).suffix.lower()
            document_text = None
            if extension in _SPREADSHEET_EXTENSIONS:
                document_text = _spreadsheet_to_text(path)
            elif extension in _TEXT_EXTENSIONS:
                try:
                    document_text = extract_text_fn(path)
                except Exception as text_error:
                    current_app.logger.warning('FAP aux: extração de texto falhou (%s): %s', name, text_error)
                    document_text = None
            if document_text:
                document_text = document_text[:_MAX_TEXT_CHARS]
            if not document_text and extension not in (_IMAGE_EXTENSIONS | {'.pdf'}):
                raise ValueError('Não foi possível extrair texto do arquivo')

            extraction = await agent.extract(
                file_path=path,
                file_name=name,
                document_text=document_text,
                benefit_anchors=anchors,
                law_firm_id=law_firm_id,
            )
            extraction_dict = extraction.model_dump(mode='json')
            store_extraction(law_firm_id, sha, name, agent.model_name, fingerprint, extraction_dict)
            results.append({'file_name': name, 'from_cache': False, 'extraction': extraction_dict, 'error': None})
        except Exception as exc:
            current_app.logger.warning('FAP aux: extração falhou (%s): %s', name, exc)
            results.append({'file_name': name, 'from_cache': False, 'extraction': None, 'error': str(exc)})

    payload = build_review_payload(results, anchors, anchor_source, skipped)
    agent_documents = build_agent_documents(results)
    return payload, agent_documents


def build_review_payload(results: list[dict], anchors: list[dict],
                         anchor_source: str, skipped: list[str]) -> dict:
    """Payload persistido em result_json['auxiliary_documents_review'] e lido pela tela."""
    theses_map = {a['benefit_number_normalized']: a.get('theses') or [] for a in anchors}
    documents: list[dict] = []
    matched_count = 0

    for item in results:
        extraction = item.get('extraction') or {}
        related: list[dict] = []
        for benefit in extraction.get('related_benefits') or []:
            if not isinstance(benefit, dict):
                continue
            normalized = _normalize_number(benefit.get('benefit_number'))
            related.append({
                'benefit_number': str(benefit.get('benefit_number') or ''),
                'benefit_number_normalized': normalized,
                'theses': theses_map.get(normalized, []),
                'in_anchor_list': normalized in theses_map,
                'match_reason': str(benefit.get('match_reason') or ''),
                'facts': [
                    {
                        'label': str(fact.get('label') or ''),
                        'value': str(fact.get('value') or ''),
                        'source_excerpt': str(fact.get('source_excerpt') or '') or None,
                    }
                    for fact in (benefit.get('facts') or []) if isinstance(fact, dict)
                ],
            })

        if item.get('error'):
            status = 'error'
        elif related:
            status = 'matched'
            matched_count += 1
        else:
            status = 'unmatched'

        documents.append({
            'file_name': item.get('file_name') or '',
            'document_type': str(extraction.get('document_type') or 'OUTRO'),
            'status': status,
            'related_benefits': related,
            'general_summary': str(extraction.get('general_summary') or ''),
            'potential_divergences': [str(d) for d in extraction.get('potential_divergences') or []],
            'from_cache': bool(item.get('from_cache')),
            'error': item.get('error'),
        })

    return {
        'anchor_source': anchor_source,
        'total_documents': len(results),
        'matched_documents': matched_count,
        'documents': documents,
        'skipped_documents': list(skipped or []),
    }


def build_agent_documents(results: list[dict]) -> list[dict]:
    """Converte extrações em [{'name', 'content_summary'}] para o prompt do revisor."""
    agent_docs: list[dict] = []
    for item in results:
        name = item.get('file_name') or 'arquivo_sem_nome'
        extraction = item.get('extraction')
        if not extraction:
            agent_docs.append({'name': name})
            continue

        lines = [f"Tipo: {extraction.get('document_type') or 'OUTRO'}"]
        summary = str(extraction.get('general_summary') or '').strip()
        if summary:
            lines.append(f"Resumo: {summary}")
        for benefit in extraction.get('related_benefits') or []:
            if not isinstance(benefit, dict):
                continue
            lines.append(
                f"Benefício {benefit.get('benefit_number')} — vínculo: {benefit.get('match_reason') or 'não informado'}")
            for fact in benefit.get('facts') or []:
                if not isinstance(fact, dict):
                    continue
                excerpt = fact.get('source_excerpt')
                suffix = f' (trecho: "{excerpt}")' if excerpt else ''
                lines.append(f"  - {fact.get('label')}: {fact.get('value')}{suffix}")
        for divergence in extraction.get('potential_divergences') or []:
            lines.append(f"Possível divergência: {divergence}")

        agent_docs.append({'name': name, 'content_summary': '\n'.join(lines)})
    return agent_docs
