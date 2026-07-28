"""Importação de peças-modelo a partir da planilha de controle do escritório.

Módulo puro: NÃO importa app/Flask nem models no nível do módulo (os testes
importam este arquivo diretamente, sem contexto de aplicação).

Duas responsabilidades:
1. `parse_spreadsheet` — lê a planilha ("Petições protocoladas (pós inicial)")
   e devolve os candidatos a peça-modelo (agrupados por arquivo).
2. `download_drive_file` — baixa anonimamente o arquivo do Google Drive
   referenciado no hyperlink de cada linha.

Ver docs/superpowers/specs/2026-07-28-importacao-pecas-planilha-design.md
para o desenho completo (modelo de dados, fluxo, decisões).
"""
from __future__ import annotations

import hashlib
import os
import re
import unicodedata
from collections import OrderedDict
from datetime import date, datetime
from typing import Any, Optional
from urllib.parse import unquote

import openpyxl
import requests
from werkzeug.utils import secure_filename

# ── Erros ─────────────────────────────────────────────────────────────


class SpreadsheetFormatError(Exception):
    """Planilha sem cabeçalho reconhecível ou sem linhas do tipo pedido."""


class DriveAccessError(Exception):
    """Download do Drive falhou (arquivo restrito, rede, etc.)."""


# ── Normalização de texto ────────────────────────────────────────────


def _strip_accents(value: str) -> str:
    nfkd = unicodedata.normalize('NFKD', value)
    return ''.join(ch for ch in nfkd if not unicodedata.combining(ch))


def _normalize_text(value: Any) -> str:
    """Strip + upper + sem acento + espaços colapsados. Usado para comparar
    cabeçalhos e o tipo de documento (tolerante a variações de digitação)."""
    if value is None:
        return ''
    text = _strip_accents(str(value)).strip().upper()
    return re.sub(r'\s+', ' ', text)


def normalize_tribunal(raw: Any) -> Optional[str]:
    """`'TRF-4'`, `'TRF 4'`, `'trf4'`, `'TRF - 4'` -> `'TRF4'`.

    Qualquer coisa que não seja TRF1..TRF6 (ex.: `'TJSC'`, `''`, `None`)
    devolve `None`.
    """
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    normalized = _strip_accents(text).upper()
    compact = re.sub(r'[^A-Z0-9]', '', normalized)
    match = re.fullmatch(r'TRF([1-6])', compact)
    if not match:
        return None
    return f'TRF{match.group(1)}'


# ── Leitura da planilha ───────────────────────────────────────────────

# Cabeçalhos esperados (nome real na planilha -> chave interna). Apenas
# 'Nº' e 'AUTORA' são obrigatórios para reconhecer a linha de cabeçalho;
# os demais são usados se presentes.
_COLUMN_HEADERS = {
    'numero': 'Nº',
    'autora': 'AUTORA',
    'tribunal': 'TRIBUNAL',
    'orgao_julgador': 'ÓRGÃO JULGADOR',
    'teses': 'TESES / TÓPICOS',
    'qnt': 'QNT.',
    'document_type': 'MANIFESTAÇÕES APÓS A INICIAL',
    'protocolado': 'PROTOCOLADO',
}
_COLUMN_HEADERS_NORM = {key: _normalize_text(val) for key, val in _COLUMN_HEADERS.items()}
_HEADER_NUMERO_NORM = _COLUMN_HEADERS_NORM['numero']
_HEADER_AUTORA_NORM = _COLUMN_HEADERS_NORM['autora']

_MAX_HEADER_SCAN_ROWS = 15

_DRIVE_ID_RE = re.compile(r'/d/([A-Za-z0-9_-]+)')

_MAX_LEN = {
    'autora': 255,
    'orgao_julgador': 255,
    'tribunal_raw': 60,
    'document_type_label': 120,
}

_DATE_FORMATS = (
    '%d/%m/%Y %H:%M:%S',
    '%d/%m/%Y',
    '%Y-%m-%d %H:%M:%S',
    '%Y-%m-%d',
    '%d-%m-%Y',
)


def _find_header(wb) -> Optional[tuple]:
    """Varre as primeiras `_MAX_HEADER_SCAN_ROWS` linhas de cada aba
    procurando uma linha que contenha, ao mesmo tempo, os cabeçalhos
    normalizados de 'Nº' e 'AUTORA'. Devolve (worksheet, header_row, col_map)
    ou None se nenhuma aba tiver esse cabeçalho."""
    for ws in wb.worksheets:
        max_row_scan = min(_MAX_HEADER_SCAN_ROWS, ws.max_row or 0)
        for row_idx in range(1, max_row_scan + 1):
            cols_by_norm: dict[str, int] = {}
            for col_idx in range(1, (ws.max_column or 0) + 1):
                norm = _normalize_text(ws.cell(row=row_idx, column=col_idx).value)
                if norm:
                    cols_by_norm.setdefault(norm, col_idx)
            if _HEADER_NUMERO_NORM in cols_by_norm and _HEADER_AUTORA_NORM in cols_by_norm:
                col_map = {
                    key: cols_by_norm[norm]
                    for key, norm in _COLUMN_HEADERS_NORM.items()
                    if norm in cols_by_norm
                }
                return ws, row_idx, col_map
    return None


def _clean_str(value: Any, max_len: Optional[int] = None) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text[:max_len] if max_len else text


def _normalize_numero(raw: Any) -> Optional[str]:
    """`325.0` (float do Excel) -> `'325'`. Preserva texto não numérico."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return str(raw)
    if isinstance(raw, int):
        return str(raw)
    if isinstance(raw, float):
        return str(int(raw)) if raw.is_integer() else str(raw)
    text = str(raw).strip()
    if not text:
        return None
    try:
        as_float = float(text.replace(',', '.'))
    except ValueError:
        return text
    return str(int(as_float)) if as_float.is_integer() else text


def _parse_protocolado(raw: Any) -> Optional[datetime]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, date):
        return datetime(raw.year, raw.month, raw.day)
    text = str(raw).strip()
    if not text:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _first_non_empty(values):
    for value in values:
        if value not in (None, ''):
            return value
    return None


def _build_candidate(
    *,
    numero,
    autora,
    tribunal_raw,
    trf_region,
    orgao_julgador,
    document_type_label,
    drive_file_id,
    teses,
    protocolado_at,
    row_number,
):
    drive_url = f'https://drive.google.com/file/d/{drive_file_id}/view' if drive_file_id else None
    return {
        'numero': numero,
        'autora': _clean_str(autora, _MAX_LEN['autora']),
        'tribunal_raw': _clean_str(tribunal_raw, _MAX_LEN['tribunal_raw']),
        'trf_region': trf_region,
        'orgao_julgador': _clean_str(orgao_julgador, _MAX_LEN['orgao_julgador']),
        'document_type_label': _clean_str(document_type_label, _MAX_LEN['document_type_label']),
        'drive_file_id': drive_file_id,
        'drive_url': drive_url,
        'teses': teses,
        'protocolado_at': protocolado_at,
        'row_number': row_number,
    }


def parse_spreadsheet(
    file_path: str,
    *,
    document_type_filter: str = 'IMPUGNAÇÃO À CONTESTAÇÃO',
) -> list[dict]:
    """Lê a planilha de controle e devolve os candidatos a peça-modelo.

    Localiza o cabeçalho (não assume posição fixa), filtra as linhas cujo
    tipo de manifestação contém `document_type_filter` (comparação sem
    acento/caixa) e agrupa por `(numero, autora, tipo_normalizado)` — a
    unidade de importação é o arquivo, então um grupo com mais de um
    `drive_file_id` distinto vira mais de um candidato.

    Levanta `SpreadsheetFormatError` (mensagem em português, pronta para a
    tela) se nenhuma aba tiver cabeçalho reconhecível ou se nenhuma linha
    do tipo pedido for encontrada.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    header = _find_header(wb)
    if header is None:
        raise SpreadsheetFormatError(
            "Não foi possível localizar o cabeçalho da planilha (colunas "
            "'Nº' e 'AUTORA') nas primeiras 15 linhas de nenhuma aba. "
            "Verifique se o arquivo enviado é a planilha de controle correta."
        )
    ws, header_row, col_map = header

    if 'document_type' not in col_map:
        raise SpreadsheetFormatError(
            "Não foi possível localizar a coluna 'MANIFESTAÇÕES APÓS A "
            "INICIAL' na planilha — sem ela não é possível filtrar o tipo "
            "de peça a importar."
        )

    filter_norm = _normalize_text(document_type_filter)

    # groups[(numero, autora, tipo_norm)] = {'rows': [row_data...], 'row_number': int}
    groups: "OrderedDict[tuple, dict]" = OrderedDict()

    doc_type_col = col_map['document_type']
    numero_col = col_map.get('numero')
    autora_col = col_map.get('autora')
    tribunal_col = col_map.get('tribunal')
    orgao_col = col_map.get('orgao_julgador')
    teses_col = col_map.get('teses')
    protocolado_col = col_map.get('protocolado')

    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        doc_cell = ws.cell(row=row_idx, column=doc_type_col)
        doc_type_raw = doc_cell.value
        doc_type_norm = _normalize_text(doc_type_raw)
        if not doc_type_norm or filter_norm not in doc_type_norm:
            continue

        numero_raw = ws.cell(row=row_idx, column=numero_col).value if numero_col else None
        autora_raw = ws.cell(row=row_idx, column=autora_col).value if autora_col else None
        numero = _normalize_numero(numero_raw)
        autora = _clean_str(autora_raw, _MAX_LEN['autora'])

        tese_raw = ws.cell(row=row_idx, column=teses_col).value if teses_col else None
        tese = _clean_str(tese_raw)

        drive_id = None
        if doc_cell.hyperlink and doc_cell.hyperlink.target:
            match = _DRIVE_ID_RE.search(doc_cell.hyperlink.target)
            if match:
                drive_id = match.group(1)

        row_entry = {
            'row': row_idx,
            'tese': tese,
            'drive_id': drive_id,
            'tribunal_raw': ws.cell(row=row_idx, column=tribunal_col).value if tribunal_col else None,
            'orgao_julgador': ws.cell(row=row_idx, column=orgao_col).value if orgao_col else None,
            'protocolado': ws.cell(row=row_idx, column=protocolado_col).value if protocolado_col else None,
            'doc_type_label': doc_type_raw,
        }

        key = (numero, autora, doc_type_norm)
        if key not in groups:
            groups[key] = {'rows': [], 'row_number': row_idx}
        groups[key]['rows'].append(row_entry)

    if not groups:
        raise SpreadsheetFormatError(
            f"Nenhuma linha do tipo '{document_type_filter}' foi encontrada "
            "na planilha."
        )

    candidates: list[dict] = []

    for (numero, autora, _tipo_norm), info in groups.items():
        rows = info['rows']
        row_number = info['row_number']

        tribunal_raw_val = _first_non_empty(r['tribunal_raw'] for r in rows)
        orgao_val = _first_non_empty(r['orgao_julgador'] for r in rows)
        protocolado_val = _first_non_empty(r['protocolado'] for r in rows)
        doc_type_label_val = _first_non_empty(r['doc_type_label'] for r in rows)

        trf_region = normalize_tribunal(tribunal_raw_val)
        protocolado_at = _parse_protocolado(protocolado_val)

        distinct_ids: list[str] = []
        for r in rows:
            if r['drive_id'] and r['drive_id'] not in distinct_ids:
                distinct_ids.append(r['drive_id'])

        if len(distinct_ids) <= 1:
            drive_id = distinct_ids[0] if distinct_ids else None
            teses = []
            for r in rows:
                if r['tese'] and r['tese'] not in teses:
                    teses.append(r['tese'])
            candidates.append(_build_candidate(
                numero=numero,
                autora=autora,
                tribunal_raw=tribunal_raw_val,
                trf_region=trf_region,
                orgao_julgador=orgao_val,
                document_type_label=doc_type_label_val,
                drive_file_id=drive_id,
                teses=teses,
                protocolado_at=protocolado_at,
                row_number=row_number,
            ))
        else:
            # Mais de um arquivo no mesmo grupo -> um candidato por id.
            # Teses de linhas sem link vão para o primeiro candidato.
            for i, drive_id in enumerate(distinct_ids):
                teses = []
                for r in rows:
                    if not r['tese']:
                        continue
                    belongs = r['drive_id'] == drive_id or (i == 0 and r['drive_id'] is None)
                    if belongs and r['tese'] not in teses:
                        teses.append(r['tese'])
                candidates.append(_build_candidate(
                    numero=numero,
                    autora=autora,
                    tribunal_raw=tribunal_raw_val,
                    trf_region=trf_region,
                    orgao_julgador=orgao_val,
                    document_type_label=doc_type_label_val,
                    drive_file_id=drive_id,
                    teses=teses,
                    protocolado_at=protocolado_at,
                    row_number=row_number,
                ))

    return candidates


# ── Download do Drive ────────────────────────────────────────────────

_DRIVE_DOWNLOAD_URL = 'https://drive.usercontent.google.com/download'
_RESTRICTED_MESSAGE = (
    'Arquivo sem compartilhamento público no Drive — libere o acesso '
    '(qualquer pessoa com o link) e reprocesse este item.'
)


def _looks_like_html(content_type: Optional[str], first_chunk_bytes: bytes) -> bool:
    """Heurística p/ detectar página HTML (login/aviso) em vez do arquivo
    binário esperado. Testável sem rede."""
    if content_type and 'text/html' in content_type.lower():
        return True
    if first_chunk_bytes:
        head = first_chunk_bytes.lstrip()[:20].lower()
        if head.startswith(b'<!doctype') or head.startswith(b'<html'):
            return True
    return False


def _filename_from_content_disposition(header_value: Optional[str]) -> Optional[str]:
    if not header_value:
        return None
    # RFC 5987: filename*=UTF-8''nome%20codificado
    match = re.search(r"filename\*\s*=\s*([^;]+)", header_value, re.IGNORECASE)
    if match:
        raw = match.group(1).strip().strip('"')
        if "''" in raw:
            encoding, _, encoded_name = raw.partition("''")
            try:
                return unquote(encoded_name, encoding=encoding or 'utf-8')
            except (LookupError, ValueError):
                return unquote(encoded_name)
        return unquote(raw)
    match = re.search(r'filename\s*=\s*"([^"]+)"', header_value, re.IGNORECASE)
    if match:
        return match.group(1)
    match = re.search(r'filename\s*=\s*([^;]+)', header_value, re.IGNORECASE)
    if match:
        return match.group(1).strip().strip('"')
    return None


def _safe_filename(candidate: Optional[str], fallback: str) -> str:
    if candidate:
        safe = secure_filename(candidate)
        if safe:
            return safe
    safe_fallback = secure_filename(fallback)
    return safe_fallback or fallback


def _open_drive_stream(url: str, timeout: int):
    try:
        response = requests.get(url, stream=True, timeout=timeout, allow_redirects=True)
        first_chunk = next(response.iter_content(chunk_size=65536), b'')
    except requests.RequestException as exc:
        raise DriveAccessError(
            f'Não foi possível conectar ao Google Drive: {exc}'
        ) from exc
    return response, first_chunk


def _save_stream(response, first_chunk: bytes, drive_file_id: str, dest_dir: str) -> tuple:
    filename = _filename_from_content_disposition(response.headers.get('Content-Disposition'))
    original_filename = filename or f'{drive_file_id}.pdf'
    safe_name = _safe_filename(filename, f'{drive_file_id}.pdf')

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    dest_path = os.path.join(dest_dir, f'{timestamp}_{safe_name}')

    sha256 = hashlib.sha256()
    try:
        with open(dest_path, 'wb') as fh:
            if first_chunk:
                fh.write(first_chunk)
                sha256.update(first_chunk)
            for chunk in response.iter_content(chunk_size=65536):
                if chunk:
                    fh.write(chunk)
                    sha256.update(chunk)
    except requests.RequestException as exc:
        raise DriveAccessError(
            f'Falha ao baixar o arquivo do Drive (id={drive_file_id}): {exc}'
        ) from exc

    return dest_path, original_filename, sha256.hexdigest()


def download_drive_file(drive_file_id: str, dest_dir: str, *, timeout: int = 90) -> tuple:
    """Baixa anonimamente do Drive. Retorna (caminho_salvo, nome_original, sha256).

    Arquivo restrito devolve HTML de login -> DriveAccessError com mensagem
    pronta para a tela.
    """
    os.makedirs(dest_dir, exist_ok=True)
    url = f'{_DRIVE_DOWNLOAD_URL}?id={drive_file_id}&export=download'

    response, first_chunk = _open_drive_stream(url, timeout)
    try:
        content_type = response.headers.get('Content-Type', '')
        if _looks_like_html(content_type, first_chunk):
            # Página pequena (login ou interstitial de vírus) — lê o corpo
            # inteiro para decidir se vale tentar de novo com confirm=t.
            try:
                body = first_chunk + response.content
            except requests.RequestException as exc:
                raise DriveAccessError(
                    f'Não foi possível baixar o arquivo do Drive (id={drive_file_id}): {exc}'
                ) from exc
            response.close()

            if b'confirm=' in body:
                retry_response, retry_first_chunk = _open_drive_stream(f'{url}&confirm=t', timeout)
                retry_content_type = retry_response.headers.get('Content-Type', '')
                if _looks_like_html(retry_content_type, retry_first_chunk):
                    retry_response.close()
                    raise DriveAccessError(_RESTRICTED_MESSAGE)
                response, first_chunk = retry_response, retry_first_chunk
            else:
                raise DriveAccessError(_RESTRICTED_MESSAGE)

        return _save_stream(response, first_chunk, drive_file_id, dest_dir)
    finally:
        response.close()
