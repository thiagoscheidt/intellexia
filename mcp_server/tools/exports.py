"""
Tools: Exportação para Excel (XLSX)
===================================
Gera planilhas no mesmo padrão visual dos exports das telas (openpyxl) e
devolve um link de download assinado (itsdangerous, validade de 1 hora),
servido pelo próprio servidor MCP em /export/<token>.

Os arquivos ficam em uploads/mcp_exports/{law_firm_id}/ e são limpos
automaticamente após 24 horas.
"""
from __future__ import annotations

import os
import re
import secrets
import time
from datetime import datetime

from itsdangerous import URLSafeTimedSerializer

EXPORT_TTL_SECONDS = 3600          # validade do link (1 hora)
EXPORT_FILE_TTL_SECONDS = 24 * 3600  # arquivos antigos são removidos após 24h
MAX_EXPORT_ROWS = 50000

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_EXPORT_DIR = os.path.join(_PROJECT_ROOT, "uploads", "mcp_exports")


def _serializer():
    from main import app

    return URLSafeTimedSerializer(app.config["SECRET_KEY"], salt="mcp-export")


def _cleanup_old_files() -> None:
    cutoff = time.time() - EXPORT_FILE_TTL_SECONDS
    for root, _dirs, files in os.walk(_EXPORT_DIR):
        for name in files:
            path = os.path.join(root, name)
            try:
                if os.path.getmtime(path) < cutoff:
                    os.remove(path)
            except OSError:
                pass


def _save_workbook(workbook, law_firm_id: int, prefix: str) -> str:
    """Salva o workbook e retorna o caminho relativo (law_firm_id/arquivo)."""
    os.makedirs(os.path.join(_EXPORT_DIR, str(law_firm_id)), exist_ok=True)
    _cleanup_old_files()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_{stamp}_{secrets.token_hex(4)}.xlsx"
    rel_path = f"{law_firm_id}/{filename}"
    workbook.save(os.path.join(_EXPORT_DIR, rel_path))
    return rel_path


def build_download_response(token: str):
    """Valida o token assinado e devolve o arquivo (usado pela rota /export)."""
    from itsdangerous import BadSignature, SignatureExpired
    from starlette.responses import FileResponse, JSONResponse

    try:
        payload = _serializer().loads(token, max_age=EXPORT_TTL_SECONDS)
    except SignatureExpired:
        return JSONResponse({"erro": "Link de download expirado. Gere a exportação novamente."}, status_code=410)
    except BadSignature:
        return JSONResponse({"erro": "Link de download inválido."}, status_code=403)

    rel_path = payload.get("f", "")
    # Sanidade: sem path traversal e restrito ao diretório de exports
    if not re.fullmatch(r"\d+/[A-Za-z0-9_.\-]+\.xlsx", rel_path):
        return JSONResponse({"erro": "Link de download inválido."}, status_code=403)
    full_path = os.path.join(_EXPORT_DIR, rel_path)
    if not os.path.isfile(full_path):
        return JSONResponse({"erro": "Arquivo não encontrado (expirado). Gere a exportação novamente."}, status_code=404)

    return FileResponse(
        full_path,
        filename=os.path.basename(full_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _download_result(rel_path: str, mcp_public_url: str, total_linhas: int) -> dict:
    token = _serializer().dumps({"f": rel_path})
    return {
        "arquivo": os.path.basename(rel_path),
        "total_linhas": total_linhas,
        "url_download": f"{mcp_public_url.rstrip('/')}/export/{token}",
        "validade_minutos": EXPORT_TTL_SECONDS // 60,
        "instrucao": "Abra o link no navegador para baixar a planilha.",
    }


# ── Escrita da planilha (padrão visual dos exports das telas) ─────────────────


def _write_sheet(sheet_title: str, main_title: str, filter_pairs: list[tuple[str, str]],
                 headers: list[str], rows: list[list], col_widths: list[int]):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1565C0")
    title_align = Alignment(horizontal="center", vertical="center")
    filter_label_font = Font(name="Calibri", size=10, bold=True, color="0D47A1")
    filter_val_font = Font(name="Calibri", size=10, color="212121")
    filter_fill = PatternFill("solid", fgColor="E3F2FD")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1976D2")
    cell_font = Font(name="Calibri", size=10)
    zebra_fill = PatternFill("solid", fgColor="F5F5F5")
    thin = Side(style="thin", color="B0BEC5")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    num_cols = len(headers)

    # Título
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = sheet.cell(row=1, column=1, value=main_title)
    cell.font, cell.fill, cell.alignment = title_font, title_fill, title_align
    sheet.row_dimensions[1].height = 28

    # Bloco de filtros
    pairs_per_row = 3
    filter_start_row = 2
    for idx, (lbl, val) in enumerate(filter_pairs):
        frow = filter_start_row + (idx // pairs_per_row)
        fcol = 1 + (idx % pairs_per_row) * 2
        lc = sheet.cell(row=frow, column=fcol, value=f"{lbl}:")
        lc.font, lc.fill, lc.border = filter_label_font, filter_fill, cell_border
        lc.alignment = Alignment(horizontal="right", vertical="center")
        vc = sheet.cell(row=frow, column=fcol + 1, value=val)
        vc.font, vc.fill, vc.border = filter_val_font, filter_fill, cell_border
        vc.alignment = Alignment(horizontal="left", vertical="center")
    filter_end_row = filter_start_row + (len(filter_pairs) - 1) // pairs_per_row
    for frow in range(filter_start_row, filter_end_row + 1):
        sheet.row_dimensions[frow].height = 18
        for fcol in range(1, num_cols + 1):
            c = sheet.cell(row=frow, column=fcol)
            if c.value is None:
                c.fill, c.border = filter_fill, cell_border

    # Cabeçalhos
    header_row = filter_end_row + 2
    for col, header in enumerate(headers, start=1):
        c = sheet.cell(row=header_row, column=col, value=header)
        c.font, c.fill, c.border = header_font, header_fill, cell_border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 26

    # Dados (zebra)
    for r_idx, row in enumerate(rows):
        excel_row = header_row + 1 + r_idx
        for col, value in enumerate(row, start=1):
            c = sheet.cell(row=excel_row, column=col, value=value)
            c.font, c.border = cell_font, cell_border
            if r_idx % 2 == 1:
                c.fill = zebra_fill

    for i, width in enumerate(col_widths, start=1):
        sheet.column_dimensions[sheet.cell(row=header_row, column=i).column_letter].width = width

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    return workbook


# ── Exports ───────────────────────────────────────────────────────────────────


def export_benefits_excel_handler(law_firm_id: int, mcp_public_url: str, **filters) -> dict:
    """Exporta benefícios FAP filtrados para XLSX; retorna link de download."""
    from mcp_server.tools.fap import list_fap_benefits_handler

    data = list_fap_benefits_handler(law_firm_id, limit=MAX_EXPORT_ROWS, **filters)
    itens = data["itens"]
    if not itens:
        return {"erro": "Nenhum benefício encontrado com esses filtros.", "total_linhas": 0}

    headers = [
        "Nº Benefício", "Tipo", "Pedido", "Status", "Segurado", "NIT", "CPF",
        "CNPJ Empregador", "Empregador", "Início", "Fim", "Data Acidente",
        "Nº CAT", "Vigências FAP", "Tópicos de Contestação",
        "1ª Instância", "2ª Instância", "Mensalidade Inicial", "Total Pago",
    ]
    rows = [
        [
            b["numero_beneficio"], b["tipo_beneficio"], b["tipo_pedido"], b["status"],
            b["segurado_nome"], b["segurado_nit"], b["segurado_cpf"],
            b["empregador_cnpj"], b["empregador_nome"],
            b["data_inicio_beneficio"], b["data_fim_beneficio"], b["data_acidente"],
            b["numero_cat"], b["vigencias_fap"], ", ".join(b["topicos_contestacao"]),
            b["status_primeira_instancia"], b["status_segunda_instancia"],
            b["mensalidade_inicial"], b["total_pago"],
        ]
        for b in itens
    ]
    filter_labels = {
        "cnpj": "CNPJ", "status": "Status", "request_type": "Pedido",
        "benefit_type": "Tipo", "fap_contestation_topic": "Tópico",
        "segurado": "Segurado", "nit": "NIT", "cpf": "CPF",
        "numero_beneficio": "Nº Benefício", "ano_vigencia": "Vigência",
    }
    filter_pairs = [(filter_labels.get(k, k), str(v)) for k, v in filters.items() if v]
    filter_pairs.append(("Total linhas", str(len(rows))))

    workbook = _write_sheet(
        "Benefícios FAP",
        f'Benefícios FAP — Exportado em {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        filter_pairs, headers, rows,
        [16, 8, 12, 12, 30, 14, 14, 18, 30, 12, 12, 12, 14, 12, 40, 14, 14, 14, 14],
    )
    rel_path = _save_workbook(workbook, law_firm_id, "beneficios_fap")
    result = _download_result(rel_path, mcp_public_url, len(rows))
    if data["total_encontrado"] > MAX_EXPORT_ROWS:
        result["aviso"] = f"Resultado truncado em {MAX_EXPORT_ROWS} linhas (total: {data['total_encontrado']})."
    return result


def export_contestacoes_excel_handler(law_firm_id: int, mcp_public_url: str, **filters) -> dict:
    """Exporta contestações FAP filtradas para XLSX; retorna link de download."""
    from mcp_server.tools.fap import list_fap_contestacoes_handler

    data = list_fap_contestacoes_handler(law_firm_id, limit=MAX_EXPORT_ROWS, **filters)
    itens = data["itens"]
    if not itens:
        return {"erro": "Nenhuma contestação encontrada com esses filtros.", "total_linhas": 0}

    headers = [
        "ID Contestação", "Empresa", "CNPJ", "CNPJ Raiz", "Vigência",
        "Instância", "Situação", "Protocolo", "Data Transmissão", "Data D.O.U.",
        "PDF Baixado", "Última Sincronização",
    ]
    rows = [
        [
            c["contestacao_id"], c["empresa_nome"], c["cnpj"], c["cnpj_raiz"],
            c["ano_vigencia"], c["instancia_descricao"], c["situacao_descricao"],
            c["protocolo"], c["data_transmissao"], c["data_dou"],
            "Sim" if c["pdf_baixado"] else "Não", c["ultima_sincronizacao"],
        ]
        for c in itens
    ]
    filter_labels = {
        "cnpj": "CNPJ", "cnpj_raiz": "CNPJ Raiz", "ano_vigencia": "Vigência",
        "situacao_codigo": "Situação", "instancia_codigo": "Instância",
    }
    filter_pairs = [(filter_labels.get(k, k), str(v)) for k, v in filters.items() if v]
    filter_pairs.append(("Total linhas", str(len(rows))))

    workbook = _write_sheet(
        "Contestações FAP",
        f'Contestações FAP — Exportado em {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        filter_pairs, headers, rows,
        [14, 32, 18, 12, 10, 26, 26, 14, 18, 12, 12, 18],
    )
    rel_path = _save_workbook(workbook, law_firm_id, "contestacoes_fap")
    result = _download_result(rel_path, mcp_public_url, len(rows))
    if data["total_encontrado"] > MAX_EXPORT_ROWS:
        result["aviso"] = f"Resultado truncado em {MAX_EXPORT_ROWS} linhas (total: {data['total_encontrado']})."
    return result
