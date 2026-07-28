"""Teste do parser de planilha e helpers de download do serviço de
importação de peças-modelo (impugnacao_import_service).

Puro Python — sem rede, sem banco, sem app Flask. Constrói uma planilha
.xlsx sintética com openpyxl (reproduzindo as manhas do arquivo real:
cabeçalho fora da linha 1, hyperlink só na primeira linha de um grupo,
grupo sem hyperlink algum) e a remove ao final.

Rodar: uv run python scripts/tests/test_impugnacao_import_service.py
"""
import sys
import tempfile
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from app.services.impugnacao_import_service import (
    _classify_restricted_html,
    _filename_from_content_disposition,
    _looks_like_html,
    _RATE_LIMIT_MESSAGE,
    _RESTRICTED_MESSAGE,
    download_drive_file,
    normalize_tribunal,
    parse_spreadsheet,
    SpreadsheetFormatError,
)
from app.services.impugnacao_reference_ingestion import apply_extracted_metadata

FAILS = []


def check(label, got, expected):
    if got != expected:
        FAILS.append(f"{label}: esperado {expected!r}, obtido {got!r}")
        print(f"  ✗ {label}: esperado {expected!r}, obtido {got!r}")
    else:
        print(f"  ✓ {label}")


def check_true(label, cond, detail=""):
    if not cond:
        FAILS.append(f"{label} {detail}")
        print(f"  ✗ {label} {detail}")
    else:
        print(f"  ✓ {label}")


# ── monta a planilha sintética ──────────────────────────────────────────

HEADERS = [
    'Nº', 'AUTORA', 'TRIBUNAL', 'ÓRGÃO JULGADOR', 'TESES / TÓPICOS', 'QNT.',
    'MANIFESTAÇÕES APÓS A INICIAL', 'PROTOCOLADO',
]

DRIVE_ID_A = '1AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
DRIVE_ID_B = '1BBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBB'


def _build_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Planilha de teses'

    # título acima do cabeçalho, como no arquivo real (cabeçalho na linha 3)
    ws.cell(row=1, column=1, value='REVISÃO DE INSUMOS - CUMPRIMENTO DE PRAZOS')
    ws.cell(row=2, column=1, value=None)
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=3, column=col_idx, value=header)

    row = 4

    # (a) peça com 3 linhas de tese, hyperlink só na primeira linha
    for i, tese in enumerate(['TESE 1 DO GRUPO A', 'TESE 2 DO GRUPO A', 'TESE 3 DO GRUPO A']):
        ws.cell(row=row, column=1, value=325.0)  # numero como float, sempre igual no grupo
        ws.cell(row=row, column=2, value='EMPRESA A LTDA')
        ws.cell(row=row, column=3, value='TRF-4')
        ws.cell(row=row, column=4, value='3ª Vara Federal de Florianópolis')
        ws.cell(row=row, column=5, value=tese)
        ws.cell(row=row, column=6, value=1.0)
        tipo_cell = ws.cell(row=row, column=7, value='IMPUGNAÇÃO À CONTESTAÇÃO')
        if i == 0:
            tipo_cell.hyperlink = f'https://drive.google.com/file/d/{DRIVE_ID_A}/view?usp=drive_link'
        ws.cell(row=row, column=8, value='13/07/2026')
        row += 1

    # (b) peça "IMPUGNAÇÃO À CONTESTAÇÃO DO INSS" com 1 linha e link
    ws.cell(row=row, column=1, value=200)
    ws.cell(row=row, column=2, value='EMPRESA B S.A.')
    ws.cell(row=row, column=3, value='TRF2')
    ws.cell(row=row, column=4, value='8ª Vara Federal do Rio de Janeiro/RJ')
    ws.cell(row=row, column=5, value='TESE ÚNICA DO GRUPO B')
    ws.cell(row=row, column=6, value=1.0)
    tipo_cell = ws.cell(row=row, column=7, value='IMPUGNAÇÃO À CONTESTAÇÃO DO INSS')
    tipo_cell.hyperlink = f'https://drive.google.com/file/d/{DRIVE_ID_B}/view?usp=drive_link'
    ws.cell(row=row, column=8, value='14/07/2026')
    row += 1

    # (c) peça de tipo diferente ("APELAÇÃO") — deve ser ignorada
    ws.cell(row=row, column=1, value=300)
    ws.cell(row=row, column=2, value='EMPRESA C LTDA')
    ws.cell(row=row, column=3, value='TJSC')
    ws.cell(row=row, column=4, value='1ª Vara Cível')
    ws.cell(row=row, column=5, value='TESE DO GRUPO C')
    ws.cell(row=row, column=6, value=1.0)
    ws.cell(row=row, column=7, value='APELAÇÃO')
    ws.cell(row=row, column=8, value='15/07/2026')
    row += 1

    # (d) peça de impugnação sem hyperlink nenhum
    ws.cell(row=row, column=1, value=400)
    ws.cell(row=row, column=2, value='EMPRESA D LTDA')
    ws.cell(row=row, column=3, value='TRF3')
    ws.cell(row=row, column=4, value='2ª Vara Federal de Campinas/SP')
    ws.cell(row=row, column=5, value='TESE DO GRUPO D')
    ws.cell(row=row, column=6, value=1.0)
    ws.cell(row=row, column=7, value='IMPUGNAÇÃO À CONTESTAÇÃO')
    ws.cell(row=row, column=8, value='16/07/2026')
    row += 1

    return wb


tmp_dir = Path(tempfile.mkdtemp(prefix='impugnacao_import_test_'))
xlsx_path = tmp_dir / 'planilha_sintetica.xlsx'
wb = _build_workbook()
wb.save(xlsx_path)

try:
    # ── parse_spreadsheet ────────────────────────────────────────────
    print("parse_spreadsheet:")
    candidates = parse_spreadsheet(str(xlsx_path))

    check_true("3 candidatos (a, b, d)", len(candidates) == 3, f"(veio {len(candidates)})")

    autoras = {c['autora'] for c in candidates}
    check_true("nenhum candidato da EMPRESA C (APELAÇÃO)", 'EMPRESA C LTDA' not in autoras)

    cand_a = next((c for c in candidates if c['autora'] == 'EMPRESA A LTDA'), None)
    cand_b = next((c for c in candidates if c['autora'] == 'EMPRESA B S.A.'), None)
    cand_d = next((c for c in candidates if c['autora'] == 'EMPRESA D LTDA'), None)

    check_true("candidato A encontrado", cand_a is not None)
    check_true("candidato B encontrado", cand_b is not None)
    check_true("candidato D encontrado", cand_d is not None)

    if cand_a:
        check("candidato A tem 3 teses", len(cand_a['teses']), 3)
        check("candidato A teses na ordem", cand_a['teses'],
              ['TESE 1 DO GRUPO A', 'TESE 2 DO GRUPO A', 'TESE 3 DO GRUPO A'])
        check("candidato A drive_file_id", cand_a['drive_file_id'], DRIVE_ID_A)
        check("candidato A numero (325.0 -> '325')", cand_a['numero'], '325')
        check("candidato A trf_region (TRF-4 -> TRF4)", cand_a['trf_region'], 'TRF4')
        check("candidato A drive_url", cand_a['drive_url'],
              f'https://drive.google.com/file/d/{DRIVE_ID_A}/view')

    if cand_b:
        check("candidato B tem 1 tese", len(cand_b['teses']), 1)
        check("candidato B drive_file_id", cand_b['drive_file_id'], DRIVE_ID_B)
        check("candidato B document_type_label", cand_b['document_type_label'],
              'IMPUGNAÇÃO À CONTESTAÇÃO DO INSS')

    if cand_d:
        check("candidato D sem link -> drive_file_id None", cand_d['drive_file_id'], None)
        check("candidato D sem link -> drive_url None", cand_d['drive_url'], None)

    # planilha sem cabeçalho reconhecível -> SpreadsheetFormatError
    print("\nSpreadsheetFormatError:")
    wb_sem_cabecalho = openpyxl.Workbook()
    ws_sem = wb_sem_cabecalho.active
    ws_sem.cell(row=1, column=1, value='nada aqui')
    xlsx_sem_cabecalho = tmp_dir / 'sem_cabecalho.xlsx'
    wb_sem_cabecalho.save(xlsx_sem_cabecalho)
    try:
        parse_spreadsheet(str(xlsx_sem_cabecalho))
        check_true("planilha sem cabeçalho levanta SpreadsheetFormatError", False)
    except Exception as exc:
        check_true(
            "planilha sem cabeçalho levanta SpreadsheetFormatError",
            isinstance(exc, SpreadsheetFormatError),
            f"(veio {type(exc).__name__})",
        )

    # arquivo .xlsx corrompido/renomeado (na verdade um .txt) -> SpreadsheetFormatError,
    # não o BadZipFile cru do openpyxl (I6)
    xlsx_nao_zip = tmp_dir / 'nao_e_zip.xlsx'
    xlsx_nao_zip.write_text('isto não é um arquivo .xlsx, é texto puro', encoding='utf-8')
    try:
        parse_spreadsheet(str(xlsx_nao_zip))
        check_true(".xlsx não-zip levanta SpreadsheetFormatError", False)
    except Exception as exc:
        check_true(
            ".xlsx não-zip levanta SpreadsheetFormatError (não BadZipFile)",
            isinstance(exc, SpreadsheetFormatError),
            f"(veio {type(exc).__name__})",
        )

    # ── normalize_tribunal ───────────────────────────────────────────
    print("\nnormalize_tribunal:")
    check("'TRF-4' -> 'TRF4'", normalize_tribunal('TRF-4'), 'TRF4')
    check("'TRF 4' -> 'TRF4'", normalize_tribunal('TRF 4'), 'TRF4')
    check("'trf4' -> 'TRF4'", normalize_tribunal('trf4'), 'TRF4')
    check("'TRF - 4' -> 'TRF4'", normalize_tribunal('TRF - 4'), 'TRF4')
    check("'TJSC' -> None", normalize_tribunal('TJSC'), None)
    check("'' -> None", normalize_tribunal(''), None)
    check("None -> None", normalize_tribunal(None), None)
    check("'TRF7' -> None (fora de 1..6)", normalize_tribunal('TRF7'), None)

    # ── _looks_like_html ──────────────────────────────────────────────
    print("\n_looks_like_html:")
    check(
        "content-type text/html -> True",
        _looks_like_html('text/html; charset=utf-8', b'<!DOCTYPE html>'),
        True,
    )
    check(
        "content-type application/pdf + %PDF -> False",
        _looks_like_html('application/pdf', b'%PDF-1.7'),
        False,
    )
    check(
        "sem content-type, corpo começa com <html -> True",
        _looks_like_html('', b'<html><head>'),
        True,
    )

    # ── _filename_from_content_disposition ─────────────────────────────
    print("\n_filename_from_content_disposition:")
    check(
        'filename="a.pdf" -> a.pdf',
        _filename_from_content_disposition('attachment; filename="a.pdf"'),
        'a.pdf',
    )
    check(
        "RFC 5987 filename*=UTF-8''Impugna%C3%A7%C3%A3o.pdf -> Impugnação.pdf",
        _filename_from_content_disposition(
            "attachment; filename*=UTF-8''Impugna%C3%A7%C3%A3o.pdf"
        ),
        'Impugnação.pdf',
    )

    # ── _classify_restricted_html (I7) ──────────────────────────────────
    print("\n_classify_restricted_html:")
    check(
        "HTML de login do Google -> mensagem de compartilhamento",
        _classify_restricted_html(
            'Faça login em accounts.google.com para continuar'.encode('utf-8')
        ),
        _RESTRICTED_MESSAGE,
    )
    check(
        "HTML com 'Solicitar acesso' -> mensagem de compartilhamento",
        _classify_restricted_html(
            'Voce precisa <a>Solicitar acesso</a> a este arquivo'.encode('utf-8')
        ),
        _RESTRICTED_MESSAGE,
    )
    check(
        "HTML genérico sem sinal de login -> mensagem de limite de requisições",
        _classify_restricted_html(
            b'<html><body>Ocorreu um erro inesperado. Tente novamente.</body></html>'
        ),
        _RATE_LIMIT_MESSAGE,
    )

    # ── apply_extracted_metadata (C1) ───────────────────────────────────
    print("\napply_extracted_metadata:")
    meta = SimpleNamespace(
        title='Título sugerido pela IA',
        case_name='Empresa Sugerida pela IA Ltda',
        process_number='1234567-12.2024.4.02.1234',
        orgao_julgador='Vara sugerida pela IA',
        judge_name='Fulano de Tal',
        trf_region='TRF2',
        generation_mode='A',
        quality_score=4.5,
    )

    # preserve_curated_fields=True: os 4 campos curados (title, case_name,
    # trf_region, orgao_julgador) já preenchidos permanecem intocados;
    # generation_mode/quality_score/process_number/judge_name vêm sempre da
    # IA, mesmo com quality_score prévio 3.00 (Decimal, truthy) — é
    # exatamente o cenário do bug C1(b).
    ref_curated = SimpleNamespace(
        title='Título Curado da Planilha',
        case_name='Empresa Curada Ltda',
        trf_region='TRF4',
        orgao_julgador='3ª Vara Federal de Florianópolis',
        process_number=None,
        judge_name=None,
        generation_mode=None,
        quality_score=Decimal('3.00'),
    )
    apply_extracted_metadata(ref_curated, meta, is_new=True, preserve_curated_fields=True)
    check("preserve=True: title permanece curado", ref_curated.title, 'Título Curado da Planilha')
    check("preserve=True: case_name permanece curado", ref_curated.case_name, 'Empresa Curada Ltda')
    check("preserve=True: trf_region permanece curado", ref_curated.trf_region, 'TRF4')
    check(
        "preserve=True: orgao_julgador permanece curado",
        ref_curated.orgao_julgador,
        '3ª Vara Federal de Florianópolis',
    )
    check("preserve=True: generation_mode vem da IA", ref_curated.generation_mode, 'A')
    check(
        "preserve=True: quality_score vem da IA mesmo com 3.00 prévio",
        ref_curated.quality_score,
        4.5,
    )
    check("preserve=True: process_number vem da IA", ref_curated.process_number, meta.process_number)
    check("preserve=True: judge_name vem da IA", ref_curated.judge_name, meta.judge_name)

    # preserve_curated_fields=False (upload manual): comportamento antigo —
    # os 5 campos curados são sempre sobrescritos pela IA.
    ref_manual = SimpleNamespace(
        title='Título Curado da Planilha',
        case_name='Empresa Curada Ltda',
        trf_region='TRF4',
        orgao_julgador='3ª Vara Federal de Florianópolis',
        process_number=None,
        judge_name=None,
        generation_mode=None,
        quality_score=Decimal('3.00'),
    )
    apply_extracted_metadata(ref_manual, meta, is_new=True, preserve_curated_fields=False)
    check("preserve=False: title sobrescrito pela IA", ref_manual.title, meta.title)
    check("preserve=False: case_name sobrescrito pela IA", ref_manual.case_name, meta.case_name)
    check("preserve=False: trf_region sobrescrito pela IA", ref_manual.trf_region, meta.trf_region)
    check(
        "preserve=False: orgao_julgador sobrescrito pela IA",
        ref_manual.orgao_julgador,
        meta.orgao_julgador,
    )
    check("preserve=False: generation_mode sobrescrito pela IA", ref_manual.generation_mode, meta.generation_mode)
    check("preserve=False: quality_score sobrescrito pela IA", ref_manual.quality_score, meta.quality_score)
    check("preserve=False: process_number sobrescrito pela IA", ref_manual.process_number, meta.process_number)
    check("preserve=False: judge_name sobrescrito pela IA", ref_manual.judge_name, meta.judge_name)

    # ── download_drive_file: sem rede, só a assinatura/erro de conexão ──
    # Não fazemos chamada de rede real; apenas garantimos que download_drive_file
    # está exportado com a assinatura esperada (verificação leve de import).
    check_true("download_drive_file é chamável", callable(download_drive_file))

finally:
    import shutil
    shutil.rmtree(tmp_dir, ignore_errors=True)


print()
if FAILS:
    print(f"FALHOU: {len(FAILS)} verificação(ões)")
    sys.exit(1)
print("OK: todos os checks passaram")
