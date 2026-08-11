"""
Classificação das contestações em Em Andamento / Transmitidas / Publicadas.

A regra antiga procurava palavras em `código + descrição` e classificava as
TRÊS situações reais como "transmitidas": a descrição de EM_ANDAMENTO é
"Iniciada, não TRANSMITida", e a busca por substring achava o "transmit" do
"não transmitida", excluindo justamente o registro que devia entrar. A coluna
"Em Andamento" ficava sempre vazia.

Executar:
    uv run python tests/test_fap_contestacoes_colunas.py
"""

import os
import sys

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from main import app

DB_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '_tmp_colunas.db')
if os.path.exists(DB_FILE):
    os.remove(DB_FILE)
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_FILE}'

from app.models import db  # noqa: E402

app.extensions.pop('sqlalchemy', None)
try:
    db._app_engines.pop(app, None)
except Exception:
    pass
db.init_app(app)

from app.blueprints import fap_panel as fp  # noqa: E402

FALHAS = []


def check(rotulo, condicao, extra=''):
    print(f"  [{'OK ' if condicao else 'FALHA'}] {rotulo}{(' — ' + str(extra)) if extra else ''}")
    if not condicao:
        FALHAS.append(rotulo)


# As três situações que a API do FAP Web devolve, com a descrição exata.
SITUACOES_REAIS = [
    ('EM_ANDAMENTO', 'Iniciada, não transmitida', fp.COL_ANDAMENTO),
    ('LIBERADA_PARA_ANALISE', 'Transmitida', fp.COL_TRANSMITIDA),
    ('PUBLICADA', 'Resultado divulgado no D.O.U', fp.COL_PUBLICADA),
]

PRIMEIRA = 'ADMINISTRATIVO_PRIMEIRA_INSTANCIA'
SEGUNDA = 'ADMINISTRATIVO_SEGUNDA_INSTANCIA'


def main():
    print('\n1. cada situação real vai para a sua coluna')
    for codigo, _desc, esperado in SITUACOES_REAIS:
        obtido = fp._coluna_da_situacao(codigo)
        check(f'{codigo} → {esperado}', obtido == esperado, obtido)

    print('\n2. a regressão que motivou a mudança')
    check('EM_ANDAMENTO não cai mais em "transmitidas" por causa do '
          '"não TRANSMITida" da descrição',
          fp._coluna_da_situacao('EM_ANDAMENTO') == fp.COL_ANDAMENTO)
    check('PUBLICADA não é mais confundida com transmitida',
          fp._coluna_da_situacao('PUBLICADA') == fp.COL_PUBLICADA)

    print('\n3. situação desconhecida fica visível, não some')
    for codigo in ('SITUACAO_NOVA_DA_DATAPREV', '', None, '   '):
        check(f'{codigo!r} → outras',
              fp._coluna_da_situacao(codigo) == fp.COL_OUTRAS,
              fp._coluna_da_situacao(codigo))

    print('\n4. instância separa as colunas 1 e 2')
    check('1ª instância → sufixo 1',
          fp._chave_celula('PUBLICADA', PRIMEIRA) == 'c_pub1',
          fp._chave_celula('PUBLICADA', PRIMEIRA))
    check('2ª instância → sufixo 2',
          fp._chave_celula('PUBLICADA', SEGUNDA) == 'c_pub2',
          fp._chave_celula('PUBLICADA', SEGUNDA))
    check('instância nula cai em 1ª (contestação)',
          fp._chave_celula('EM_ANDAMENTO', None) == 'c_and1')
    check('caixa não importa no código da situação',
          fp._chave_celula('publicada', PRIMEIRA) == 'c_pub1')

    print('\n5. os baldes cobrem todas as combinações, sem sobra')
    baldes = fp._baldes_vazios()
    check('8 baldes (4 colunas × 2 instâncias)', len(baldes) == 8, len(baldes))
    todas_chaves = {fp._chave_celula(c, i)
                    for c in ['EM_ANDAMENTO', 'LIBERADA_PARA_ANALISE', 'PUBLICADA', 'XPTO']
                    for i in [PRIMEIRA, SEGUNDA]}
    check('toda situação possível tem balde', todas_chaves <= set(baldes),
          todas_chaves - set(baldes))

    # ── a tela: contagens e colunas com dados de verdade ─────────────────
    print('\n6. a tela distribui os registros nas três colunas')
    from datetime import datetime
    with app.app_context():
        from app.models import LawFirm, User, FapWebContestacao
        assert DB_FILE in str(db.engine.url), 'ABORTADO: fora do sandbox'
        db.create_all()
        db.session.add(LawFirm(id=1, name='Escritório', cnpj='00000000000191'))
        db.session.add(User(id=1, law_firm_id=1, name='Admin', email='a@b.c',
                            password_hash='x', role='admin'))
        # 1ª instância: 2 em andamento, 3 transmitidas, 5 publicadas
        # 2ª instância: 1 em andamento, 1 transmitida, 2 publicadas
        plano = ([('EM_ANDAMENTO', PRIMEIRA)] * 2 + [('LIBERADA_PARA_ANALISE', PRIMEIRA)] * 3
                 + [('PUBLICADA', PRIMEIRA)] * 5 + [('EM_ANDAMENTO', SEGUNDA)]
                 + [('LIBERADA_PARA_ANALISE', SEGUNDA)] + [('PUBLICADA', SEGUNDA)] * 2)
        for i, (sit, inst) in enumerate(plano, start=1):
            db.session.add(FapWebContestacao(
                id=i, law_firm_id=1, contestacao_id=1000 + i,
                cnpj='11111111000100', cnpj_raiz='11111111', ano_vigencia=2026,
                situacao_codigo=sit, instancia_codigo=inst, protocolo=f'P{i}',
                last_synced_at=datetime(2026, 1, 1), created_at=datetime(2026, 1, 1),
            ))
        db.session.commit()

    cliente = app.test_client()
    with cliente.session_transaction() as sessao:
        sessao['user_id'] = 1
        sessao['law_firm_id'] = 1
        sessao['user_role'] = 'admin'

    resposta = cliente.get('/fap-panel/contestacoes?ano_vigencia=2026')
    html = resposta.get_data(as_text=True)
    check('tela responde', resposta.status_code == 200, resposta.status_code)
    for rotulo in ('Em Andamento', 'Transmitidas', 'Publicadas'):
        check(f'coluna "{rotulo}" no cabeçalho', rotulo in html)
    check('cards mostram 1ª: 2 em andamento', '1ª: 2' in html)
    check('cards mostram 1ª: 5 publicadas', '1ª: 5' in html)
    check('"Outras" não aparece quando está vazia', 'Inst. Outras' not in html)

    print('\n7. situação desconhecida faz a coluna "Outras" aparecer')
    with app.app_context():
        from app.models import FapWebContestacao
        db.session.add(FapWebContestacao(
            id=999, law_firm_id=1, contestacao_id=9999,
            cnpj='11111111000100', cnpj_raiz='11111111', ano_vigencia=2026,
            situacao_codigo='CODIGO_QUE_NAO_EXISTIA', instancia_codigo=PRIMEIRA,
            protocolo='P999', last_synced_at=datetime(2026, 1, 1),
            created_at=datetime(2026, 1, 1),
        ))
        db.session.commit()

    html = cliente.get('/fap-panel/contestacoes?ano_vigencia=2026').get_data(as_text=True)
    check('coluna "Outras" aparece', 'Inst. Outras' in html)
    check('e o registro é contado nela', '1ª: 1' in html)

    print('\n8. o export agrupado não diverge da tela')
    resposta = cliente.get('/fap-panel/contestacoes/export-excel-agrupado?ano_vigencia=2026')
    check('export responde', resposta.status_code == 200, resposta.status_code)
    check('devolve um .xlsx',
          resposta.headers.get('Content-Type', '').find('spreadsheet') >= 0,
          resposta.headers.get('Content-Type'))
    if resposta.status_code == 200:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resposta.get_data()))
        textos = {str(c.value) for linha in wb.active.iter_rows() for c in linha if c.value}
        for esperado in ('Publicadas\n(Qtd)', 'Transmitidas\n(Qtd)', 'Em Andamento\n(Qtd)'):
            check(f'planilha tem a coluna {esperado.splitlines()[0]!r}',
                  esperado in textos)
        check('planilha traz o bloco "Outras" quando existe',
              any(str(t).startswith('Outras') for t in textos))

    print('\n' + '=' * 62)
    print('RESULTADO:', 'TUDO OK' if not FALHAS else f'{len(FALHAS)} FALHA(S): {FALHAS}')
    if os.path.exists(DB_FILE):
        os.remove(DB_FILE)
    return 1 if FALHAS else 0


if __name__ == '__main__':
    sys.exit(main())
