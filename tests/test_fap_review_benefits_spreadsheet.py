"""
Teste da leitura da planilha de benefícios do Revisor FAP com múltiplas abas.

Regressão: planilhas com uma aba por vigência (2021, 2022, ...) eram lidas só
na aba ativa — as demais eram ignoradas silenciosamente.

Uso: uv run python tests/test_fap_review_benefits_spreadsheet.py
"""

import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.environ.setdefault('OPENAI_API_KEY', 'test-key')

from openpyxl import Workbook  # noqa: E402

from app.blueprints.fap_review import _parse_benefits_spreadsheet  # noqa: E402

PASSED = 0
FAILED = 0


def check(label: str, condition: bool, detail: str = "") -> None:
    global PASSED, FAILED
    if condition:
        PASSED += 1
        print(f"  ✓ {label}")
    else:
        FAILED += 1
        print(f"  ✗ {label} {detail}")


def build_multi_tab_spreadsheet(path: Path) -> None:
    workbook = Workbook()

    sheet_2021 = workbook.active
    sheet_2021.title = '2021'
    sheet_2021.append(['Número do Benefício', 'TESES'])
    sheet_2021.append(['6407132987', 'BENEFÍCIO CANCELADO + DIB=DCB'])
    sheet_2021.append(['1111111111', ''])  # sem tese: fora da conferência

    sheet_2022 = workbook.create_sheet('2022')
    sheet_2022.append(['Número do Benefício', 'TESES'])
    sheet_2022.append(['6222222222', 'ACIDENTE DE TRAJETO'])

    # Aba sem as colunas esperadas: deve ser ignorada sem erro
    notes = workbook.create_sheet('Anotações')
    notes.append(['Observações gerais'])
    notes.append(['texto livre'])

    sheet_2024 = workbook.create_sheet('2024')
    sheet_2024.append(['Número do Benefício', 'TESES'])
    sheet_2024.append(['6444444444', 'B31 INDEVIDO'])

    # Simula o comportamento comum: última aba editada fica ativa ao salvar
    workbook.active = workbook.sheetnames.index('2024')
    workbook.save(str(path))


def run():
    print("[1] Planilha com múltiplas abas")
    with tempfile.TemporaryDirectory() as tmp:
        xlsx_path = Path(tmp) / 'beneficios.xlsx'
        build_multi_tab_spreadsheet(xlsx_path)
        rows = _parse_benefits_spreadsheet(str(xlsx_path))

    numbers = sorted(row['benefit_number_normalized'] for row in rows)
    check("lê linhas de todas as abas com tese", len(rows) == 3, f"(obteve {len(rows)}: {numbers})")
    check("benefício da aba 2021 presente", '6407132987' in numbers)
    check("benefício da aba 2022 presente", '6222222222' in numbers)
    check("benefício da aba 2024 presente", '6444444444' in numbers)
    check("linha sem tese fora", '1111111111' not in numbers)
    sheets = {row.get('sheet_name') for row in rows}
    check("aba de origem registrada", sheets == {'2021', '2022', '2024'}, f"(obteve {sheets})")

    print("[2] Nenhuma aba com as colunas esperadas → erro claro")
    with tempfile.TemporaryDirectory() as tmp:
        xlsx_path = Path(tmp) / 'invalida.xlsx'
        workbook = Workbook()
        workbook.active.append(['Coluna qualquer'])
        workbook.save(str(xlsx_path))
        try:
            _parse_benefits_spreadsheet(str(xlsx_path))
            check("levantou ValueError", False, "(não levantou)")
        except ValueError:
            check("levantou ValueError", True)


if __name__ == '__main__':
    run()
    print(f"\nResultado: {PASSED} ok, {FAILED} falhas")
    sys.exit(1 if FAILED else 0)
